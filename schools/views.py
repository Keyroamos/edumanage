from django.shortcuts import render, redirect, get_object_or_404
from .view_assessment_api import api_assessment_batch_save, api_academic_analytics
from .view_student_academic_api import api_student_academic_records
from .views_academic_mgmt import api_promote_students, api_update_term
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.http import HttpResponse, JsonResponse, FileResponse
from django.utils import timezone
from django.core.exceptions import PermissionDenied, ValidationError
from django.conf import settings
import logging
from django.db.models import (
    Count,
    Q, Sum, Count, Avg, F, ExpressionWrapper, FloatField, Value, 
    IntegerField, CharField, Case, When, DecimalField, DurationField,
    Subquery, OuterRef, Max, Prefetch
)
from django.db.models.functions import (
    Coalesce, TruncMonth, Cast
)
from collections import Counter, defaultdict
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.units import inch, mm
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from io import BytesIO
import os
import base64
import qrcode
import traceback
from django.core.paginator import Paginator
from django.views.generic import ListView, CreateView, DetailView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin, PermissionRequiredMixin
from django.urls import reverse, reverse_lazy
from django.views.decorators.csrf import csrf_exempt, ensure_csrf_cookie
from django.views.decorators.cache import never_cache
from django.views.decorators.http import require_http_methods
import json  # Add this import
from django.db import models
from collections import Counter
from .paystack_utils import initiate_paystack_stk, verify_paystack_payment # Switched to Paystack

from .models import (
    Student, Teacher, Grade, Assessment, AssessmentResult, 
    Payment, Employee, Salary, Leave, Term, Subject, Schedule,
    Attendance, TERM_CHOICES, SMSMessage, CommunicationTemplate, Department,
    EmployeeAttendance, Deduction, Vehicle, Route, StudentTransportAssignment, TransportFee,
    MealPricing, StudentMealPayment, MealConsumption, SyncQueue, SyncStatus,
    Branch, SalaryAdvance
)
from django.db import transaction
from django.contrib.auth.models import User, Group
from .forms import (
    StudentForm, TeacherForm, AssessmentForm, AssessmentResultFormSet, PaymentForm,
    EmployeeForm, SalaryForm, LeaveForm, AssessmentResultForm,
    ScheduleForm, SMSMessageForm, AdminStaffForm, VehicleForm, RouteForm,
    StudentTransportAssignmentForm, TransportFeeForm, MealPricingForm, StudentMealPaymentForm
)
from .decorators import admin_required, accountant_required, can_manage_students
from .utils import generate_payment_receipt, generate_receipt_qr
from django.views.decorators.http import require_http_methods
from datetime import datetime, date
from decimal import Decimal, InvalidOperation

# CSRF Token Endpoint
@ensure_csrf_cookie
@require_http_methods(["GET"])
def get_csrf_token(request):
    """
    Endpoint to set CSRF cookie for the frontend
    """
    return JsonResponse({'detail': 'CSRF cookie set'})

# Authentication Views
@ensure_csrf_cookie
@csrf_exempt
@require_http_methods(["POST"])
def api_login(request):
    """API endpoint for user login with support for portal passwords"""
    try:
        print(f"DEBUG: api_login view entered. Method: {request.method}")
        print(f"DEBUG: Params: {request.GET}")
        data = json.loads(request.body)
        login_field = data.get('username') or data.get('email')
        password = data.get('password')
        portal_slug = data.get('portal_slug')
        is_admin_portal = data.get('is_admin_portal', False)
        
        print(f"DEBUG: Login attempt for: {login_field}, portal_slug: {portal_slug}, is_admin_portal: {is_admin_portal}")
        
        if not login_field or not password:
            print("DEBUG: Missing credentials")
            return JsonResponse({'error': 'Username/Email and password are required'}, status=400)

        user = None
        role = 'admin'
        redirect_url = '/dashboard'
        school = None
        
        # 1. GET SCHOOL CONTEXT (If portal_slug provided)
        if portal_slug:
            from config.models import SchoolConfig
            school = SchoolConfig.objects.filter(portal_slug=portal_slug).first()
            if not school:
                if is_admin_portal:
                    print(f"DEBUG: Stale/Invalid portal slug ({portal_slug}) ignored for Admin Portal login.")
                    portal_slug = None # Ignore stale/invalid slug for admin portal
                else:
                    print(f"DEBUG: Invalid portal slug provided: {portal_slug}")
                    return JsonResponse({'error': f'Invalid portal access link ({portal_slug})'}, status=403)

        # 2. ATTEMPT STANDARD AUTHENTICATION (Individual User Password)
        user = authenticate(request, username=login_field, password=password)
        if user is None:
            # Try by email
            from django.contrib.auth.models import User
            user_obj = User.objects.filter(email=login_field).first()
            if user_obj:
                print(f"DEBUG: Found user by email: {user_obj.username}")
                user = authenticate(request, username=user_obj.username, password=password)
            else:
                print("DEBUG: No user found with this email/username")

        if user:
            print(f"DEBUG: Standard Auth Success for: {user.username}")
        else:
            print("DEBUG: Standard Auth Failed")

        # 3. ATTEMPT SPECIAL PORTAL AUTHENTICATION (Shared Portal Password)
        # Only if standard auth failed and we have a portal_slug and NOT logging into Admin Portal
        if user is None and portal_slug and not is_admin_portal:
            # --- STUDENT PORTAL AUTH ---
            student = Student.objects.filter(school=school, admission_number=login_field).first()
            if not student:
                 student = Student.objects.filter(school=school, user__username=login_field).first()
            
            if student:
                if password == school.student_portal_password:
                    if student.user:
                        user = student.user
                        role = 'student'
                        redirect_url = f'/student/{student.id}'
                        # Manually set backend for direct login
                        user.backend = 'django.contrib.auth.backends.ModelBackend'
                    else:
                        return JsonResponse({'error': 'Student account not fully configured (no user profile)'}, status=400)
                else:
                    return JsonResponse({'error': 'Invalid student portal password'}, status=401)
            
            # --- EMPLOYEE PORTAL AUTH ---
            if not user:
                employee = Employee.objects.filter(school=school, email=login_field).select_related('teacher', 'nonteachingstaff').first()
                if employee:
                    required_password = None
                    emp_role = None
                    
                    if employee.position == 'TEACHER':
                        required_password = school.teacher_portal_password
                        emp_role = 'teacher'
                    elif employee.position == 'ACCOUNTANT':
                        required_password = school.accountant_portal_password
                        emp_role = 'accountant'
                    elif employee.position == 'FOOD_MANAGER':
                        required_password = school.food_portal_password
                        emp_role = 'food manager'
                    elif employee.position == 'TRANSPORT_MANAGER':
                        required_password = school.transport_portal_password
                        emp_role = 'transport manager'
                    elif employee.position == 'DRIVER':
                        required_password = school.driver_portal_password
                        emp_role = 'driver'
                    else:
                        required_password = school.accountant_portal_password
                        emp_role = 'staff'

                    if password == required_password:
                        # Use the property to get user
                        user = employee.user
                        if user:
                            user.backend = 'django.contrib.auth.backends.ModelBackend'
                            role = 'teacher' if emp_role == 'teacher' else 'staff'
                            if emp_role == 'driver': role = 'driver'
                            
                            # Set dashboard based on role/emp_role
                            if emp_role == 'teacher':
                                redirect_url = f'/teacher/{employee.id}'
                            elif emp_role == 'driver':
                                redirect_url = '/driver-portal'
                            elif emp_role == 'food manager':
                                redirect_url = '/food-portal/dashboard'
                            elif emp_role == 'transport manager':
                                redirect_url = '/transport-portal/dashboard'
                            else:
                                redirect_url = '/finance-portal/dashboard'
                        else:
                            return JsonResponse({'error': 'Employee account not fully configured (no user profile)'}, status=400)
                    else:
                        return JsonResponse({'error': f'Invalid {emp_role} portal password. If you changed your password, please use your new password.'}, status=401)
        
        # 3. FINALIZE LOGIN
        if user is not None and user.is_active:
            print(f"DEBUG: Authenticated user: {user.username}, Superuser: {user.is_superuser}")
            # Enforce School Authorization if logging into a specific portal
            if portal_slug:
                # 0. Check if school is active
                if not school.is_active and not user.is_superuser:
                    return JsonResponse({
                        'error': 'This school account has been deactivated. Please contact support.'
                    }, status=403)

                # Verify user belongs to this school
                is_authorized = False
                
                # 1. School Admin or Superuser
                user_config = getattr(user, 'school_config', None)
                print(f"DEBUG: User has school_config: {user_config.school_name if user_config else 'No'}")
                print(f"DEBUG: Request school: {school.school_name}")

                if (user_config and user_config == school) or user.is_superuser:
                    is_authorized = True
                    print("DEBUG: Authorized as Admin/Superuser")
                
                # 1b. Lenient Admin Login: If it's the admin portal and they are an admin of ANY school,
                # let them in and we'll just switch the school context to theirs.
                elif is_admin_portal and user_config:
                    is_authorized = True
                    school = user_config # Switch context to the user's actual school
                    print(f"DEBUG: Admin login switch context to school: {school.school_name}")
                
                # 2. Specialized Roles (Only if NOT logging into Admin Portal)
                elif not is_admin_portal:
                    print(f"DEBUG: Checking specialized roles for portal: {portal_slug}")
                    # 2. Student
                    if hasattr(user, 'student_profile') and user.student_profile.school == school:
                        is_authorized = True
                        print("DEBUG: Authorized as Student")
                    
                    # 3. Teacher (Teacher inherits from Employee, usually has .teacher related_name)
                    elif hasattr(user, 'teacher') and user.teacher.school == school:
                        is_authorized = True
                    
                    # 4. Non-Teaching Staff
                    elif hasattr(user, 'staff_profile') and user.staff_profile.school == school:
                        is_authorized = True
                    
                    # 5. Driver
                    elif hasattr(user, 'transport_driver_profile') and user.transport_driver_profile.school == school:
                        is_authorized = True
                
                print(f"DEBUG: Final Portal Authorization Result: {is_authorized}")
                if not is_authorized:
                    if is_admin_portal:
                        error_msg = f"Access denied: Your account '{user.username}' is not authorized as an administrator for this school profile."
                    else:
                        error_msg = f"Access denied: Your account '{user.username}' is not authorized for the {portal_slug or 'requested'} portal."
                    return JsonResponse({'error': error_msg}, status=403)

            login(request, user)
            
            # If standard auth was used, determine role based on profile properties
            if not portal_slug:
                # Determine role and redirect
                role = 'admin'
                redirect_url = '/dashboard'
                
                if not user.is_superuser and not hasattr(user, 'school_config'):
                    if hasattr(user, 'student_profile'):
                        role = 'student'
                        redirect_url = f'/student/{user.student_profile.id}' 
                    elif hasattr(user, 'teacher'):
                        role = 'teacher'
                        redirect_url = f'/teacher/{user.teacher.id}'
                    elif hasattr(user, 'transport_driver_profile'):
                        role = 'driver'
                        redirect_url = '/driver-portal'
                    elif user.is_staff:
                        role = 'staff'
                        redirect_url = '/finance-portal/dashboard'
                    else:
                        print(f"DEBUG: User '{user.username}' has no recognized role and is not superuser")
                        return JsonResponse({'error': f'Access denied: User {user.username} has no recognized role (Admin/Staff/Student/Teacher)'}, status=403)
            
            # 4. Check if resolving school is active
            try:
                from config.models import SchoolConfig
                user_school = SchoolConfig.get_config(user=user)
                if user_school and not user_school.is_active and not user.is_superuser:
                    logout(request)
                    return JsonResponse({'error': 'Your school account is currently deactivated. Please contact support.'}, status=403)
            except Exception:
                pass

            # Get the user's school portal slug
            user_portal_slug = None
            try:
                from config.models import SchoolConfig
                user_school = SchoolConfig.get_config(user=user)
                if user_school:
                    user_portal_slug = user_school.portal_slug
            except Exception:
                pass
            
            return JsonResponse({
                'success': True,
                'user': {
                    'id': user.id,
                    'username': user.username,
                    'email': user.email,
                    'first_name': user.first_name,
                    'last_name': user.last_name,
                    'role': role,
                    'is_superuser': user.is_superuser,
                    'teacher_id': user.teacher.id if hasattr(user, 'teacher') else None,
                    'student_id': user.student_profile.id if hasattr(user, 'student_profile') else None
                },
                'redirect_url': redirect_url,
                'portal_slug': user_portal_slug if user_portal_slug else None,
                'subscription': {
                    'plan': user_school.subscription_plan if user_school else 'Basic',
                    'status': user_school.subscription_status if user_school else 'Free',
                    'trial_end': user_school.trial_end_date.isoformat() if user_school and user_school.trial_end_date else None
                } if user_school else None
            })
        else:
            print("DEBUG: Authentication failed")
            return JsonResponse({'error': 'Invalid username/email or password'}, status=401)
            
    except json.JSONDecodeError:
        print("DEBUG: JSON Decode Error")
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        import traceback
        print(f"DEBUG: Exception: {str(e)}")
        print(traceback.format_exc())
        with open('login_debug.log', 'w') as f:
            f.write(f"Error: {str(e)}\n\nTraceback:\n{traceback.format_exc()}")
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["POST"])
def api_school_signup(request):
    """
    API endpoint for school registration.
    Creates a new school and an associate admin user.
    """
    try:
        from django.contrib.auth.models import User
        from config.models import SchoolConfig, SystemSettings
        import json
        
        # Check if registration is open
        sys_settings = SystemSettings.load()
        if not sys_settings.registration_open:
            return JsonResponse({'error': 'Registration is currently closed. Please contact support.'}, status=403)

        # Handle multipart or JSON
        ct = request.content_type or ''
        if 'multipart/form-data' in ct:
            data = request.POST.dict()
            files = request.FILES
        else:
            data = json.loads(request.body)
            files = {}
            
        school_name = data.get('school_name')
        school_code = data.get('school_code')
        school_email = data.get('school_email')
        school_phone = data.get('school_phone')
        school_address = data.get('school_address')
        password = data.get('password') or data.get('school_password')
        plan = data.get('plan', 'Basic') # Default to Basic if not provided
        
        if not school_name or not school_code or not school_email or not password:
            return JsonResponse({'error': 'All important data (Name, Code, Email, Password) must be provided'}, status=400)
            
        if User.objects.filter(username=school_email).exists() or User.objects.filter(email=school_email).exists():
           return JsonResponse({'error': 'A user with this email already exists'}, status=400)

        from django.utils import timezone
        from datetime import timedelta
        
        trial_start = timezone.now()
        # Force 7 days trial period as per requirement (overriding potentially stale DB setting)
        trial_days = 7
        # trial_days = sys_settings.trial_days if sys_settings.trial_days > 0 else 7
        trial_end = trial_start + timedelta(days=trial_days)

        with transaction.atomic():
            # 1. Create User
            user = User.objects.create_user(
                username=school_email,
                email=school_email,
                password=password,
                first_name=school_name,
                is_staff=True,
                is_superuser=True # Making them superuser for now so they have full access to their "school"
            )
            
            # 2. Create School Config
            school = SchoolConfig.objects.create(
                admin=user,
                school_name=school_name,
                school_code=school_code,
                school_email=school_email,
                school_phone=school_phone,
                school_address=school_address,
                subscription_plan=plan,
                subscription_status='Trial',
                trial_start_date=trial_start,
                trial_end_date=trial_end,
                current_year=timezone.now().year
            )
            
            if 'school_logo' in files:
                school.school_logo = files['school_logo']
                school.save()
                
        return JsonResponse({
            'success': True,
            'message': 'School registered successfully!',
            'redirect_url': '/login'
        })
        
    except Exception as e:
        print(f"SIGNUP ERROR: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["GET"])
def api_student_list(request):
    """API endpoint for fetching student list with filters"""
    try:
        from django.core.paginator import Paginator
        from django.db.models import Q
        from config.models import SchoolConfig
        
        # Get Current School
        school = SchoolConfig.get_config(user=request.user, request=request)
        
        # Base Query - Filtered by School
        students = Student.objects.filter(school=school).select_related('grade').all().order_by('-created_at')
        
        # Filtering
        search = request.GET.get('search', '')
        if search:
            students = students.filter(
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(admission_number__icontains=search)
            )
            
        grade_id = request.GET.get('grade')
        if grade_id:
            students = students.filter(grade_id=grade_id)
            
        status = request.GET.get('status') # e.g., 'active' or 'graduated' (if you had a status field, assuming logic here)
            
        # Pagination
        page = request.GET.get('page', 1)
        per_page = request.GET.get('per_page', 100)
        paginator = Paginator(students, per_page)
        
        try:
            paginated_students = paginator.page(page)
        except:
            paginated_students = paginator.page(1)
            
        # Serialization
        data = [{
            'id': s.id,
            'admission_number': s.admission_number,
            'first_name': s.first_name,
            'last_name': s.last_name,
            'full_name': s.get_full_name(),
            'grade': s.grade.name if s.grade else 'N/A',
            'gender': s.gender,
            'parent_phone': s.parent_phone,
            'balance': float(s.get_balance() or 0),
            'status': 'ACTIVE', # Placeholder, replace with actual status logic
            'photo': s.photo.url if s.photo else None,
            'branch': s.branch.name if s.branch else 'N/A'
        } for s in paginated_students]
        
        # Calculate stats on the full filtered queryset (not just paginated)
        stats = {
            'fully_paid': 0,
            'total_count': students.count()
        }
        
        try:
            # Annotation to calculate balance in DB
            # Balance = (Fees due based on current_term) - (Total Payments)
            
            # 1. Total Payments
            students_with_payments = students.annotate(
                total_paid_annot=Coalesce(Sum('payments__amount'), Value(0), output_field=DecimalField())
            )
            
            # 2. Total Due (Dynamic based on current_term)
            students_with_balance = students_with_payments.annotate(
                fees_due_annot=Case(
                    When(current_term=1, then=F('term1_fees')),
                    When(current_term=2, then=F('term1_fees') + F('term2_fees')),
                    When(current_term=3, then=F('term1_fees') + F('term2_fees') + F('term3_fees')),
                    default=Value(0),
                    output_field=DecimalField()
                )
            ).annotate(
                balance_annot=F('fees_due_annot') - F('total_paid_annot')
            )
            
            # 3. Count where balance <= 0
            stats['fully_paid'] = students_with_balance.filter(balance_annot__lte=0).count()
            
            # 4. Gender Stats
            gender_stats = students.values('gender').annotate(count=Count('id'))
            for item in gender_stats:
                if item['gender'] in ['M', 'Male']:
                    stats['male_count'] = item['count']
                elif item['gender'] in ['F', 'Female']:
                    stats['female_count'] = item['count']
            
            stats.setdefault('male_count', 0)
            stats.setdefault('female_count', 0)

        except Exception as e:
            print(f"Error calculating stats: {e}")
            # Fallback (optional, or just leave as 0)
        
        return JsonResponse({
            'students': data,
            'total': paginator.count,
            'pages': paginator.num_pages,
            'current_page': int(page),
            'stats': stats
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["GET"])
@login_required
def api_grades(request):
    """API endpoint for fetching grades"""
    try:
        from config.models import SchoolConfig
        school = SchoolConfig.get_config(user=request.user, request=request)
        
        grades = Grade.objects.filter(school=school).select_related('class_teacher').all().order_by('id') 
        
        # If no grades exist for this school, clone from global templates or defaults
        if not grades.exists():
            global_grades = Grade.objects.filter(school__isnull=True)
            if global_grades.exists():
                for gg in global_grades:
                    Grade.objects.create(
                        school=school,
                        name=gg.name,
                        description=gg.description,
                        term1_fees=gg.term1_fees,
                        term2_fees=gg.term2_fees,
                        term3_fees=gg.term3_fees,
                        is_active=gg.is_active
                    )
            else:
                # Fallback: Create from standard defaults
                for code, label in Grade.GRADE_CHOICES:
                    Grade.objects.create(
                        school=school,
                        name=code,
                        description=label,
                        term1_fees=0,
                        term2_fees=0,
                        term3_fees=0,
                        is_active=True
                    )

            # Re-fetch
            grades = Grade.objects.filter(school=school).select_related('class_teacher').all().order_by('id')
            
        data = [{
            'id': g.id, 
            'name': g.get_name_display(), # Send "Grade 1" instead of "G1"
            'code': g.name,
            'class_teacher_id': g.class_teacher.id if g.class_teacher else None,
            'term1_fees': float(g.term1_fees),
            'term2_fees': float(g.term2_fees),
            'term3_fees': float(g.term3_fees),
        } for g in grades]
        return JsonResponse({'grades': data})
    except Exception as e:
        import traceback, logging
        logging.error(f"Grade API Error: {str(e)}")
        logging.error(traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["POST"])
@login_required
def api_grade_update_fees(request, pk):
    """
    API endpoint to update fees for a specific grade and propagate to students.
    """
    try:
        import json
        data = json.loads(request.body)
        grade = Grade.objects.get(pk=pk)
        
        term1_fees = data.get('term1_fees')
        term2_fees = data.get('term2_fees')
        term3_fees = data.get('term3_fees')
        
        if term1_fees is not None:
            grade.term1_fees = term1_fees
        if term2_fees is not None:
            grade.term2_fees = term2_fees
        if term3_fees is not None:
            grade.term3_fees = term3_fees
            
        grade.save()
        
        # Propagate to students
        update_fields = {}
        if term1_fees is not None:
            update_fields['term1_fees'] = term1_fees
        if term2_fees is not None:
            update_fields['term2_fees'] = term2_fees
        if term3_fees is not None:
            update_fields['term3_fees'] = term3_fees
            
        if update_fields:
            Student.objects.filter(grade=grade).update(**update_fields)
            
        return JsonResponse({'success': True, 'message': 'Fees updated successfully for grade and students'})
        
    except Grade.DoesNotExist:
        return JsonResponse({'error': 'Grade not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["GET", "POST"])
@login_required
def api_branches(request):
    """API endpoint for fetching and creating branches"""
    from config.models import SchoolConfig
    school = SchoolConfig.get_config(user=request.user, request=request)
    
    if request.method == 'GET':
        try:
            # Check for any branch (active or inactive) to avoid unique constraint violations during seeding
            any_branches = Branch.objects.filter(school=school)
            
            # Seed default 'Main School' branch if absolutely none exist
            if not any_branches.exists():
                Branch.objects.create(
                    school=school,
                    name='Main School',
                    address=school.school_address or '',
                    contact_phone=school.school_phone or ''
                )
            
            # Now fetch only active branches for the UI
            branches = Branch.objects.filter(school=school, is_active=True).order_by('name')
            data = [{'id': b.id, 'name': b.name, 'address': b.address, 'phone': b.contact_phone} for b in branches]
            return JsonResponse({'branches': data})
        except Exception as e:
            import traceback
            print(f"ERROR in api_branches GET: {str(e)}")
            print(traceback.format_exc())
            return JsonResponse({'error': str(e)}, status=500)
    
    elif request.method == 'POST':
        try:
            import json
            data = json.loads(request.body)
            name = data.get('name')
            if not name:
                return JsonResponse({'error': 'Branch name is required'}, status=400)
            
            if Branch.objects.filter(school=school, name=name).exists():
                 return JsonResponse({'error': 'Branch already exists'}, status=400)

            branch = Branch.objects.create(
                school=school,
                name=name,
                address=data.get('address', ''),
                contact_phone=data.get('phone', '')
            )
            return JsonResponse({'success': True, 'branch': {'id': branch.id, 'name': branch.name}})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["DELETE"])
@login_required
def api_branch_delete(request, pk):
    try:
        branch = Branch.objects.get(pk=pk)
        branch.delete()
        return JsonResponse({'success': True})
    except Branch.DoesNotExist:
         return JsonResponse({'error': 'Branch not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["DELETE"])
@login_required
def api_student_bulk_delete(request):
    """Delete all students and their associated users for the current school"""
    try:
        from config.models import SchoolConfig
        school = SchoolConfig.get_config(user=request.user, request=request)
        
        # Get all students for this school with their users
        students = Student.objects.filter(school=school).select_related('user')
        count = students.count()
        for student in students:
            if student.user:
                student.user.delete()
            student.delete()
        return JsonResponse({'success': True, 'message': f'All {count} students deleted successfully'})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["GET", "POST", "DELETE"])
@login_required
def api_student_detail(request, pk):
    """API endpoint for fetching and updating single student details"""
    try:
        from django.db.models import Sum
        from config.models import SchoolConfig
        school = SchoolConfig.get_config(user=request.user, request=request)
        student = Student.objects.select_related('grade').get(pk=pk, school=school)

        if request.method == 'DELETE':
            # Also delete associated user if it exists
            if student.user:
                student.user.delete()
            student.delete()
            return JsonResponse({'success': True, 'message': 'Student deleted successfully'})
        
        if request.method == 'POST':
            # Handle Update
            try:
                # Personal
                if 'first_name' in request.POST: student.first_name = request.POST['first_name']
                if 'last_name' in request.POST: student.last_name = request.POST['last_name']
                if 'date_of_birth' in request.POST: student.date_of_birth = request.POST['date_of_birth']
                if 'gender' in request.POST: student.gender = 'Male' if request.POST['gender'] == 'M' else 'Female'
                if 'location' in request.POST: student.location = request.POST['location']
                if 'grade' in request.POST and request.POST['grade']:
                    student.grade_id = request.POST['grade']
                if 'academic_year' in request.POST: student.academic_year = request.POST['academic_year']
                if 'current_term' in request.POST: student.current_term = request.POST['current_term']
                if 'branch' in request.POST and request.POST['branch']:
                    student.branch_id = request.POST['branch']

                # Guardian
                if 'parent_name' in request.POST: student.parent_name = request.POST['parent_name']
                if 'parent_phone' in request.POST: student.parent_phone = request.POST['parent_phone']
                if 'parent_email' in request.POST: student.parent_email = request.POST['parent_email']

                # Photo
                if 'photo' in request.FILES:
                    student.photo = request.FILES['photo']
                
                student.save()
                return JsonResponse({'success': True, 'message': 'Student updated successfully'})
            except Exception as e:
                return JsonResponse({'error': str(e)}, status=400)
        
        # GET Request - Return Details
        # Attendance Stats
        attendance_stats = student.get_attendance_stats()
        
        # Financial History
        payments = student.payments.all().order_by('-date')
        payment_list = [{
            'id': p.id,
            'amount': float(p.amount),
            'date': p.date.strftime('%Y-%m-%d'),
            'method': p.payment_method,
            'reference': p.transaction_id,
            'term': p.term or student.current_term 
        } for p in payments]
        
        # Simple Assessment Summary
        position = student.get_class_position() or 'N/A'
        
        data = {
            'id': student.id,
            'personal': {
                'admission_number': student.admission_number,
                'first_name': student.first_name,
                'last_name': student.last_name,
                'full_name': student.get_full_name(),
                'gender': student.gender,
                'dob': student.date_of_birth.strftime('%Y-%m-%d'),
                'age': student.age,
                'photo': student.get_photo_url(),
                'grade': student.grade.name if student.grade else 'N/A',
                'grade_id': student.grade.id if student.grade else '',
                'grade_id': student.grade.id if student.grade else '',
                'location': student.location,
                'branch': student.branch.name if student.branch else 'N/A',
                'branch_id': student.branch.id if student.branch else '',
            },
            'academic': {
                'grade_id': student.grade.id if student.grade else '',
                'position': position,
            },
            'guardian': {
                'name': student.parent_name,
                'phone': student.parent_phone,
                'email': student.parent_email,
            },
            'financial': {
                'balance': float(student.get_balance()),
                'total_fees': float(student.get_total_fees()),
                'total_paid': float(student.get_total_paid()),
                'status': student.get_payment_status(),
                'holistic': student.get_holistic_financials(),
                'history': payment_list
            },
            'attendance': attendance_stats,
        }
        return JsonResponse({'student': data})
        
    except Student.DoesNotExist:
        return JsonResponse({'error': 'Student not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["POST"])
@login_required
def api_student_payment_add(request, student_id):
    """API endpoint to record a payment for a student"""
    try:
        import uuid
        data = json.loads(request.body)
        student = get_object_or_404(Student, pk=student_id)
        
        amount = Decimal(str(data.get('amount', '0')))
        method = data.get('payment_method', 'CASH')
        reference = data.get('reference', '')
        term = data.get('term', student.current_term or 1)
        description = data.get('description', f"Fee Payment - Term {term}")
        
        if amount <= 0:
            return JsonResponse({'error': 'Amount must be greater than zero'}, status=400)
            
        with transaction.atomic():
            # 1. Create the Payment record (Main app)
            payment = Payment.objects.create(
                student=student,
                amount=amount,
                payment_method=method,
                transaction_id=reference,
                reference_number=reference or f'REF-{uuid.uuid4().hex[:8].upper()}',
                term=term,
                status='COMPLETED'
            )
            
            # 2. Create Transaction record (Finance app) if it exists
            try:
                from finance.models import StudentFinanceAccount, Transaction
                account, _ = StudentFinanceAccount.objects.get_or_create(student=student)
                Transaction.objects.create(
                    account=account,
                    type='PAYMENT',
                    amount=amount,
                    description=description,
                    payment_method=method,
                    reference=reference or payment.reference_number,
                    term=term,
                    academic_year=student.academic_year,
                    recorded_by=request.user
                )
            except (ImportError, Exception) as e:
                # Log error but don't fail the whole transaction if finance app is missing
                print(f"Finance integration error: {str(e)}")
        
        return JsonResponse({
            'success': True, 
            'message': 'Payment recorded successfully',
            'payment_id': payment.id
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["DELETE"])
@login_required
def api_staff_bulk_delete(request):
    """Delete all staff members (Teachers, Drivers, Non-teaching)"""
    try:
        # Delete Teachers (which are Employees)
        teachers = Teacher.objects.all().select_related('user')
        for teacher in teachers:
            if teacher.user:
                teacher.user.delete()
            teacher.delete()
            
        # Delete Non-Teaching Staff (which are Employees)
        from .hr_models import NonTeachingStaff
        staff_members = NonTeachingStaff.objects.all()
        for s in staff_members:
            if hasattr(s, 'user') and s.user:
                s.user.delete()
            s.delete()
            
        # Delete Drivers
        from transport.models import TransportDriver
        drivers = TransportDriver.objects.all().select_related('user')
        for driver in drivers:
            if driver.user:
                driver.user.delete()
            driver.delete()
            
        return JsonResponse({'success': True, 'message': 'All staff members deleted successfully'})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
@csrf_exempt
@require_http_methods(["POST"])
@login_required
def api_student_create(request):
    """API endpoint for creating a new student"""
    try:
        import json
        from decimal import Decimal
        from config.models import SchoolConfig
        
        # Determine if we have multipart data (file upload) or JSON
        if request.content_type.startswith('multipart/form-data'):
            data = request.POST.copy()
            files = request.FILES
        else:
            data = json.loads(request.body)
            files = {}
        
        # Get School Config
        config = SchoolConfig.get_config(user=request.user, request=request)
        
        # Let's instantiate the form
        form = StudentForm(data, files)
        form.instance.school = config 
        
        if form.is_valid():
            # Validate Payment Details
            payment_method = data.get('payment_method', 'CASH')
            ref_number = data.get('reference_number')
            
            if payment_method in ['MPESA', 'BANK'] and not ref_number:
                return JsonResponse({'error': 'Reference number is required for M-Pesa or Bank payments'}, status=400)

            from django.db import transaction, IntegrityError
            import time
            
            # Retry logic for unique constraint (admission number collision)
            max_retries = 5
            student = None
            
            for attempt in range(max_retries):
                try:
                    # Generate/Ensure Admission Number OUTSIDE the atomic block
                    # This ensures that if the student save fails, the counter increment in SchoolConfig is NOT rolled back
                    # if we were to put it inside, the rollback would reset the counter, causing an infinite loop of collisions
                    if not data.get('admission_number') or (attempt > 0 and not data.get('admission_number')):
                         # Force a refresh of config from DB to get latest counter
                        config.refresh_from_db()
                        admission_number = config.generate_admission_number(grade=data.get('grade'))
                    else:
                        admission_number = data.get('admission_number')

                    with transaction.atomic():
                        student = form.save(commit=False)
                        student.school = config
                        student.admission_number = admission_number
                        student.admission_fee = config.admission_fee
                        student.save()
                        form.save_m2m()
                        
                        # If we get here, save was successful
                        break 
                except IntegrityError as e:
                    if "admission_number" in str(e):
                        if attempt < max_retries - 1:
                            # If we have retries left, generate a new number and try again
                            print(f"Collision detected for {admission_number}, retrying...")
                            continue
                        else:
                            return JsonResponse({'error': 'Could not generate a unique admission number. Please try manual entry.'}, status=400)
                    else:
                        raise e  # Re-raise other integrity errors

            # Payment Handling
            import uuid
            payment_method = data.get('payment_method', 'CASH')
            amount = config.admission_fee
            
            ref_number = data.get('reference_number')
            if not ref_number:
                    ref_number = f'REF-{uuid.uuid4().hex[:8].upper()}'
                
            # Create Payment Record
            Payment.objects.create(
                school=config,
                student=student,
                amount=config.admission_fee,
                payment_method=payment_method,
                transaction_id=ref_number,
                reference_number=ref_number,
                term=student.current_term or 1,
                status='COMPLETED' if payment_method == 'CASH' else 'PENDING'
            )
            
            # Create User Account for Student
            try:
                from django.contrib.auth.models import User, Group
                if not student.user and student.admission_number:
                    username = student.admission_number
                    # Use student portal password from config
                    password = config.student_portal_password
                    
                    # Check if user exists by username
                    user = User.objects.filter(username=username).first()
                    if not user:
                        user = User.objects.create_user(
                            username=username,
                            password=password,
                            first_name=student.first_name,
                            last_name=student.last_name
                        )
                         
                        # Add to Student Group
                        group, _ = Group.objects.get_or_create(name='Students')
                        user.groups.add(group)
                    
                    student.user = user
                    student.save()
            except Exception as e:
                import traceback
                print(f"Error creating student user: {e}")
                print(traceback.format_exc())
                # Non-fatal error, but good to log
            
            return JsonResponse({
                'success': True, 
                'student_id': student.id,
                'admission_number': student.admission_number,
                    'username': student.admission_number
                })
        else:
            return JsonResponse({'error': 'Validation Failed', 'errors': form.errors}, status=400)
            
    except Exception as e:
        import traceback
        error_info = traceback.format_exc()
        print(f"API Error: {error_info}")
        return JsonResponse({
            'error': str(e),
            'details': error_info
        }, status=500)


@csrf_exempt
@require_http_methods(["GET"])
def api_dashboard(request):
    """API endpoint for dashboard statistics"""
    try:
        from django.db import transaction
        import uuid
        from django.db.models import Sum, Count, F, Value, DecimalField, Q
        from django.db.models.functions import Coalesce
        from datetime import  timedelta
        import calendar
        
        from config.models import SchoolConfig
        
        # Get Currrent School
        school = SchoolConfig.get_config(user=request.user, request=request)
        
        today = timezone.now().date()
        
        # Student Stats
        all_students = Student.objects.filter(school=school)
        total_students = all_students.count()
        new_students = all_students.filter(
            created_at__month=timezone.now().month,
            created_at__year=timezone.now().year
        ).count()
        
        # Teacher Stats
        teacher_stats = Teacher.objects.filter(school=school).aggregate(
            total=Count('id'),
            active=Count('id', filter=Q(status='ACTIVE'))
        )
        
        # Attendance Stats
        todays_attendance = Attendance.objects.filter(school=school, date=today)
        present_today = todays_attendance.filter(status='PRESENT').count()
        attendance_rate = (present_today / total_students * 100) if total_students > 0 else 0
        
        # Fee Stats (School Fees Target)
        # 1. Tuition Target (Cumulative year)
        tuition_target = all_students.aggregate(
            total=Sum(F('admission_fee') + F('term1_fees') + F('term2_fees') + F('term3_fees'))
        )['total'] or 0

        # 2. Collected Tuition
        tuition_collected = all_students.aggregate(
            total=Coalesce(Sum('payments__amount'), Value(0), output_field=DecimalField(max_digits=12, decimal_places=2))
        )['total'] or 0

        # 3. Transport Target (System estimation)
        try:
            from transport.models import TransportAssignment, TransportTransaction
            transport_target = TransportAssignment.objects.filter(account__school=school, active=True).aggregate(
                total=Sum(F('route__cost_per_term') * 3)
            )['total'] or 0
            transport_collected = TransportTransaction.objects.filter(school=school, type='PAYMENT').aggregate(sum=Sum('amount'))['sum'] or 0
        except:
            transport_target = 0
            transport_collected = 0

        # 4. Food Target (System estimation)
        try:
            from food.models import FoodSubscription, FoodTransaction
            food_target = FoodSubscription.objects.filter(account__school=school, active=True).aggregate(
                total=Sum(F('meal_item__cost') * 3)
            )['total'] or 0
            food_collected = FoodTransaction.objects.filter(school=school, type='PAYMENT').aggregate(sum=Sum('amount'))['sum'] or 0
            
            # Legacy food income (if any)
            try:
                food_income_fee = FoodFee.objects.filter(school=school, status='COMPLETED').aggregate(sum=Sum('amount'))['sum'] or 0
                food_income_meal = StudentMealPayment.objects.filter(school=school, status='COMPLETED').aggregate(sum=Sum('amount'))['sum'] or 0
                food_collected = float(food_collected) + float(food_income_fee) + float(food_income_meal)
            except:
                pass
        except:
            food_target = 0
            food_collected = 0
        
        # Combined KPI Totals
        expected_total = float(tuition_target) + float(transport_target) + float(food_target)
        collected_total = float(tuition_collected) + float(transport_collected) + float(food_collected)
        
        # Override local variables for simple JSON return if needed
        total_fees = expected_total
        fees_collected = collected_total
        outstanding = max(0, expected_total - collected_total)
        
        # Chart Data (Last 6 months)
        chart_labels = []
        chart_data = []
        
        for i in range(5, -1, -1):
            current = today - timedelta(days=today.day - 1) - timedelta(days=30*i)
            month_start = current.replace(day=1)
            month_end = current.replace(day=calendar.monthrange(current.year, current.month)[1])
            
            month_total = Payment.objects.filter(
                school=school,
                date__gte=month_start,
                date__lte=month_end
            ).aggregate(
                total=Coalesce(Sum('amount'), Value('0.00'), output_field=DecimalField(max_digits=10, decimal_places=2))
            )['total']
            
            chart_labels.append(current.strftime('%b'))
            chart_data.append(float(month_total))
            
        recent_payments = Payment.objects.filter(school=school).select_related('student').order_by('-date')[:5]
        recent_payments_data = [{
            'id': p.id,
            'student_name': p.student.get_full_name(),
            'amount': float(p.amount),
            'date': p.date.strftime('%Y-%m-%d'),
            'method': p.payment_method
        } for p in recent_payments]

        return JsonResponse({
            'stats': {
                'total_students': total_students,
                'new_students': new_students,
                'total_teachers': teacher_stats['total'],
                'attendance_rate': round(attendance_rate, 1),
                'total_revenue': float(collected_total or 0),
                'revenue_growth': 12.5, # Placeholder or calc real growth
            },
            'chart_data': {
                'labels': chart_labels,
                'revenue': chart_data
            },
            'recent_payments': recent_payments_data
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["GET"])
@login_required
def api_subjects(request):
    """API endpoint for fetching subjects"""
    try:
        from config.models import SchoolConfig
        school = SchoolConfig.get_config(user=request.user, request=request)
        from django.db.models import Q
        subjects = Subject.objects.filter(Q(school=school) | Q(school__isnull=True)).order_by('name')
        data = [{'id': s.id, 'name': s.name, 'code': s.code} for s in subjects]
        return JsonResponse({'subjects': data})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["GET"])
def api_teacher_list(request):
    """API endpoint for fetching teachers list"""
    try:
        from django.core.paginator import Paginator, EmptyPage
        from config.models import SchoolConfig
        
        # Get Current School
        school = SchoolConfig.get_config(user=request.user, request=request)
        
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 100))
        search = request.GET.get('search', '')
        status = request.GET.get('status', '')

        teachers = Teacher.objects.filter(school=school).select_related('grade').prefetch_related('subjects').order_by('first_name')

        if search:
            teachers = teachers.filter(
                Q(first_name__icontains=search) | 
                Q(last_name__icontains=search) |
                Q(tsc_number__icontains=search) |
                Q(email__icontains=search)
            )

        if status:
            teachers = teachers.filter(status=status.upper())

        paginator = Paginator(teachers, per_page)
        
        try:
            current_page = paginator.page(page)
        except EmptyPage:
            current_page = paginator.page(paginator.num_pages)
            
        data = []
        for t in current_page:
            try:
                subjects = [s.name for s in t.subjects.all()]
            except Exception:
                subjects = []
                
            try:
                avatar = t.profile_picture.url if t.profile_picture and t.profile_picture.name else None
            except Exception:
                avatar = None

            data.append({
                'id': t.id,
                'full_name': t.get_full_name(),
                'tsc_number': t.tsc_number or 'N/A',
                'email': t.email,
                'phone': t.phone,
                'role': t.get_position_display(),
                'subjects': subjects,
                'status': t.status,
                'avatar': avatar,
            })

        return JsonResponse({
            'teachers': data,
            'total': paginator.count,
            'pages': paginator.num_pages,
            'current_page': current_page.number
        })
    except Exception as e:
        import traceback
        with open('api_teacher_error.log', 'a') as f:
            f.write(f"Error: {str(e)}\n{traceback.format_exc()}\n")
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["GET", "POST", "DELETE"])
@login_required
def api_teacher_detail(request, pk):
    """API endpoint for fetching and updating single teacher details"""
    try:
        # Try to get teacher by ID first
        teacher = Teacher.objects.filter(pk=pk).first()
        
        # If not found, try to find by User ID (in case frontend passed User ID)
        if not teacher:
            teacher = Teacher.objects.filter(user__id=pk).first()
            
        if not teacher:
            return JsonResponse({'error': 'Teacher not found'}, status=404)
        
        if request.method == 'DELETE':
            # Also delete associated user if it exists
            if teacher.user:
                teacher.user.delete()
            teacher.delete()
            return JsonResponse({'success': True, 'message': 'Teacher deleted successfully'})
        
        if request.method == 'POST':
            # Handle Update
            try:
                # Personal Information
                if request.POST.get('first_name'): teacher.first_name = request.POST['first_name']
                if request.POST.get('last_name'): teacher.last_name = request.POST['last_name']
                if request.POST.get('email'): teacher.email = request.POST['email']
                if request.POST.get('phone'): teacher.phone = request.POST['phone']
                if request.POST.get('national_id'): teacher.national_id = request.POST['national_id']
                if request.POST.get('address'): teacher.address = request.POST['address']
                if request.POST.get('dob'): teacher.date_of_birth = request.POST['dob']
                
                gender_map = {'Male': 'M', 'Female': 'F'}
                if request.POST.get('gender'): 
                    gender_val = request.POST['gender']
                    teacher.gender = gender_map.get(gender_val, gender_val)

                if request.POST.get('marital_status'): teacher.marital_status = request.POST['marital_status']
                if request.POST.get('religion'): teacher.religion = request.POST['religion']
                if request.POST.get('nationality'): teacher.nationality = request.POST['nationality']

                # Professional Details
                if request.POST.get('tsc_number'): teacher.tsc_number = request.POST['tsc_number']
                if request.POST.get('date_joined'): teacher.date_joined = request.POST['date_joined']
                if request.POST.get('status'): teacher.status = request.POST['status']
                if request.POST.get('position'): teacher.position = request.POST['position']
                if request.POST.get('qualifications'): teacher.qualifications = request.POST['qualifications']
                if request.POST.get('years_of_experience'): teacher.years_of_experience = request.POST['years_of_experience']

                if request.POST.get('department'):
                    teacher.department_id = request.POST['department']
                
                if request.POST.get('branch'):
                    teacher.branch_id = request.POST['branch']

                # Class Teacher Logic
                if 'is_class_teacher' in request.POST:
                    # Check if 'true' string or boolean true
                    val = request.POST['is_class_teacher']
                    teacher.is_class_teacher = val == 'true' or val == True or val == 'True'
                
                if request.POST.get('grade'):
                    teacher.grade_id = request.POST['grade']
                elif teacher.is_class_teacher == False:
                    teacher.grade = None # Remove grade if not class teacher

                # Handle Subjects (M2M)
                # FormData sends multiple keys for 'subjects', use getlist
                if 'subjects' in request.POST:
                    # If subjects are being updated, we expect them in the payload
                    # Note: getlist might be empty if all unchecked, frontend usually handles this
                    subject_ids = request.POST.getlist('subjects')
                    teacher.subjects.set(subject_ids)
                
                # Handle Photo Upload
                if 'avatar' in request.FILES:
                    teacher.profile_picture = request.FILES['avatar']
                
                # Handle Certificate
                if 'certificate' in request.FILES:
                    teacher.certificate = request.FILES['certificate']
                
                teacher.save()
                return JsonResponse({'success': True, 'message': 'Teacher updated successfully'})
            except Exception as e:
                import traceback
                traceback.print_exc()
                return JsonResponse({'error': str(e)}, status=400)
        
        
        # Calculate some stats
        students_count = 0
        try:
            if teacher.grade:
                students_count = Student.objects.filter(grade=teacher.grade).count()
        except:
            students_count = 0
        
        # Fetch upcoming schedule
        schedules = Schedule.objects.filter(teacher=teacher).select_related('subject', 'grade').order_by('day', 'start_time')
        
        schedule_data = []
        for s in schedules:
            try:
                # Safe time formatting
                start = s.start_time.strftime('%H:%M') if s.start_time else "00:00"
                end = s.end_time.strftime('%H:%M') if s.end_time else "00:00"

                schedule_data.append({
                    'id': s.id,
                    'day': s.get_day_display(),
                    'subject': s.subject.name if s.subject else "Unknown Subject",
                    'grade': s.grade.name if s.grade else "Unknown Grade",
                    'time': f"{start} - {end}",
                    'term': s.get_term_display()
                })
            except Exception:
                continue

        # Get salary info from finance app if available
        basic_salary = float(teacher.basic_salary or 0)
        net_salary = basic_salary
        allowances = 0
        deductions = 0
        nssf = 0
        loans = 0
        advances = 0

        if teacher.user:
            try:
                from finance.models import SalaryStructure
                ss, _ = SalaryStructure.objects.get_or_create(user=teacher.user)
                basic_salary = float(ss.base_salary)
                allowances = float(ss.allowances)
                deductions = float(ss.deductions)
                nssf = float(ss.nssf)
                loans = float(ss.loans)
                advances = float(ss.get_advances())
                net_salary = float(ss.net_salary())
            except Exception as e:
                print(f"Error fetching salary structure: {e}")
                # Fallback to model property if it exists
                try: net_salary = float(teacher.net_salary)
                except: pass

        data = {
            'personal': {
                'id': teacher.id,
                'full_name': teacher.get_full_name(),
                'first_name': teacher.first_name,
                'last_name': teacher.last_name,
                'email': teacher.email,
                'phone': teacher.phone,
                'national_id': teacher.national_id,
                'gender': teacher.gender,
                'gender_display': teacher.get_gender_display(),
                'dob': teacher.date_of_birth.strftime('%Y-%m-%d') if teacher.date_of_birth else None,
                'marital_status': teacher.marital_status,
                'religion': teacher.religion,
                'nationality': teacher.nationality,
                'address': teacher.address,
                'avatar': teacher.profile_picture.url if teacher.profile_picture else None,
            },
            'professional': {
                'is_class_teacher': teacher.is_class_teacher,
                'grade_assigned': teacher.grade.name if teacher.grade else None,
                'grade_id': teacher.grade.id if teacher.grade else None,
                'tsc_number': teacher.tsc_number,
                'position': teacher.position,
                'position_display': teacher.get_position_display(),
                'department_id': teacher.department.id if teacher.department else None,
                'department_name': teacher.department.name if teacher.department else None,
                'branch_id': teacher.branch.id if teacher.branch else None,
                'branch_name': teacher.branch.name if teacher.branch else None,
                'date_joined': teacher.date_joined.strftime('%Y-%m-%d') if teacher.date_joined else None,
                'status': teacher.status,
                'qualifications': teacher.qualifications,
                'qualifications_display': teacher.get_qualifications_display(),
                'years_of_experience': teacher.years_of_experience,
                'subjects': [{'id': s.id, 'name': s.name, 'code': s.code} for s in teacher.subjects.all()],
                'subject_ids': [s.id for s in teacher.subjects.all()],
                'certificate': teacher.certificate.url if teacher.certificate else None,
                'certificate_name': teacher.certificate.name.split('/')[-1] if teacher.certificate else None,
                'students_count': students_count
            },
            'financial': {
                'basic_salary': basic_salary,
                'allowances': allowances,
                'deductions': deductions,
                'nssf': nssf,
                'loans': loans,
                'advances': advances,
                'net_salary': net_salary,
            },
            'leaves': [{
                'type': l.get_leave_type_display(),
                'start_date': l.start_date.strftime('%Y-%m-%d'),
                'end_date': l.end_date.strftime('%Y-%m-%d'),
                'status': l.status,
                'reason': l.reason
            } for l in teacher.leaves.all().order_by('-start_date')[:5]],
            'schedule': schedule_data
        }
        
        # Add advances safely
        try:
            advances_data = []
            if hasattr(teacher, 'salary_advances'):
                 # Check if the relationship exists
                 advances = teacher.salary_advances.all().order_by('-date_requested')[:5]
                 for a in advances:
                     advances_data.append({
                        'amount': float(a.amount),
                        'date': a.date_requested.strftime('%Y-%m-%d'),
                        'status': a.status,
                        'reason': a.reason
                     })
            data['advances'] = advances_data
        except Exception:
             # Fallback if the relationship is missing or query fails
            data['advances'] = []

        return JsonResponse({'teacher': data})
    except Teacher.DoesNotExist:
        return JsonResponse({'error': 'Teacher not found'}, status=404)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': f"Internal Server Error: {str(e)}"}, status=500)

@csrf_exempt
@require_http_methods(["GET"])
@login_required
def api_finance_stats(request):
    """API endpoint for finance dashboard statistics"""
    try:
        from django.db.models import Sum, Count, F, Value, DecimalField
        from django.db.models.functions import Coalesce, TruncMonth
        from datetime import  timedelta
        import calendar
        
        today = timezone.now().date()
        year = request.GET.get('year', today.year)
        
        # 1. Summary Stats
        # Total Fees Expected (Sum of term_fees for all active students)
        total_expected = Student.objects.aggregate(
            total=Coalesce(Sum('term_fees'), Value(0), output_field=DecimalField())
        )['total']
        
        # Total Collected (Sum of all payments)
        total_collected = Payment.objects.aggregate(
            total=Coalesce(Sum('amount'), Value(0), output_field=DecimalField())
        )['total']
        
        # Calculate pending
        total_expected_float = float(total_expected or 0)
        total_collected_float = float(total_collected or 0)
        total_pending = max(0, total_expected_float - total_collected_float)
        
        # Collection Rate
        collection_rate = (total_collected_float / total_expected_float * 100) if total_expected_float > 0 else 0

        # 2. Monthly Trends (Last 12 months)
        monthly_labels = []
        monthly_revenue = []
        
        for i in range(11, -1, -1):
            current = today - timedelta(days=30*i) # Approx month
            month_start = current.replace(day=1)
            month_end = current.replace(day=calendar.monthrange(current.year, current.month)[1])
            
            month_total = Payment.objects.filter(
                date__gte=month_start,
                date__lte=month_end
            ).aggregate(
                total=Coalesce(Sum('amount'), Value(0), output_field=DecimalField())
            )['total']
            
            monthly_labels.append(current.strftime('%b'))
            monthly_revenue.append(float(month_total or 0))

        # 3. Payment Methods Breakdown
        payment_methods = Payment.objects.values('payment_method').annotate(
            total=Sum('amount'),
            count=Count('id')
        ).order_by('-total')
        
        methods_data = [{
            'name': pm['payment_method'],
            'value': float(pm['total'] or 0),
            'count': pm['count']
        } for pm in payment_methods]
        
        # 4. Recent Transactions
        recent_tx = Payment.objects.select_related('student').order_by('-date')[:10]
        recent_data = [{
            'id': t.id,
            'student': t.student.get_full_name(),
            'admission_no': t.student.admission_number,
            'amount': float(t.amount),
            'date': t.date.strftime('%Y-%m-%d'),
            'method': t.payment_method,
            'reference': t.transaction_id or 'N/A'
        } for t in recent_tx]

        return JsonResponse({
            'summary': {
                'expected': total_expected_float,
                'collected': total_collected_float,
                'pending': total_pending,
                'rate': round(collection_rate, 1)
            },
            'chart_data': {
                'labels': monthly_labels,
                'revenue': monthly_revenue
            },
            'methods': methods_data,
            'recent_transactions': recent_data
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["POST"])
@login_required
def api_payment_create(request):
    """API endpoint for recording a new payment"""
    try:
        import json
        import uuid
        from decimal import Decimal
        from django.shortcuts import get_object_or_404
        from django.utils import timezone
        from .models import Student, Payment # Assuming these models are in the same app
        
        data = json.loads(request.body)
        
        # Validation
        student_id = data.get('student_id')
        if not student_id:
             return JsonResponse({'error': 'Student is required'}, status=400)
             
        student = get_object_or_404(Student, pk=student_id)
        
        amount = data.get('amount')
        if not amount:
             return JsonResponse({'error': 'Amount is required'}, status=400)
             
        payment_method = data.get('payment_method', 'CASH')
        reference = data.get('reference_number')
        
        if payment_method != 'CASH' and not reference:
             return JsonResponse({'error': 'Reference number is required for non-cash payments'}, status=400)
             
        # Create Payment
        payment = Payment.objects.create(
            student=student,
            amount=Decimal(str(amount)),
            payment_method=payment_method,
            transaction_id=reference or f'CSH-{uuid.uuid4().hex[:8].upper()}',
            term=data.get('term', student.current_term)
        )
        
        return JsonResponse({'success': True, 'payment_id': payment.id})
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["GET"])
@login_required
def api_transactions_list(request):
    """API endpoint for fetching all transactions"""
    try:
        transactions = Payment.objects.select_related('student').order_by('-date', '-created_at')
        
        data = [{
            'id': t.id,
            'student': t.student.get_full_name(),
            'admission_no': t.student.admission_number,
            'amount': float(t.amount),
            'date': t.date.strftime('%Y-%m-%d'),
            'method': t.payment_method,
            'reference': t.transaction_id or 'N/A'
        } for t in transactions]
        
        return JsonResponse({'transactions': data})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["GET", "POST"])
@login_required
def api_schedule(request):
    """API endpoint for fetching and creating schedule/timetable"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Extract fields
            teacher_id = data.get('teacher')
            subject_id = data.get('subject')
            grade_id = data.get('grade')
            day = data.get('day')
            start_time = data.get('start_time')
            end_time = data.get('end_time')
            
            # Validation
            if not all([teacher_id, subject_id, grade_id, day, start_time, end_time]):
                return JsonResponse({'error': 'Missing required fields'}, status=400)
                
            # Check conflicts
            # 1. Teacher conflict
            teacher_conflict = Schedule.objects.filter(
                teacher_id=teacher_id,
                day=day,
                start_time__lt=end_time,
                end_time__gt=start_time
            ).exists()
            
            if teacher_conflict:
                return JsonResponse({'error': 'Teacher is already booked for this time slot'}, status=400)
                
            # 2. Grade/Class conflict
            grade_conflict = Schedule.objects.filter(
                grade_id=grade_id,
                day=day,
                start_time__lt=end_time,
                end_time__gt=start_time
            ).exists()
            
            if grade_conflict:
                return JsonResponse({'error': 'Class already has a session in this time slot'}, status=400)
            
            # Create schedule
            schedule = Schedule.objects.create(
                school=school,
                teacher_id=teacher_id,
                subject_id=subject_id,
                grade_id=grade_id,
                day=day,
                start_time=start_time,
                end_time=end_time
            )
            
            return JsonResponse({'success': True, 'id': schedule.id})
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    try:
        from config.models import SchoolConfig
        school = SchoolConfig.get_config(user=request.user, request=request)
        
        grade_id = request.GET.get('grade')
        teacher_id = request.GET.get('teacher')
        
        schedules = Schedule.objects.filter(school=school).select_related('teacher', 'subject', 'grade', 'teacher__user').all()
        
        if grade_id:
            schedules = schedules.filter(grade_id=grade_id)
        if teacher_id:
            schedules = schedules.filter(teacher_id=teacher_id)
            
        # Group by day
        days_order = {
            'MONDAY': 1, 'TUESDAY': 2, 'WEDNESDAY': 3, 
            'THURSDAY': 4, 'FRIDAY': 5, 'SATURDAY': 6, 'SUNDAY': 7
        }
        
        data = []
        for s in schedules:
            data.append({
                'id': s.id,
                'day': s.day,
                'day_order': days_order.get(s.day, 8),
                'start_time': s.start_time.strftime('%H:%M'),
                'end_time': s.end_time.strftime('%H:%M'),
                'subject': s.subject.name,
                'subject_code': s.subject.code,
                'teacher': s.teacher.get_full_name(),
                'grade': s.grade.name,
                'avatar': s.teacher.profile_picture.url if s.teacher.profile_picture else None,
                'teacher_id': s.teacher.id # Added for linking
            })
            
        # Sort by day order then start time
        data.sort(key=lambda x: (x['day_order'], x['start_time']))
        
        return JsonResponse({'schedule': data})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["DELETE"])
@login_required
def api_schedule_delete(request, pk):
    """API endpoint to delete a schedule entry"""
    try:
        schedule = get_object_or_404(Schedule, pk=pk)
        schedule.delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# API endpoints for school configuration (consolidated at end of file)

@csrf_exempt
@require_http_methods(["POST"])
@login_required
def api_regenerate_portal_slug(request):
    """API endpoint to regenerate portal slug for security"""
    try:
        from config.models import SchoolConfig
        import uuid
        
        config = SchoolConfig.get_config(user=request.user, request=request)
        
        # Generate new unique slug
        new_slug = str(uuid.uuid4())[:12]
        while SchoolConfig.objects.filter(portal_slug=new_slug).exists():
            new_slug = str(uuid.uuid4())[:12]
        
        config.portal_slug = new_slug
        config.save()
        
        return JsonResponse({
            'success': True,
            'portal_slug': config.portal_slug,
            'message': 'Portal access link regenerated successfully'
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["GET"])
@login_required
def api_search(request):
    """Global search endpoint across students, teachers, and subjects"""
    try:
        from config.models import SchoolConfig
        school = SchoolConfig.get_config(user=request.user, request=request)
        
        query = request.GET.get('q', '').strip()
        
        if not query or len(query) < 2:
            return JsonResponse({'results': []})
        
        results = []
        
        # Search Students
        students = Student.objects.filter(
            Q(school=school),
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(admission_number__icontains=query)
        ).select_related('grade')[:5]
        
        for student in students:
            results.append({
                'type': 'student',
                'id': student.id,
                'title': student.get_full_name(),
                'subtitle': f"Grade {student.grade.name} • {student.admission_number}",
                'url': f'/students/{student.id}',
                'avatar': student.photo.url if student.photo else None
            })
        
        # Search Teachers
        teachers = Teacher.objects.filter(
            Q(school=school),
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(email__icontains=query) |
            Q(tsc_number__icontains=query)
        )[:5]
        
        for teacher in teachers:
            results.append({
                'type': 'teacher',
                'id': teacher.id,
                'title': teacher.get_full_name(),
                'subtitle': f"{teacher.get_position_display()} • {teacher.email}",
                'url': f'/teachers/{teacher.id}',
                'avatar': teacher.profile_picture.url if teacher.profile_picture else None
            })

        # Search Non-Teaching Staff
        from .models import NonTeachingStaff
        staff_members = NonTeachingStaff.objects.filter(
            Q(school=school),
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(position__icontains=query)
        )[:5]
        
        for staff in staff_members:
            results.append({
                'type': 'staff',
                'id': staff.id,
                'title': staff.get_full_name(),
                'subtitle': f"{staff.position} • {staff.email}",
                'url': f'/hr/staff/{staff.id}',
                'avatar': None
            })
        
        # Search Subjects
        subjects = Subject.objects.filter(
            Q(school=school),
            Q(name__icontains=query) |
            Q(code__icontains=query)
        )[:3]
        
        for subject in subjects:
            results.append({
                'type': 'subject',
                'id': subject.id,
                'title': subject.name,
                'subtitle': f"Code: {subject.code}",
                'url': '#',
                'avatar': None
            })
        
        return JsonResponse({'results': results})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)



@csrf_exempt
@require_http_methods(["POST"])
@login_required
@csrf_exempt
@require_http_methods(["POST"])
@login_required
def api_teacher_request_leave(request, pk):
    """API endpoint for teachers to request leave"""
    try:
        teacher = get_object_or_404(Teacher, pk=pk)
        data = json.loads(request.body)
        
        Leave.objects.create(
            employee=teacher,
            leave_type=data.get('leave_type'),
            start_date=data.get('start_date'),
            end_date=data.get('end_date'),
            reason=data.get('reason'),
            status='PENDING'
        )
        return JsonResponse({'success': True, 'message': 'Leave request submitted successfully'})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@csrf_exempt
@require_http_methods(["POST"])
@login_required
def api_teacher_request_advance(request, pk):
    """API endpoint for teachers to request salary advance"""
    try:
        teacher = get_object_or_404(Teacher, pk=pk)
        data = json.loads(request.body)
        
        SalaryAdvance.objects.create(
            employee=teacher,
            amount=data.get('amount'),
            reason=data.get('reason'),
            status='PENDING'
        )
        return JsonResponse({'success': True, 'message': 'Salary advance request submitted successfully'})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@csrf_exempt
@require_http_methods(["POST"])
@login_required
def api_teacher_create(request):
    """API endpoint for creating a new teacher"""
    try:
        import json
        
        # Determine if we have multipart data (file upload)
        if request.content_type.startswith('multipart/form-data'):
            data = request.POST.copy()
            files = request.FILES
        else:
            data = json.loads(request.body)
            files = {}
            
        form = TeacherForm(data, files)
        
        if form.is_valid():
            from config.models import SchoolConfig
            school = SchoolConfig.get_config(user=request.user, request=request)
            form.instance.school = school # Required for default password logic
            
            teacher = form.save(commit=False)
            teacher.school = school # CRITICAL: Associate with school
            
            # Default to Main School if no branch selected
            if not teacher.branch:
                main_branch = Branch.objects.filter(school=school, name__icontains='Main').first()
                if main_branch:
                    teacher.branch = main_branch
            
            teacher.save()
            form.save_m2m()
            
            # Auto-create user account
            from .user_utils import create_staff_user
            create_staff_user(teacher, school)
            
            return JsonResponse({'success': True, 'teacher_id': teacher.id})
        else:
            # Return validation errors
            return JsonResponse({'error': 'Validation Failed', 'errors': form.errors}, status=400)
            
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@require_http_methods(["GET", "POST"])
def login_view(request):
    """Custom login view with email/username login support"""
    if request.method == 'POST':
        login_field = request.POST.get('username')  # Could be username or email
        password = request.POST.get('password')
        
        # Try to authenticate directly (username)
        user = authenticate(request, username=login_field, password=password)
        
        # If direct auth fails, try email
        if user is None:
            try:
                # Try to find user by email
                user_obj = User.objects.filter(email=login_field).first()
                if user_obj:
                    # Try to authenticate with found username
                    user = authenticate(request, username=user_obj.username, password=password)
            except Exception:
                user = None
        
        if user is not None and user.is_active:
            login(request, user)
            
            # Check if user is a student
            if hasattr(user, 'student_profile'):
                messages.success(request, f'Welcome back, {user.student_profile.get_full_name()}!')
                return redirect('student_detail', pk=user.student_profile.pk)
            
            # Check if user is a teacher
            elif hasattr(user, 'teacher'):
                messages.success(request, f'Welcome back, {user.teacher.get_full_name()}!')
                return redirect('teacher_detail', pk=user.teacher.pk)
            
            # For admin and other staff
            else:
                messages.success(request, f'Welcome back, {user.get_full_name() or user.username}!')
                return redirect('dashboard')
        else:
            messages.error(request, 'Invalid username/email or password')
    
    return render(request, 'schools/login.html')

def logout_view(request):
    logout(request)
    return redirect('/login/')

@login_required
def export_teachers(request):
    """Export teachers list in Excel or PDF format"""
    format_type = request.GET.get('format', 'excel')
    
    # Get filtered teachers
    teachers = Teacher.objects.select_related('user', 'grade', 'employee_ptr').prefetch_related('subjects')
    
    if format_type == 'excel':
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Teachers"
        
        # Add headers
        headers = ['ID', 'Name', 'Email', 'Phone', 'Grade', 'Subjects', 'Status']
        ws.append(headers)
        
        # Add data
        for teacher in teachers:
            ws.append([
                teacher.employee_ptr.id if hasattr(teacher, 'employee_ptr') else '',
                teacher.get_full_name(),
                teacher.user.email if teacher.user else '',
                teacher.phone,
                teacher.grade.name if teacher.grade else '',
                ', '.join(s.name for s in teacher.subjects.all()),
                teacher.get_status_display()
            ])
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=teachers.xlsx'
        wb.save(response)
        return response
        
    elif format_type == 'pdf':
        # Create PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=teachers.pdf'
        
        # Create PDF document
        doc = SimpleDocTemplate(response, pagesize=letter)
        elements = []
        
        # Add title
        styles = getSampleStyleSheet()
        elements.append(Paragraph("Teachers List", styles['Title']))
        elements.append(Spacer(1, 12))
        
        # Create table data
        data = [['ID', 'Name', 'Email', 'Grade', 'Status']]
        for teacher in teachers:
            data.append([
                teacher.employee_ptr.id if hasattr(teacher, 'employee_ptr') else '',
                teacher.get_full_name(),
                teacher.user.email if teacher.user else '',
                teacher.grade.name if teacher.grade else '',
                teacher.get_status_display()
            ])
        
        # Create table
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 12),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        return response
    
    return HttpResponse('Invalid format specified', status=400)

@login_required
def teacher_list(request):
    """List all teachers with summary statistics"""
    # Get all teachers with their related data
    teachers = Teacher.objects.select_related('grade').prefetch_related('subjects')
    
    # Handle location filter
    location_filter = request.GET.get('location', '')
    if location_filter:
        teachers = teachers.filter(location=location_filter)
    
    # Calculate average experience
    avg_experience = teachers.aggregate(Avg('years_of_experience'))['years_of_experience__avg'] or 0
    
    # Count class teachers
    class_teacher_count = teachers.filter(is_class_teacher=True).count()
    
    # Count teachers on leave (using status field)
    on_leave_count = teachers.filter(status='ON_LEAVE').count()
    
    context = {
        'teachers': teachers,
        'avg_experience': avg_experience,
        'class_teacher_count': class_teacher_count,
        'on_leave_count': on_leave_count,
        'location_filter': location_filter,
        'title': 'Teachers'
    }
    return render(request, 'schools/teacher_list.html', context)

def calculate_fee_statistics(term=None, year=None, student=None):
    """Calculate fee statistics for dashboard or specific student"""
    
    # If student is provided, calculate for that student only
    if student:
        from django.db.models import Sum
        from django.db.models.functions import Coalesce
        from django.db.models import Value, DecimalField
        
        total_paid = student.payments.aggregate(
            total=Coalesce(Sum('amount'), Value(0), output_field=DecimalField(max_digits=10, decimal_places=2))
        )['total'] or 0
        
        term_fee = student.term_fees or 0
        balance = term_fee - total_paid
        
        return {
            'total_fees': term_fee,
            'total_collected': total_paid,
            'total_pending': max(0, balance),
            'students_with_balance': 1 if balance > 0 else 0,
            'fully_paid_students': 1 if balance <= 0 and term_fee > 0 else 0
        }
    
    # Get all students with their fee information
    students = Student.objects.all()
    if term:
        students = students.filter(current_term=term)
    
    # Calculate total expected fees and payments for each student using raw SQL
    students_with_fees = students.annotate(
        total_paid=Coalesce(
            Sum('payments__amount'), 
            Value(0), 
            output_field=DecimalField(max_digits=10, decimal_places=2)
        ),
        balance=Case(
            When(
                term_fees__isnull=False,
                then=F('term_fees') - F('total_paid')
            ),
            default=Value(0),
            output_field=DecimalField(max_digits=10, decimal_places=2)
        )
    )
    
    # Calculate summary statistics
    stats = students_with_fees.aggregate(
        total_expected=Coalesce(
            Sum('term_fees'), 
            Value(0), 
            output_field=DecimalField(max_digits=10, decimal_places=2)
        ),
        total_collected=Coalesce(
            Sum('total_paid'), 
            Value(0), 
            output_field=DecimalField(max_digits=10, decimal_places=2)
        ),
        total_pending=Coalesce(
            Sum('balance'), 
            Value(0), 
            output_field=DecimalField(max_digits=10, decimal_places=2)
        )
    )
    
    # Count students with balance using explicit comparison
    students_with_balance = students_with_fees.filter(
        balance__gt=0
    ).count()
    
    # Count fully paid students
    fully_paid_students = students_with_fees.filter(
        balance__lte=0,
        term_fees__gt=0  # Only count students who have fees set
    ).count()
    
    # Get monthly collections
    monthly_data = [0] * 12
    payments = Payment.objects.filter(student__in=students)
    if year:
        payments = payments.filter(date__year=year)
        
    for payment in payments:
        if payment.date:
            month_index = payment.date.month - 1
            monthly_data[month_index] += float(payment.amount or 0)
    
    # Get payment methods distribution
    payment_methods = Payment.objects.filter(
        student__in=students
    ).values('payment_method').annotate(
        count=Count('id'),
        total=Sum('amount')
    ).order_by('payment_method')
    
    formatted_payment_methods = []
    for method in payment_methods:
        formatted_payment_methods.append({
            'method': method['payment_method'],
            'count': method['count'],
            'total': float(method['total'] or 0)
        })

    return {
        'total_expected': stats['total_expected'],
        'total_collected': stats['total_collected'],
        'total_pending': stats['total_pending'],
        'students_with_balance': students_with_balance,
        'fully_paid_students': fully_paid_students,
        'monthly_data': monthly_data,
        'payment_methods': formatted_payment_methods,
        'collection_percentage': round(
            (float(stats['total_collected']) / float(stats['total_expected']) * 100) 
            if stats['total_expected'] > 0 else 0, 1
        )
    }

def get_term_dates(term, year):
    """Helper function to get term date ranges"""
    if term == 1:
        return {
            'start_date': f'{year}-01-01',
            'end_date': f'{year}-04-30'
        }
    elif term == 2:
        return {
            'start_date': f'{year}-05-01',
            'end_date': f'{year}-08-31'
        }
    else:  # term 3
        return {
            'start_date': f'{year}-09-01',
            'end_date': f'{year}-12-31'
        }

@login_required
def dashboard(request):
    """View for main dashboard"""
    from django.db.models import Sum, Count, F, Value, DecimalField, Q
    from django.db.models.functions import Coalesce, TruncMonth
    from django.utils import timezone
    import calendar
    from datetime import datetime, timedelta
    
    today = timezone.now().date()
    
    # Get student statistics - all enrolled students
    total_students = Student.objects.count()
    
    # Get student counts by school location
    main_school_students = Student.objects.filter(location='MAIN').count()
    annex_school_students = Student.objects.filter(location='ANNEX').count()
    
    new_students = Student.objects.filter(
        created_at__month=timezone.now().month,
        created_at__year=timezone.now().year
    ).count()
    
    # Get teacher statistics
    teacher_stats = Teacher.objects.aggregate(
        total=Count('id'),
        active=Count('id', filter=Q(status='ACTIVE'))
    )
    
    # Get today's attendance
    todays_attendance = Attendance.objects.filter(date=today)
    present_today = todays_attendance.filter(status='PRESENT').count()
    absent_today = todays_attendance.filter(status='ABSENT').count()
    late_today = todays_attendance.filter(status='LATE').count()
    
    # Calculate attendance rate
    attendance_rate = (present_today / total_students * 100) if total_students > 0 else 0
    
    # Calculate fee collection statistics
    students_with_fees = Student.objects.annotate(
        total_paid=Coalesce(
            Sum('payments__amount'), 
            Value(0), 
            output_field=DecimalField(max_digits=10, decimal_places=2)
        )
    )
    
    # Calculate total expected fees
    total_fees = students_with_fees.aggregate(
        total=Coalesce(Sum('term_fees'), Value(0), 
        output_field=DecimalField(max_digits=10, decimal_places=2))
    )['total']
    
    # Calculate total collected fees
    total_collected = students_with_fees.aggregate(
        total=Coalesce(Sum('total_paid'), Value(0), 
        output_field=DecimalField(max_digits=10, decimal_places=2))
    )['total']
    
    # Calculate admission fees statistics
    ADMISSION_FEE = Decimal('500.00')
    # Count admission fee payments (payments of exactly 500)
    admission_fees_collected = Payment.objects.filter(
        amount=ADMISSION_FEE
    ).aggregate(
        total=Coalesce(Sum('amount'), Value(0), output_field=DecimalField(max_digits=10, decimal_places=2))
    )['total']
    
    # Expected admission fees (one per student)
    total_admission_fees_expected = total_students * ADMISSION_FEE
    
    # Calculate detailed fee statistics
    fee_stats = {
        'total_expected': total_fees,
        'total_collected': total_collected,
        'balance': total_fees - total_collected,
        'students_with_balance': students_with_fees.filter(
            total_paid__lt=F('term_fees')
        ).count()
    }
    
    # Get monthly collection data for chart (last 6 months)
    chart_labels = []
    chart_data = []
    
    for i in range(5, -1, -1):
        # Calculate start and end of month
        current = today - timedelta(days=today.day - 1) - timedelta(days=30*i)
        month_start = current.replace(day=1)
        month_end = current.replace(day=calendar.monthrange(current.year, current.month)[1])
        
        # Get total collections for this month
        month_total = Payment.objects.filter(
            date__gte=month_start,
            date__lte=month_end
        ).aggregate(
            total=Coalesce(
                Sum('amount'),
                Value('0.00'),
                output_field=DecimalField(max_digits=10, decimal_places=2)
            )
        )['total']
        
        chart_labels.append(current.strftime('%b %Y'))
        chart_data.append(float(month_total))
    
    # Get payment methods distribution
    payment_methods = Payment.objects.values('payment_method').annotate(
        total=Count('id')
    ).order_by('-total')
    
    # Calculate payment methods data for chart
    payment_method_labels = [p['payment_method'] for p in payment_methods]
    payment_method_data = [p['total'] for p in payment_methods]
    
    # Transport Statistics
    vehicles = Vehicle.objects.all()
    routes = Route.objects.filter(is_active=True)
    transport_assignments = StudentTransportAssignment.objects.filter(is_active=True)
    transport_fees = TransportFee.objects.all()
    
    total_vehicles = vehicles.count()
    active_vehicles = vehicles.filter(status='ACTIVE').count()
    total_routes = routes.count()
    total_transport_students = transport_assignments.count()
    total_capacity = vehicles.aggregate(total=Sum('capacity'))['total'] or 0
    used_capacity = sum(v.get_current_students_count() for v in vehicles)
    capacity_percentage = (used_capacity / total_capacity * 100) if total_capacity > 0 else 0
    
    # Transport Revenue
    transport_revenue = transport_fees.filter(status='COMPLETED').aggregate(total=Sum('amount'))['total'] or 0
    transport_pending = transport_fees.filter(status='PENDING').aggregate(total=Sum('amount'))['total'] or 0
    
    # Recent transport fees
    recent_transport_fees = transport_fees.select_related('student', 'route').order_by('-date')[:3]
    
    # Total pending payments (transport only)
    total_pending_payments = transport_pending
    
    context = {
        'today': today,
        'total_students': total_students,
        'main_school_students': main_school_students,
        'annex_school_students': annex_school_students,
        'new_students': new_students,
        'total_teachers': teacher_stats['total'],
        'active_teachers': teacher_stats['active'],
        'present_today': present_today,
        'absent_today': absent_today,
        'late_today': late_today,
        'attendance_rate': round(attendance_rate, 1),
        'total_fees': total_fees,
        'total_collected': total_collected,
        'fees_percent': (total_collected / total_fees) * 100 if total_fees > 0 else 0,
        'admission_fee': ADMISSION_FEE,
        'admission_fees_collected': admission_fees_collected,
        'admission_fees_expected': total_admission_fees_expected,
        'admission_fees_percent': (admission_fees_collected / total_admission_fees_expected * 100) if total_admission_fees_expected > 0 else 0,
        'recent_admissions': Student.objects.select_related('grade').order_by('-created_at')[:3],
        'recent_payments': Payment.objects.select_related(
            'student', 
            'student__grade'
        ).order_by('-date')[:3],
        'fee_stats': fee_stats,
        'expected_term_fees': total_fees,
        'total_payments': total_collected,
        'fees_balance': fee_stats['balance'],
        'chart_labels': chart_labels,
        'chart_data': chart_data,
        'payment_method_labels': payment_method_labels,
        'payment_method_data': payment_method_data,
        # Transport data
        'total_vehicles': total_vehicles,
        'active_vehicles': active_vehicles,
        'total_routes': total_routes,
        'total_transport_students': total_transport_students,
        'capacity_percentage': round(capacity_percentage, 1),
        'transport_revenue': transport_revenue,
        'transport_pending': transport_pending,
        'recent_transport_fees': recent_transport_fees,
        'total_pending_payments': total_pending_payments,
        'current_term': 'Term 1, 2024',  # Update this based on your term system
        'title': 'Dashboard',
        'page_title': 'Dashboard'
    }
    
    return render(request, 'schools/dashboard.html', context)

def calculate_attendance_rate(date):
    """Calculate attendance rate for a specific date"""
    total_students = Student.objects.count()
    if total_students == 0:
        return 0
        
    present_students = Attendance.objects.filter(
        date=date,
        status='PRESENT'
    ).count()
    
    return round((present_students / total_students) * 100, 1)

def get_current_term():
    """Get current term number"""
    current_term = Term.objects.filter(is_current=True).first()
    if current_term:
        # Extract just the number from term name/number
        return current_term.number
    return 1  # Default to term 1 if no current term set

def calculate_term_progress(assessments, current_term):
    """Calculate the term progress based on completed assessments"""
    # Filter assessments for current term
    term_assessments = assessments.filter(term=current_term)
    
    # Define required assessments
    required_assessments = {
        'OPENER': False,
        'MIDTERM': False,
        'ENDTERM': False
    }
    
    # Check which assessments are completed
    for assessment in term_assessments:
        required_assessments[assessment.assessment_type] = True
    
    # Calculate progress percentage
    completed = sum(1 for status in required_assessments.values() if status)
    total = len(required_assessments)
    
    return int((completed / total) * 100) if total > 0 else 0

@login_required
def student_list(request):
    """List all students with summary statistics"""
    # Get all students with related data
    students = Student.objects.select_related('grade').all()
    
    # Handle filters
    filters = {
        'search': request.GET.get('search', ''),
        'grade': request.GET.get('grade', ''),
        'gender': request.GET.get('gender', ''),
        'location': request.GET.get('location', ''),
        'fee_status': request.GET.get('fee_status', '')
    }
    
    # Apply filters
    if filters['search']:
        search_query = filters['search']
        students = students.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(admission_number__icontains=search_query)
        )
    
    if filters['grade']:
        students = students.filter(grade_id=filters['grade'])
    
    if filters['gender']:
        students = students.filter(gender=filters['gender'])
    
    if filters['location']:
        students = students.filter(location=filters['location'])
    
    if filters['fee_status']:
        current_term = get_current_term()
        term_dates = get_term_dates(current_term, timezone.now().year)
        
        students = students.annotate(
            total_paid=Coalesce(
                Sum('payments__amount', 
                    filter=Q(payments__date__range=(term_dates['start_date'], term_dates['end_date']))),
                Value(0, output_field=models.DecimalField())
            ),
            balance=ExpressionWrapper(
                F('term_fees') - F('total_paid'),
                output_field=models.DecimalField(max_digits=10, decimal_places=2)
            )
        )
        
        if filters['fee_status'] == 'paid':
            students = students.filter(balance__lte=0)
        elif filters['fee_status'] == 'partial':
            students = students.filter(balance__gt=0, total_paid__gt=0)
        elif filters['fee_status'] == 'unpaid':
            students = students.filter(total_paid=0)
    
    # Calculate statistics
    total_students = students.count()
    male_count = students.filter(gender='M').count()
    female_count = students.filter(gender='F').count()
    
    # Calculate percentages
    male_percentage = (male_count / total_students * 100) if total_students > 0 else 0
    female_percentage = (female_count / total_students * 100) if total_students > 0 else 0
    
    # Get grade statistics
    total_grades = Grade.objects.count()
    grades = Grade.objects.annotate(student_count=Count('students'))
    grades_with_students = grades.filter(student_count__gt=0).count()
    
    # Calculate grade utilization
    grade_utilization_percentage = (grades_with_students / total_grades * 100) if total_grades > 0 else 0
    
    context = {
        'students': students,
        'total_students': total_students,
        'male_count': male_count,
        'female_count': female_count,
        'male_percentage': male_percentage,
        'female_percentage': female_percentage,
        'total_grades': total_grades,
        'grade_utilization_percentage': grade_utilization_percentage,
        'grades': grades,  # For grade filter dropdown
        'filters': filters,  # Pass filters back to template
        'title': 'Students',
        'page_title': 'Students'
    }
    return render(request, 'schools/student_list.html', context)

@login_required
@can_manage_students
def student_create(request):
    """Create a new student - requires admission fee payment"""
    from decimal import Decimal
    from django.utils import timezone
    import uuid
    import json
    import tempfile
    import os
    
    ADMISSION_FEE = Decimal('500.00')
    
    # Handle payment confirmation step
    if request.method == 'POST' and 'confirm_payment' in request.POST:
        # Get student data from session
        student_data_json = request.session.get('pending_student_data')
        if not student_data_json:
            messages.error(request, 'Student data not found. Please fill the form again.')
            return redirect('student_create')
        
        try:
            student_data = json.loads(student_data_json)
        except:
            messages.error(request, 'Invalid student data. Please fill the form again.')
            return redirect('student_create')
        
        # Get payment details
        payment_method = request.POST.get('payment_method')
        if not payment_method:
            messages.error(request, 'Payment method is required')
            return render(request, 'schools/student_create_payment.html', {
                'admission_fee': ADMISSION_FEE,
                'student_data': student_data,
                'form': temp_form
            })
        
        # Generate reference number
        reference_number = None
        phone_number = None
        transaction_id = None
        
        if payment_method == 'MPESA':
            reference_number = request.POST.get('mpesa_reference_number') or request.POST.get('reference_number')
            phone_number = request.POST.get('phone_number')
            transaction_id = reference_number
            if not reference_number:
                messages.error(request, 'M-Pesa reference number is required')
                temp_form = StudentForm(student_data)
                return render(request, 'schools/student_create_payment.html', {
                    'admission_fee': ADMISSION_FEE,
                    'student_data': student_data,
                    'form': temp_form
                })
        elif payment_method == 'BANK':
            reference_number = request.POST.get('bank_reference_number') or request.POST.get('reference_number')
            transaction_id = reference_number
            if not reference_number:
                messages.error(request, 'Bank reference number is required')
                temp_form = StudentForm(student_data)
                return render(request, 'schools/student_create_payment.html', {
                    'admission_fee': ADMISSION_FEE,
                    'student_data': student_data,
                    'form': temp_form
                })
        elif payment_method == 'CASH':
            reference_number = f'CASH-ADM-{timezone.now().strftime("%Y%m%d%H%M%S")}-{uuid.uuid4().hex[:6].upper()}'
        
        if not reference_number:
            messages.error(request, 'Reference number is required')
            temp_form = StudentForm(student_data)
            return render(request, 'schools/student_create_payment.html', {
                'admission_fee': ADMISSION_FEE,
                'student_data': student_data,
                'form': temp_form
            })
        
        # Recreate form with session data
        from django.core.files.uploadedfile import InMemoryUploadedFile, TemporaryUploadedFile
        from django.core.files.base import ContentFile
        from django.core.files.uploadedfile import SimpleUploadedFile
        
        form_data = student_data.copy()
        
        # Handle photo file if it was uploaded
        photo_path = request.session.get('pending_student_photo_path')
        files_dict = {}
        if photo_path and os.path.exists(photo_path):
            with open(photo_path, 'rb') as f:
                photo_content = f.read()
                photo_file = SimpleUploadedFile(
                    name=os.path.basename(photo_path),
                    content=photo_content,
                    content_type='image/jpeg'
                )
                files_dict['photo'] = photo_file
        
        # Create form with data and files
        if files_dict:
            form = StudentForm(form_data, files_dict)
        else:
            form = StudentForm(form_data)
        
        # Set school on form instance before saving
        from config.models import SchoolConfig
        config = SchoolConfig.get_config(user=request.user, request=request)
        form.instance.school = config
        
        if form.is_valid():
            # Save student
            student = form.save()
            
            # Clean up temporary photo file
            if photo_path and os.path.exists(photo_path):
                try:
                    os.remove(photo_path)
                except:
                    pass
            
            # Create admission fee payment
            current_term = student.current_term
            payment = Payment.objects.create(
                student=student,
                amount=ADMISSION_FEE,
                payment_method=payment_method,
                reference_number=reference_number,
                phone_number=phone_number,
                transaction_id=transaction_id,
                term=current_term,
                status='COMPLETED'
            )
            
            # Send receipt
            try:
                from .utils import send_payment_receipt
                send_payment_receipt(payment)
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error sending receipt: {str(e)}")
            
            # Clear session data
            if 'pending_student_data' in request.session:
                del request.session['pending_student_data']
            if 'pending_student_photo_path' in request.session:
                del request.session['pending_student_photo_path']
            
            if hasattr(student, 'temp_credentials'):
                messages.success(
                    request,
                    f'Student registered successfully! Admission fee of KES {ADMISSION_FEE:,.2f} has been recorded. IMPORTANT - Login Credentials:\n'
                    f'Username: {student.temp_credentials["username"]}\n'
                    f'Password: {student.temp_credentials["password"]}\n'
                    f'Please save these credentials and share them with the student/parent.'
                )
            else:
                messages.success(request, f'Student registered successfully! Admission fee of KES {ADMISSION_FEE:,.2f} has been recorded.')
            
            return redirect('student_detail', pk=student.pk)
        else:
            messages.error(request, 'Error saving student. Please try again.')
            return redirect('student_create')
    
    # Handle initial form submission
    if request.method == 'POST':
        form = StudentForm(request.POST, request.FILES)
        from config.models import SchoolConfig
        config = SchoolConfig.get_config(user=request.user, request=request)
        form.instance.school = config
        
        if form.is_valid():
            # Store form data in session as JSON
            student_data = {}
            for key, value in request.POST.items():
                if key != 'csrfmiddlewaretoken':
                    student_data[key] = value
            
            # Ensure generated admission number is persisted in session
            if form.instance.admission_number:
                student_data['admission_number'] = form.instance.admission_number
            
            request.session['pending_student_data'] = json.dumps(student_data)
            
            # Handle photo file upload - save temporarily
            if 'photo' in request.FILES:
                photo_file = request.FILES['photo']
                # Save to temporary file
                temp_dir = tempfile.gettempdir()
                temp_filename = f'student_photo_{uuid.uuid4().hex}.jpg'
                temp_path = os.path.join(temp_dir, temp_filename)
                
                with open(temp_path, 'wb+') as destination:
                    for chunk in photo_file.chunks():
                        destination.write(chunk)
                
                request.session['pending_student_photo_path'] = temp_path
            
            # Recreate form to pass to template (for grade display)
            temp_form = StudentForm(student_data)
            
            # Show payment confirmation page
            return render(request, 'schools/student_create_payment.html', {
                'admission_fee': ADMISSION_FEE,
                'student_data': student_data,
                'form': temp_form  # Pass form to access grade field
            })
    else:
        form = StudentForm()
        # Clear any pending data
        if 'pending_student_data' in request.session:
            del request.session['pending_student_data']
        if 'pending_student_photo_path' in request.session:
            photo_path = request.session['pending_student_photo_path']
            if os.path.exists(photo_path):
                try:
                    os.remove(photo_path)
                except:
                    pass
            del request.session['pending_student_photo_path']

    grades = Grade.objects.filter(is_active=True).order_by('name')
    
    context = {
        'form': form,
        'title': 'Register New Student',
        'grades': grades,
        'page_title': 'Register New Student',
        'admission_fee': ADMISSION_FEE
    }
    return render(request, 'schools/student_create.html', context)

@login_required
def student_edit(request, pk):
    student = get_object_or_404(Student, pk=pk)
    
    if request.method == 'POST':
        form = StudentForm(request.POST, request.FILES, instance=student)
        
        
        if form.is_valid():
            student = form.save()
            
            # Update associated user information
            if student.user:
                student.user.first_name = form.cleaned_data['first_name']
                student.user.last_name = form.cleaned_data['last_name']
                student.user.email = form.cleaned_data.get('parent_email', '')
                student.user.save()
            
            success_message = f'Student {student.get_full_name()} was updated successfully!'
            messages.success(request, success_message)
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'status': 'success',
                    'message': success_message,
                    'redirect_url': reverse('student_detail', kwargs={'pk': student.pk})
                })
            
            return redirect('student_detail', pk=pk)
        else:
            messages.error(request, 'Please correct the errors below.')
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'status': 'error',
                    'message': 'Please correct the errors below.',
                    'errors': {field: errors[0] for field, errors in form.errors.items()}
                }, status=400)
    else:
        form = StudentForm(instance=student)
    
    context = {
        'form': form,
        'student': student,
        'title': f'Edit {student.get_full_name()}',
        'submit_text': 'Update Student'
    }
    
    return render(request, 'schools/student_form.html', context)

@login_required
def student_detail(request, pk):
    student = get_object_or_404(Student.objects.select_related('grade', 'grade__class_teacher', 'grade__class_teacher__user'), pk=pk)
    
    # Get student's current term (automatically updated when terms advance)
    student_current_term = student.current_term or 1
    current_year = timezone.now().year
    
    # Get all assessments for the student with related data
    assessments = Assessment.objects.filter(student=student).select_related('student').prefetch_related(
        'results',
        'results__subject'
    ).order_by('-date')
    
    # Get current term assessments (for academic overview)
    current_term_assessments = assessments.filter(term=student_current_term)
    
    # Group assessments by type (all terms)
    weekly_assessments = assessments.filter(assessment_type='weekly')
    opener_assessments = assessments.filter(assessment_type='opener')
    midterm_assessments = assessments.filter(assessment_type='mid-term')
    endterm_assessments = assessments.filter(assessment_type='end-term')
    
    # Calculate term progress using student's current term
    term_progress = calculate_term_progress(assessments, student_current_term)
    
    # Get attendance statistics
    attendance_stats = student.get_attendance_stats()
    attendance_rate = attendance_stats.get('present_percentage', 0) if attendance_stats else 0
    
    # Get fee statistics for this student using student's current term
    fee_stats = calculate_fee_statistics(term=student_current_term, year=current_year, student=student)
    
    # Get performance trends for current term
    performance_trend = calculate_trend_direction(current_term_assessments)
    trend_description = get_trend_description(current_term_assessments)
    
    # Get best performing subject for current term
    best_subject = get_best_performing_subject(current_term_assessments)
    
    # Get areas needing improvement for current term
    improvement_areas = get_improvement_areas(current_term_assessments)
    
    # Get academic overview for current term
    academic_overview = {
        'most_common_level': calculate_most_common_level(current_term_assessments),
        'best_subject': best_subject,
        'improvement_areas': improvement_areas,  # always a list
        'trend': performance_trend
    }
    
    # Get terms for filter
    terms = [
        {'value': 'all', 'label': 'All Terms'},
        {'value': '1', 'label': 'Term 1'},
        {'value': '2', 'label': 'Term 2'},
        {'value': '3', 'label': 'Term 3'}
    ]
    
    # Get selected filters
    selected_term = request.GET.get('term', 'all')
    selected_type = request.GET.get('type', 'all')
    
    # Filter assessments based on selected filters
    filtered_assessments = assessments
    if selected_term != 'all':
        filtered_assessments = filtered_assessments.filter(term=selected_term)
    if selected_type != 'all':
        filtered_assessments = filtered_assessments.filter(assessment_type=selected_type)
    
    # Get transport assignments
    transport_assignments = StudentTransportAssignment.objects.filter(
        student=student
    ).select_related('route', 'vehicle').order_by('-start_date')
    
    active_assignment = transport_assignments.filter(is_active=True).first()
    
    # Get transport fees
    transport_fees = TransportFee.objects.filter(
        student=student
    ).select_related('route').order_by('-date')
    
    # Calculate transport fee statistics
    from django.db.models import Sum
    total_transport_paid = transport_fees.filter(status='COMPLETED').aggregate(
        total=Sum('amount')
    )['total'] or 0
    
    # Ensure transport_fees is a queryset (not None)
    if transport_fees is None:
        transport_fees = TransportFee.objects.none()
    
    # Get fees by term
    transport_fees_by_term = {}
    for term_num in [1, 2, 3]:
        term_fees = transport_fees.filter(term=term_num, status='COMPLETED').aggregate(
            total=Sum('amount')
        )['total'] or 0
        if active_assignment:
            term_fee_due = active_assignment.route.fee_per_term
            balance = float(term_fee_due) - float(term_fees)
        else:
            term_fee_due = 0
            balance = 0
        transport_fees_by_term[term_num] = {
            'paid': float(term_fees),
            'due': float(term_fee_due) if active_assignment else 0,
            'balance': balance
        }
    
    # Get attendance statistics
    attendance_stats = student.get_attendance_stats()
    
    # Get meal statistics
    from .models import StudentMealPayment
    from django.db.models import Sum
    meal_payments = StudentMealPayment.objects.filter(student=student, status='COMPLETED', is_active=True)
    meal_stats = {
        'active_payments': meal_payments.filter(days_remaining__gt=0).count(),
        'total_balance': meal_payments.aggregate(total=Sum('balance'))['total'] or 0,
        'total_paid': meal_payments.aggregate(total=Sum('amount'))['total'] or 0,
    }
    
    context = {
        'student': student,
        'assessments': assessments,  # Add this for the assessment history template
        'weekly_assessments': weekly_assessments,
        'opener_assessments': opener_assessments,
        'midterm_assessments': midterm_assessments,
        'endterm_assessments': endterm_assessments,
        'filtered_assessments': filtered_assessments,
        'term_progress': term_progress,
        'attendance_rate': attendance_rate,
        'fee_stats': fee_stats,
        'performance_trend': performance_trend,
        'trend_description': trend_description,
        'best_subject': best_subject,
        'transport_assignments': transport_assignments,
        'active_assignment': active_assignment,
        'transport_fees': transport_fees,
        'total_transport_paid': total_transport_paid,
        'transport_fees_by_term': transport_fees_by_term,
        'attendance_stats': attendance_stats,
        'meal_stats': meal_stats,
        'improvement_areas': improvement_areas,  # always a list
        'current_term': student_current_term,
        'academic_overview': academic_overview,
        'terms': terms,
        'selected_term': selected_term,
        'selected_type': selected_type,
        'is_class_teacher': hasattr(request.user, 'teacher') and request.user.teacher.is_class_teacher,
        'student_grade': student.grade
    }
    
    return render(request, 'schools/student_detail.html', context)

@login_required
def student_academic(request, pk):
    """Student Academic Page - Detailed academic information"""
    from datetime import datetime
    student = get_object_or_404(Student.objects.select_related('grade', 'grade__class_teacher', 'grade__class_teacher__user'), pk=pk)
    
    # Get student's current term (automatically updated when terms advance)
    student_current_term = student.current_term or 1
    
    # Get all assessments for the student with related data
    assessments = Assessment.objects.filter(student=student).select_related('student').prefetch_related(
        'results',
        'results__subject'
    ).order_by('-date')
    
    # Get current term assessments (for academic overview)
    current_term_assessments = assessments.filter(term=student_current_term)
    
    # Group assessments by type (all terms)
    weekly_assessments = assessments.filter(assessment_type='weekly')
    opener_assessments = assessments.filter(assessment_type='opener')
    midterm_assessments = assessments.filter(assessment_type='mid-term')
    endterm_assessments = assessments.filter(assessment_type='end-term')
    
    # Calculate term progress using student's current term
    term_progress = calculate_term_progress(assessments, student_current_term)
    
    # Get performance trends for current term
    performance_trend = calculate_trend_direction(current_term_assessments)
    trend_description = get_trend_description(current_term_assessments)
    
    # Get best performing subject for current term
    best_subject = get_best_performing_subject(current_term_assessments)
    
    # Get areas needing improvement for current term
    improvement_areas = get_improvement_areas(current_term_assessments)
    
    # Get academic overview for current term
    academic_overview = {
        'most_common_level': calculate_most_common_level(current_term_assessments),
        'best_subject': best_subject,
        'improvement_areas': improvement_areas,
        'trend': performance_trend
    }
    
    # Get terms for filter
    terms = [
        {'value': 'all', 'label': 'All Terms'},
        {'value': '1', 'label': 'Term 1'},
        {'value': '2', 'label': 'Term 2'},
        {'value': '3', 'label': 'Term 3'}
    ]
    
    # Get selected filters
    selected_term = request.GET.get('term', 'all')
    selected_type = request.GET.get('type', 'all')
    
    # Filter assessments based on selected filters
    filtered_assessments = assessments
    if selected_term != 'all':
        filtered_assessments = filtered_assessments.filter(term=selected_term)
    if selected_type != 'all':
        filtered_assessments = filtered_assessments.filter(assessment_type=selected_type)
    
    context = {
        'student': student,
        'assessments': assessments,
        'weekly_assessments': weekly_assessments,
        'opener_assessments': opener_assessments,
        'midterm_assessments': midterm_assessments,
        'endterm_assessments': endterm_assessments,
        'filtered_assessments': filtered_assessments,
        'term_progress': term_progress,
        'performance_trend': performance_trend,
        'trend_description': trend_description,
        'best_subject': best_subject,
        'improvement_areas': improvement_areas,
        'current_term': student_current_term,
        'academic_overview': academic_overview,
        'terms': terms,
        'selected_term': selected_term,
        'selected_type': selected_type,
        'is_class_teacher': hasattr(request.user, 'teacher') and request.user.teacher.is_class_teacher,
        'student_grade': student.grade
    }
    
    return render(request, 'schools/student_academic.html', context)

@login_required
def student_finance(request, pk):
    """Student Finance Page - Detailed financial information with analytics"""
    from django.db.models import Sum, Count, Avg
    from django.utils import timezone
    from datetime import datetime, timedelta
    import calendar
    
    student = get_object_or_404(Student.objects.select_related('grade'), pk=pk)
    
    # Get all payments
    payments = Payment.objects.filter(student=student).order_by('-date')
    
    # Get transport assignments and fees
    transport_assignments = StudentTransportAssignment.objects.filter(
        student=student
    ).select_related('route', 'vehicle').order_by('-start_date')
    
    active_assignment = transport_assignments.filter(is_active=True).first()
    
    # Get transport fees
    transport_fees = TransportFee.objects.filter(
        student=student
    ).select_related('route').order_by('-date')
    
    # Calculate transport fee statistics
    total_transport_paid = transport_fees.filter(status='COMPLETED').aggregate(
        total=Sum('amount')
    )['total'] or 0
    
    # Financial Analytics
    today = timezone.now().date()
    
    # Monthly payment trends (last 12 months)
    monthly_data = []
    monthly_labels = []
    for i in range(11, -1, -1):
        month_date = today - timedelta(days=30*i)
        month_start = month_date.replace(day=1)
        if month_date.month == 12:
            month_end = month_date.replace(year=month_date.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            month_end = month_date.replace(month=month_date.month + 1, day=1) - timedelta(days=1)
        
        # Convert to datetime for filtering DateTimeField
        month_start_dt = timezone.make_aware(datetime.combine(month_start, datetime.min.time()))
        month_end_dt = timezone.make_aware(datetime.combine(month_end, datetime.max.time()))
        
        month_total = payments.filter(
            date__gte=month_start_dt,
            date__lte=month_end_dt
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        monthly_data.append(float(month_total))
        monthly_labels.append(month_start.strftime('%b %Y'))
    
    # Payment method breakdown
    payment_method_stats = payments.values('payment_method').annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')
    
    payment_method_labels = [p['payment_method'] for p in payment_method_stats]
    payment_method_amounts = [float(p['total'] or 0) for p in payment_method_stats]
    payment_method_counts = [p['count'] for p in payment_method_stats]
    
    # Combine method data for template
    payment_method_data = [
        {
            'method': p['payment_method'],
            'amount': float(p['total'] or 0),
            'count': p['count']
        }
        for p in payment_method_stats
    ]
    
    # Payment status breakdown
    payment_status_stats = payments.values('status').annotate(
        total=Sum('amount'),
        count=Count('id')
    )
    status_data = {s['status']: {'total': float(s['total'] or 0), 'count': s['count']} for s in payment_status_stats}
    
    # Recent payments (last 30 days)
    recent_payments = payments.filter(
        date__date__gte=today - timedelta(days=30)
    )
    recent_total = recent_payments.aggregate(total=Sum('amount'))['total'] or 0
    
    # Average payment amount
    avg_payment = payments.aggregate(avg=Avg('amount'))['avg'] or 0
    
    # Largest payment
    largest_payment = payments.aggregate(max=Sum('amount'))['max'] or 0
    if payments.exists():
        largest_payment = float(payments.order_by('-amount').first().amount)
    
    # Payment frequency (payments per month average)
    if payments.exists():
        first_payment_date = payments.order_by('date').first().date.date()
        months_diff = (today.year - first_payment_date.year) * 12 + (today.month - first_payment_date.month)
        if months_diff > 0:
            payment_frequency = payments.count() / months_diff
        else:
            payment_frequency = payments.count()
    else:
        payment_frequency = 0
    
    # Year-to-date totals
    year_start = today.replace(month=1, day=1)
    ytd_payments = payments.filter(date__date__gte=year_start)
    ytd_total = ytd_payments.aggregate(total=Sum('amount'))['total'] or 0
    
    # Last payment date
    last_payment = payments.first()
    last_payment_date = last_payment.date.date() if last_payment else None
    days_since_last = (today - last_payment_date).days if last_payment_date else None
    
    # Financial health score (0-100)
    total_fees = student.get_total_fees()
    total_paid = student.get_total_paid()
    balance = student.get_total_balance()
    
    if total_fees > 0:
        payment_percentage = (total_paid / total_fees) * 100
        # Health score based on payment percentage and recency
        health_score = min(100, payment_percentage)
        if days_since_last and days_since_last > 90:
            health_score = max(0, health_score - 10)
    else:
        payment_percentage = 0
        health_score = 100
    
    import json
    
    context = {
        'student': student,
        'payments': payments,
        'transport_assignments': transport_assignments,
        'active_assignment': active_assignment,
        'transport_fees': transport_fees,
        'total_transport_paid': total_transport_paid,
        # Analytics (JSON serialized for JavaScript)
        'monthly_data': json.dumps(monthly_data),
        'monthly_labels': json.dumps(monthly_labels),
        'payment_method_labels': json.dumps(payment_method_labels),
        'payment_method_amounts': json.dumps(payment_method_amounts),
        'payment_method_counts': payment_method_counts,
        'payment_method_data': payment_method_data,
        'status_data': status_data,
        'recent_total': float(recent_total),
        'avg_payment': float(avg_payment),
        'largest_payment': largest_payment,
        'payment_frequency': round(payment_frequency, 1),
        'ytd_total': float(ytd_total),
        'last_payment_date': last_payment_date,
        'days_since_last': days_since_last,
        'payment_percentage': round(payment_percentage, 1),
        'health_score': round(health_score, 1),
        'total_fees': float(total_fees),
        'total_paid': float(total_paid),
        'balance': float(balance),
    }
    
    return render(request, 'schools/student_finance.html', context)

@login_required
def student_attendance(request, pk):
    """Student Attendance Page - Detailed attendance information"""
    from datetime import datetime, timedelta
    student = get_object_or_404(Student.objects.select_related('grade'), pk=pk)
    
    # Get attendance statistics
    attendance_stats = student.get_attendance_stats()
    
    # Get attendance records
    attendance_records = Attendance.objects.filter(student=student).order_by('-date')
    
    # Get current month for calendar
    current_month = request.GET.get('month', timezone.now().month)
    current_year = request.GET.get('year', timezone.now().year)
    
    # Get months for filter
    months = [
        {'value': str(i), 'label': datetime(2000, i, 1).strftime('%B')}
        for i in range(1, 13)
    ]
    
    # Get today's date
    today = timezone.now().date()
    
    context = {
        'student': student,
        'attendance_stats': attendance_stats,
        'attendance_records': attendance_records,
        'current_month': int(current_month),
        'current_year': int(current_year),
        'months': months,
        'today': today,
    }
    
    return render(request, 'schools/student_attendance.html', context)

@login_required
def student_transport(request, pk):
    """Student Transport Page - Detailed transport information"""
    from django.db.models import Sum
    student = get_object_or_404(Student.objects.select_related('grade'), pk=pk)
    
    # Get transport assignments
    transport_assignments = StudentTransportAssignment.objects.filter(
        student=student
    ).select_related('route', 'vehicle').order_by('-start_date')
    
    active_assignment = transport_assignments.filter(is_active=True).first()
    
    # Get transport fees
    transport_fees = TransportFee.objects.filter(
        student=student
    ).select_related('route').order_by('-date')
    
    # Calculate transport fee statistics
    total_transport_paid = transport_fees.filter(status='COMPLETED').aggregate(
        total=Sum('amount')
    )['total'] or 0
    
    # Get fees by term
    transport_fees_by_term = {}
    for term_num in [1, 2, 3]:
        term_fees = transport_fees.filter(term=term_num, status='COMPLETED').aggregate(
            total=Sum('amount')
        )['total'] or 0
        if active_assignment:
            term_fee_due = active_assignment.route.fee_per_term
            balance = float(term_fee_due) - float(term_fees)
        else:
            term_fee_due = 0
            balance = 0
        transport_fees_by_term[term_num] = {
            'paid': float(term_fees),
            'due': float(term_fee_due) if active_assignment else 0,
            'balance': balance
        }
    
    context = {
        'student': student,
        'transport_assignments': transport_assignments,
        'active_assignment': active_assignment,
        'transport_fees': transport_fees,
        'total_transport_paid': total_transport_paid,
        'transport_fees_by_term': transport_fees_by_term,
    }
    
    return render(request, 'schools/student_transport.html', context)

@login_required
def student_meals(request, pk):
    """Student Meals Page - Detailed meal payment and consumption information"""
    from django.db.models import Sum, Count, Q
    from django.utils import timezone
    from datetime import datetime, timedelta
    from decimal import Decimal
    
    student = get_object_or_404(Student.objects.select_related('grade'), pk=pk)
    
    # Get all meal payments for this student
    meal_payments = StudentMealPayment.objects.filter(
        student=student
    ).order_by('-payment_date', '-created_at')
    
    # Get active meal payments
    active_payments = meal_payments.filter(
        status='COMPLETED',
        is_active=True,
        days_remaining__gt=0
    )
    
    # Calculate statistics
    total_paid = meal_payments.filter(status='COMPLETED').aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0.00')
    
    total_balance = active_payments.aggregate(
        total=Sum('balance')
    )['total'] or Decimal('0.00')
    
    total_days_remaining = active_payments.aggregate(
        total=Sum('days_remaining')
    )['total'] or 0
    
    total_days_consumed = meal_payments.filter(status='COMPLETED').aggregate(
        total=Sum('days_consumed')
    )['total'] or 0
    
    # Get meal consumption history
    from .models import MealConsumption
    meal_consumptions = MealConsumption.objects.filter(
        student=student
    ).select_related('meal_payment', 'served_by').order_by('-consumption_date', '-created_at')
    
    # Get consumption statistics by meal type
    consumption_by_meal_raw = meal_consumptions.values('meal_type').annotate(
        count=Count('id')
    ).order_by('meal_type')
    
    # Add display names to consumption statistics
    meal_type_dict = dict(StudentMealPayment.MEAL_TYPES)
    consumption_by_meal = []
    for stat in consumption_by_meal_raw:
        consumption_by_meal.append({
            'meal_type': stat['meal_type'],
            'meal_type_display': meal_type_dict.get(stat['meal_type'], stat['meal_type']),
            'count': stat['count']
        })
    
    # Get recent consumption (last 30 days)
    thirty_days_ago = timezone.now().date() - timedelta(days=30)
    recent_consumptions = meal_consumptions.filter(consumption_date__gte=thirty_days_ago)
    
    # Group payments by meal type
    payments_by_meal_type = {}
    for meal_type_code, meal_type_name in StudentMealPayment.MEAL_TYPES:
        type_payments = meal_payments.filter(meal_type=meal_type_code, status='COMPLETED')
        payments_by_meal_type[meal_type_code] = {
            'name': meal_type_name,
            'payments': type_payments,
            'total_paid': type_payments.aggregate(total=Sum('amount'))['total'] or Decimal('0.00'),
            'active_count': type_payments.filter(is_active=True, days_remaining__gt=0).count(),
            'total_balance': type_payments.filter(is_active=True).aggregate(total=Sum('balance'))['total'] or Decimal('0.00'),
        }
    
    context = {
        'student': student,
        'meal_payments': meal_payments,
        'active_payments': active_payments,
        'total_paid': total_paid,
        'total_balance': total_balance,
        'total_days_remaining': total_days_remaining,
        'total_days_consumed': total_days_consumed,
        'meal_consumptions': meal_consumptions[:50],  # Limit to recent 50
        'recent_consumptions': recent_consumptions,
        'consumption_by_meal': consumption_by_meal,
        'payments_by_meal_type': payments_by_meal_type,
        'title': f'{student.get_full_name()} - Meal Payments',
        'page_title': 'Student Meal Payments & Consumption'
    }
    
    return render(request, 'schools/student_meals.html', context)

def calculate_most_common_level(assessments):
    """Calculate the most common performance level across all assessments"""
    if not assessments:
        return '1'
    
    all_levels = []
    for assessment in assessments:
        all_levels.extend([result.performance_level for result in assessment.results.all()])
    
    if not all_levels:
        return '1'
        
    # Count occurrences of each level
    level_counts = Counter(all_levels)
    # Return the most common level
    return level_counts.most_common(1)[0][0]

def get_best_performing_subject(assessments):
    """Get the subject with the highest consistent performance"""
    if not assessments:
        return "No Data"
    
    subject_levels = defaultdict(list)
    for assessment in assessments:
        for result in assessment.results.all():
            subject_levels[result.subject.name].append(int(result.performance_level))
    
    # Calculate average level for each subject
    subject_averages = {
        subject: sum(levels)/len(levels) 
        for subject, levels in subject_levels.items()
    }
    
    if not subject_averages:
        return "No Data"
        
    # Return subject with highest average
    return max(subject_averages.items(), key=lambda x: x[1])[0]

def get_improvement_areas(assessments):
    """Get subjects needing improvement (consistently level 2 or below)"""
    if not assessments:
        return []
    from collections import defaultdict
    subject_levels = defaultdict(list)
    for assessment in assessments:
        for result in assessment.results.all():
            subject_levels[result.subject.name].append(int(result.performance_level))
    # Find subjects with average level ≤ 2
    improvement_needed = [
        subject for subject, levels in subject_levels.items()
        if sum(levels)/len(levels) <= 2
    ]
    return improvement_needed

def calculate_trend_value(assessments):
    """Calculate the percentage change in performance between recent assessments"""
    if len(assessments) < 2:
        return 0
    
    recent = assessments[0].get_average_performance()
    previous = assessments[1].get_average_performance()
    
    if previous == 0:  # Avoid division by zero
        return 0
        
    return ((recent - previous) / previous) * 100

def calculate_trend_direction(assessments):
    """Calculate if the performance trend is positive or negative"""
    if len(assessments) < 2:
        return 0
    recent = assessments[0].get_average_performance()
    previous = assessments[1].get_average_performance()
    return 1 if recent > previous else -1 if recent < previous else 0

def get_trend_description(assessments):
    """Get a description of the performance trend"""
    if len(assessments) < 2:
        return "No trend data"
    
    trend = calculate_trend_value(assessments)
    if trend > 0:
        return f"Up {abs(trend):.1f}% from previous"
    elif trend < 0:
        return f"Down {abs(trend):.1f}% from previous"
    return "No change"

@login_required
def student_delete(request, pk):
    """Delete a student"""
    student = get_object_or_404(Student, pk=pk)
    
    if request.method == 'POST':
        student.delete()
        messages.success(request, 'Student deleted successfully')
        return redirect('student_list')
    
    context = {
        'student': student,
        'title': f'Delete {student.get_full_name()}',
        'page_title': f'Delete {student.get_full_name()}'
    }
    
    return render(request, 'schools/student_confirm_delete.html', context)

@login_required
def export_students(request):
    """Export filtered students data to Excel with proper formatting"""
    # Get filter parameters
    grade_id = request.GET.get('grade')
    gender = request.GET.get('gender')
    fee_status = request.GET.get('fee_status')
    search_query = request.GET.get('search')

    # Start with all students
    students = Student.objects.select_related('grade').order_by('grade__name', 'admission_number')

    # Apply filters
    if grade_id:
        students = students.filter(grade_id=grade_id)
    if gender:
        students = students.filter(gender=gender)
    if search_query:
        students = students.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(admission_number__icontains=search_query)
        )
    if fee_status:
        if fee_status == 'paid':
            students = students.filter(payments__amount__gte=F('grade__term_fees'))
        elif fee_status == 'partial':
            students = students.filter(
                payments__amount__gt=0,
                payments__amount__lt=F('grade__term_fees')
            )
        elif fee_status == 'unpaid':
            students = students.filter(
                Q(payments__amount__isnull=True) |
                Q(payments__amount=0)
            )

    # Limit export size to prevent memory issues and process kills
    MAX_EXPORT_SIZE = 1000
    student_count = students.count()
    
    if student_count > MAX_EXPORT_SIZE:
        messages.error(
            request, 
            f'Export limited to {MAX_EXPORT_SIZE} students. You have {student_count} students matching your filters. '
            f'Please use more specific filters to reduce the number.'
        )
        return redirect('student_list')
    
    # Limit queryset
    students = students[:MAX_EXPORT_SIZE]


    # Create workbook and active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Students List"

    # Add title and date
    ws['A1'] = "STUDENTS LIST"
    ws['A2'] = f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M')}"

    # Add filter information if any filters are applied
    filter_info = []
    if grade_id:
        grade = Grade.objects.get(id=grade_id)
        filter_info.append(f"Grade: {grade.name}")
    if gender:
        filter_info.append(f"Gender: {'Male' if gender == 'M' else 'Female'}")
    if fee_status:
        filter_info.append(f"Fee Status: {fee_status.title()}")
    if search_query:
        filter_info.append(f"Search: {search_query}")
    
    if filter_info:
        ws['A3'] = "Filters: " + ", ".join(filter_info)

    # Style header cells
    header_cells = [ws['A1'], ws['A2']]
    if filter_info:
        header_cells.append(ws['A3'])
    
    for cell in header_cells:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='left')

    # Add headers starting at row 5 to leave space for title and filters
    current_row = 5
    headers = ['Admission Number', 'Student Name', 'Grade', 'Gender', 'Parent Contact', 'Term', 'Fee Balance']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    # Add data
    current_row += 1
    for student in students:
        row_data = [
            student.admission_number,
            student.get_full_name(),
            student.grade.name if student.grade else '',
            student.get_gender_display(),
            student.parent_phone,
            f"Term {student.current_term}",
            student.get_balance()
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = value
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        current_row += 1

    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=students_list.xlsx'
    wb.save(response)
    return response

@csrf_exempt
@require_http_methods(["POST"])
@login_required
def api_students_bulk_import(request):
    """API endpoint for bulk importing students from Excel"""
    try:
        from openpyxl import load_workbook
        from config.models import SchoolConfig
        from django.contrib.auth.models import User, Group
        from django.db import transaction
        import uuid
        from datetime import datetime, date
        
        if 'file' not in request.FILES:
            return JsonResponse({'error': 'No file uploaded'}, status=400)
            
        file = request.FILES['file']
        config = SchoolConfig.get_config(user=request.user, request=request)
        
        wb = load_workbook(file, data_only=True)
        ws = wb.active
        
        # Define header mapping
        headers = {}
        for idx, cell in enumerate(ws[1]):
            if cell.value:
                headers[cell.value.strip().lower()] = idx
        
        # Expected header names (lowercase)
        column_definitions = {
            'registration_no': ['registration no.', 'admission number', 'adm no', 'adm', 'reg no', 'admission no', 'stno'],
            'first_name': ['first name', 'given name', 'fname', 'name', 'full name', 'student name'],
            'last_name': ['last name', 'surname', 'family name', 'lname'],
            'gender': ['gender (m/f)', 'gender', 'sex', 'm/f'],
            'grade': ['grade', 'class', 'form', 'level', 'current class', 'student class'],
            'date_of_birth': ['date of birth (yyyy-mm-dd)', 'date of birth', 'dob', 'birth date']
        }
        
        # Find critical columns
        col_map = {}
        for key, aliases in column_definitions.items():
            for alias in aliases:
                if alias in headers:
                    col_map[key] = headers[alias]
                    break
        
        # Only First Name is absolutely critical for a profile to exist
        if 'first_name' not in col_map:
            return JsonResponse({
                'error': f'Missing Name column. Please ensure your Excel has a "Name" or "First Name" column. Found: {list(headers.keys())}'
            }, status=400)

        results = {
            'total': 0,
            'imported': 0,
            'skipped': 0,
            'errors': []
        }

        # Cache grades for efficiency
        grades_map = {g.name.strip().lower(): g for g in Grade.objects.filter(school=config)}
        students_group, _ = Group.objects.get_or_create(name='Students')

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        results['total'] = len(rows)
        
        # Limit import size to prevent memory issues and process kills
        MAX_IMPORT_SIZE = 500
        if len(rows) > MAX_IMPORT_SIZE:
            return JsonResponse({
                'error': f'Import limited to {MAX_IMPORT_SIZE} students per file. Your file has {len(rows)} rows. '
                        f'Please split your file into smaller batches.',
                'success': False
            }, status=400)

        seen_admission_numbers = set()
        for row_idx, row in enumerate(rows):

            try:
                with transaction.atomic():
                    # Get values safely using mapping
                    def get_val(key):
                        idx = col_map.get(key)
                        if idx is not None and idx < len(row):
                            return row[idx]
                        return None

                    # Basic Info
                    raw_name = get_val('first_name')
                    if not raw_name: continue # Skip completely empty rows
                    
                    name_parts = str(raw_name).strip().split(' ', 1)
                    first_name = name_parts[0]
                    last_name = name_parts[1] if len(name_parts) > 1 else ""
                    
                    # If specific last name column exists, it overrides
                    specific_last_name = get_val('last_name')
                    if specific_last_name:
                        last_name = str(specific_last_name).strip()
                    
                    # Admission Number
                    admission_number = str(get_val('registration_no')).strip() if get_val('registration_no') else None
                    if not admission_number:
                        admission_number = config.generate_admission_number()
                    
                    # Check for duplicates (in DB or this file)
                    if admission_number in seen_admission_numbers:
                        results['skipped'] += 1
                        results['errors'].append(f"Row {row_idx+2}: Duplicate Admission Number #{admission_number} found in Excel.")
                        continue
                    
                    if Student.objects.filter(school=config, admission_number=admission_number).exists():
                        results['skipped'] += 1
                        results['errors'].append(f"Row {row_idx+2}: Admission Number #{admission_number} already exists in system.")
                        continue
                    
                    seen_admission_numbers.add(admission_number)

                    # Gender
                    gender_raw = str(get_val('gender')).strip().upper() if get_val('gender') else 'M'
                    gender = 'M'
                    if gender_raw.startswith('F'): gender = 'F'
                    elif gender_raw.startswith('O'): gender = 'O'
                    
                    # Grade (Optional now)
                    grade_obj = None
                    grade_raw = get_val('grade')
                    if grade_raw:
                        grade_name = str(grade_raw).strip().lower()
                        grade_obj = grades_map.get(grade_name)
                        
                        if not grade_obj:
                            # If they provided a grade but we don't recognize it, log it but keep going if possible?
                            # No, better to log error for that row.
                            results['skipped'] += 1
                            results['errors'].append(f"Row {row_idx+2}: Class '{grade_raw}' not found.")
                            continue

                    # Date of Birth
                    dob = date(2015, 1, 1) # Sensible default
                    dob_val = get_val('date_of_birth')
                    if dob_val:
                        if isinstance(dob_val, (date, datetime)):
                            dob = dob_val if isinstance(dob_val, date) else dob_val.date()
                        else:
                            try:
                                dob = datetime.strptime(str(dob_val).strip(), '%Y-%m-%d').date()
                            except:
                                pass

                    # Prep User Data (replica of StudentForm logic)
                    base_username = f"STD{admission_number}"
                    username = base_username
                    counter = 1
                    while User.objects.filter(username=username).exists():
                        username = f"{base_username}{counter}"
                        counter += 1
                    
                    raw_password = config.student_portal_password
                    
                    # Create User
                    user = User.objects.create_user(
                        username=username,
                        password=raw_password,
                        first_name=first_name,
                        last_name=last_name
                    )
                    user.groups.add(students_group)
                    
                    # Create Student
                    student = Student.objects.create(
                        school=config,
                        user=user,
                        admission_number=admission_number,
                        first_name=first_name,
                        last_name=last_name,
                        gender=gender,
                        grade=grade_obj,
                        date_of_birth=dob,
                        admission_fee=config.admission_fee,
                        current_term=int(config.current_term.replace('TERM_', '')) if 'TERM_' in config.current_term else 1,
                        term1_fees=grade_obj.term1_fees if grade_obj else 0,
                        term2_fees=grade_obj.term2_fees if grade_obj else 0,
                        term3_fees=grade_obj.term3_fees if grade_obj else 0
                    )

                    # Create Admission Fee Payment Record
                    ref_number = f'REF-BULK-{uuid.uuid4().hex[:6].upper()}'
                    Payment.objects.create(
                        school=config,
                        student=student,
                        amount=config.admission_fee,
                        payment_method='CASH',
                        transaction_id=ref_number,
                        reference_number=ref_number,
                        term=student.current_term,
                        status='COMPLETED'
                    )

                    results['imported'] += 1

            except Exception as row_err:
                results['skipped'] += 1
                results['errors'].append(f"Row {row_idx+2}: {str(row_err)}")

        return JsonResponse({
            'success': True,
            'summary': results
        })

    except Exception as e:
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def import_students(request):
    """Old template-based import redirect"""
    return redirect('student_list')

@csrf_exempt
@login_required
def download_student_template(request):
    """Download template for student import"""
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Student Template"
        
        # Add headers
        headers = [
            'Registration No.',
            'First Name',
            'Last Name',
            'Gender (M/F)',
            'Date of Birth (YYYY-MM-DD)',
            'Grade',
            'Parent Name',
            'Parent Email',
            'Parent Phone',
            'Address'
        ]
        ws.append(headers)
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=student_template.xlsx'
        wb.save(response)
        return response
    except Exception as e:
        print(f"Template Download Error: {e}")
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@admin_required
def teacher_create(request):
    if request.method == 'POST':
        form = TeacherForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                teacher = form.save()
                messages.success(request, f'Teacher {teacher.get_full_name()} has been created successfully!')
                return redirect('teacher_list')
            except Exception as e:
                messages.error(request, f'Error creating teacher: {str(e)}')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = TeacherForm()
    
    context = {
        'form': form,
        'title': 'Add New Teacher',
        'page_title': 'Add New Teacher'
    }
    return render(request, 'schools/teacher_form.html', context)

@login_required
@admin_required
def teacher_edit(request, pk):
    teacher = get_object_or_404(Teacher, pk=pk)
    if request.method == 'POST':
        form = TeacherForm(request.POST, request.FILES, instance=teacher)
        if form.is_valid():
            try:
                with transaction.atomic():
                    teacher = form.save()
                    if hasattr(teacher, 'changes') and teacher.changes:
                        changes_text = ', '.join(teacher.changes)
                        messages.success(
                            request, 
                            f'Teacher updated successfully. Changed: {changes_text}'
                        )
                    else:
                        messages.info(request, 'No changes were made')
                    return redirect('teacher_detail', pk=teacher.pk)
            except Exception as e:
                messages.error(request, f'Error updating teacher: {str(e)}')
    else:
        form = TeacherForm(instance=teacher)
    
    context = {
        'form': form,
        'teacher': teacher,
        'title': f'Edit {teacher.get_full_name()}',
        'page_title': f'Edit {teacher.get_full_name()}'
    }
    return render(request, 'schools/teacher_form.html', context)

@login_required
@admin_required
def teacher_delete(request, pk):
    teacher = get_object_or_404(Teacher, pk=pk)
    
    if request.method == 'POST':
        try:
            teacher_name = teacher.get_full_name()
            teacher.delete()
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': f'{teacher_name} has been deleted successfully.'
                })
            messages.success(request, f'{teacher_name} has been deleted successfully.')
            return redirect('teacher_list')
        except Exception as e:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'error': str(e)
                })
            messages.error(request, f'Error deleting teacher: {str(e)}')
            return redirect('teacher_list')
    
    context = {'teacher': teacher}
    return render(request, 'schools/teacher_confirm_delete.html', context)

@login_required
def teacher_detail(request, pk):
    teacher = get_object_or_404(Teacher, pk=pk)
    
    # Get teacher salaries with prefetch
    salaries = Salary.objects.filter(employee=teacher)\
        .prefetch_related('allowance_items', 'deduction_items')\
        .order_by('-month')
    
    # Calculate net salary for each salary record
    salaries_with_net = []
    for salary in salaries:
        total_allowances = salary.allowance_items.aggregate(total=Sum('amount'))['total'] or 0
        total_deductions = salary.deduction_items.aggregate(total=Sum('amount'))['total'] or 0
        net_salary = salary.amount + total_allowances - total_deductions
        salaries_with_net.append({
            'salary': salary,
            'total_allowances': total_allowances,
            'total_deductions': total_deductions,
            'net_salary': net_salary
        })
    
    # Check if user is viewing their own profile
    # Check both ways: if user has teacher profile and it matches, or if teacher's user matches request.user
    is_own_profile = (
        (hasattr(request.user, 'teacher') and request.user.teacher.pk == teacher.pk) or
        (teacher.user == request.user)
    )
    
    context = {
        'teacher': teacher,
        'schedules': teacher.schedule_set.all(),
        'leaves': teacher.leaves.all(),
        'salaries': salaries,
        'salaries_with_net': salaries_with_net,
        'salary_stats': {
            'total_earnings': salaries.aggregate(total=Sum('amount'))['total'] or 0,
            'avg_salary': salaries.aggregate(avg=Avg('amount'))['avg'] or 0,
            'total_allowances': salaries.aggregate(
                total=Sum('allowance_items__amount'))['total'] or 0,
            'total_deductions': salaries.aggregate(
                total=Sum('deduction_items__amount'))['total'] or 0,
        },
        'title': f'Teacher: {teacher.get_full_name()}',
        'page_title': f'Teacher: {teacher.get_full_name()}',
        'is_own_profile': is_own_profile,
    }
    return render(request, 'schools/teacher_detail.html', context)

@login_required
def teacher_leave(request, pk):
    """Handle teacher leave requests"""
    teacher = get_object_or_404(Teacher, pk=pk)
    
    if request.method == 'POST':
        form = LeaveForm(request.POST)
        if form.is_valid():
            leave = form.save(commit=False)
            leave.teacher = teacher
            leave.save()
            messages.success(request, 'Leave request submitted successfully')
            return redirect('teacher_detail', pk=pk)
    else:
        form = LeaveForm()
    
    # Get leave history
    leave_history = Leave.objects.filter(teacher=teacher).order_by('-start_date')
    
    context = {
        'form': form,
        'teacher': teacher,
        'leave_history': leave_history,
        'title': f'Leave Request - {teacher.get_full_name()}',
        'page_title': f'Leave Request - {teacher.get_full_name()}'
    }
    
    return render(request, 'schools/teacher_leave.html', context)

@login_required
def manage_schedule(request, teacher_id):
    """Manage teacher's schedule"""
    teacher = get_object_or_404(Teacher, pk=teacher_id)
    
    if request.method == 'POST':
        # Handle schedule updates
        schedule_data = json.loads(request.body)
        try:
            # Update or create schedule entries
            for entry in schedule_data:
                Schedule.objects.update_or_create(
                    teacher=teacher,
                    date=entry['date'],
                    start_time=entry['start_time'],
                    defaults={
                        'end_time': entry['end_time'],
                        'subject_id': entry['subject_id'],
                        'grade_id': entry['grade_id']
                    }
                )
            return JsonResponse({'status': 'success'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    
    # Get current schedule
    schedules = Schedule.objects.filter(
        teacher=teacher,
        date__gte=timezone.now().date()
    ).select_related('subject', 'grade').order_by('date', 'start_time')
    context = {
        'teacher': teacher,
        'schedules': schedules,
        'subjects': teacher.subjects.all(),
        'grades': Grade.objects.all(),
        'title': f'Manage Schedule - {teacher.get_full_name()}',
        'page_title': f'Manage Schedule - {teacher.get_full_name()}'
    }
    
    return render(request, 'schools/manage_schedule.html', context)

@login_required
def assessment_list(request):
    current_term = get_current_term()
    current_year = timezone.now().year
    
    # Get students ordered by grade and name
    students = Student.objects.select_related('grade').prefetch_related(
        Prefetch(
            'assessments',
            queryset=Assessment.objects.select_related('student').prefetch_related(
                'results'
            ).order_by('-date'),
            to_attr='recent_assessments'
        )
    ).order_by('grade__name', 'first_name')  # Order by grade first, then name
    
    # Apply filters
    if request.GET.get('term') and request.GET.get('term') != 'all':
        term_filter = request.GET.get('term')
        students = students.filter(assessments__term=term_filter).distinct()
    
    if request.GET.get('type') and request.GET.get('type') != 'all':
        type_filter = request.GET.get('type')
        students = students.filter(assessments__assessment_type=type_filter).distinct()
    
    if request.GET.get('grade') and request.GET.get('grade') != 'all':
        grade_filter = request.GET.get('grade')
        students = students.filter(grade_id=grade_filter)
    
    # Calculate summary statistics
    total_assessments = Assessment.objects.count()
    recent_assessments = Assessment.objects.filter(
        date__month=timezone.now().month
    ).count()
    
    # Get this term's assessments
    assessments_this_term = Assessment.objects.filter(
        term=current_term,
        date__year=current_year
    ).count()
    
    # Calculate average performance
    avg_performance = AssessmentResult.objects.aggregate(
        avg=Avg('performance_level')
    )['avg'] or 0
    
    # Count unique students assessed
    students_assessed = Assessment.objects.values('student').distinct().count()
    
    context = {
        # Summary statistics
        'total_assessments': total_assessments,
        'recent_assessments': recent_assessments,
        'avg_performance': round(float(avg_performance), 1),
        'assessments_this_term': assessments_this_term,
        'students_assessed': students_assessed,
        'current_term': current_term,
        
        # Student list with assessments
        'students': students,
        
        # Filter options
        'terms': list(range(1, 4)),  # Terms 1-3
        'grades': Grade.objects.all(),
        'assessment_types': Assessment.ASSESSMENT_TYPES,
        
        # Applied filters
        'filters': {
            'term': request.GET.get('term', 'all'),
            'type': request.GET.get('type', 'all'),
            'grade': request.GET.get('grade', 'all'),
        },
        
        'title': 'Assessments',
        'page_title': 'Assessments'
    }
    
    return render(request, 'schools/assessment_list.html', context)

@login_required
def fees_dashboard(request):
    # Calculate collection progress
    total_collected = Payment.objects.aggregate(
        total=Coalesce(Sum('amount'), 0, output_field=DecimalField(max_digits=10, decimal_places=2))
    )['total']
    total_expected = Grade.objects.aggregate(
        total=Coalesce(
            Sum(F('term1_fees') + F('term2_fees') + F('term3_fees')), 
            0,
            output_field=DecimalField(max_digits=10, decimal_places=2)
        )
    )['total']
    total_pending = max(0, total_expected - total_collected)
    
    # Calculate admission fees statistics
    ADMISSION_FEE = Decimal('500.00')
    total_students = Student.objects.count()
    admission_fees_collected = Payment.objects.filter(
        amount=ADMISSION_FEE
    ).aggregate(
        total=Coalesce(Sum('amount'), 0, output_field=DecimalField(max_digits=10, decimal_places=2))
    )['total']
    total_admission_fees_expected = total_students * ADMISSION_FEE
    admission_fees_pending = max(0, total_admission_fees_expected - admission_fees_collected)
    
    # Calculate students with balance and fully paid students
    students_with_balance = Student.objects.annotate(
        total_fees=F('grade__term1_fees') + F('grade__term2_fees') + F('grade__term3_fees'),
        total_paid=Coalesce(
            Subquery(
                Payment.objects.filter(student=OuterRef('pk'))
                .values('student')
                .annotate(total=Sum('amount'))
                .values('total')
            ),
            0,
            output_field=DecimalField(max_digits=10, decimal_places=2)
        ),
        balance=F('total_fees') - F('total_paid')
    ).filter(balance__gt=0)
    
    # Fee statistics for the cards
    fee_stats = {
        'total_collected': total_collected,
        'total_pending': total_pending,
        'total_expected': total_expected,
        'students_with_balance': students_with_balance.count(),
        'fully_paid_students': Student.objects.count() - students_with_balance.count(),
        'admission_fee': ADMISSION_FEE,
        'admission_fees_collected': admission_fees_collected,
        'admission_fees_expected': total_admission_fees_expected,
        'admission_fees_pending': admission_fees_pending,
        'admission_fees_percent': (admission_fees_collected / total_admission_fees_expected * 100) if total_admission_fees_expected > 0 else 0
    }
    
    collection_progress = {
        'collected': float(total_collected),
        'pending': float(total_pending),
        'total': float(total_expected),
        'collection_rate': round((float(total_collected) / float(total_expected) * 100) if total_expected > 0 else 0, 1),
        'paid_students': Payment.objects.values('student').distinct().count(),
        'pending_students': Student.objects.exclude(
            id__in=Payment.objects.values('student')
        ).count()
    }

    # Calculate monthly data for payment trends
    monthly_data = [0] * 12
    current_year = timezone.now().year
    payments = Payment.objects.filter(date__year=current_year)
    
    for payment in payments:
        if payment.date:
            month_index = payment.date.month - 1
            monthly_data[month_index] += float(payment.amount or 0)
    
    # Get payment methods distribution
    payment_methods = Payment.objects.values('payment_method').annotate(
        count=Count('id'),
        total=Coalesce(Sum('amount'), 0, output_field=DecimalField(max_digits=10, decimal_places=2))
    ).order_by('-count')

    payment_method_labels = [p['payment_method'] for p in payment_methods]
    payment_method_data = [p['count'] for p in payment_methods]

    # Calculate grade-wise fee collection
    grades = Grade.objects.all()
    grade_labels = [grade.name for grade in grades]
    grade_collected = []
    grade_pending = []

    for grade in grades:
        collected = Payment.objects.filter(student__grade=grade).aggregate(
            total=Coalesce(Sum('amount'), 0, output_field=DecimalField(max_digits=10, decimal_places=2))
        )['total']
        expected = grade.term1_fees + grade.term2_fees + grade.term3_fees
        pending = max(0, expected - collected)
        
        grade_collected.append(float(collected))
        grade_pending.append(float(pending))
    
    # Get recent payments
    recent_payments = Payment.objects.select_related('student', 'student__grade').order_by('-date')[:3]

    # Get students with pending fees
    students_with_pending = Student.objects.annotate(
        total_term_fees=F('grade__term1_fees') + F('grade__term2_fees') + F('grade__term3_fees'),
        total_paid=Coalesce(
            Subquery(
                Payment.objects.filter(student=OuterRef('pk'))
                .values('student')
                .annotate(total=Sum('amount'))
                .values('total')
            ),
            0,
            output_field=DecimalField(max_digits=10, decimal_places=2)
        ),
        balance=F('total_term_fees') - F('total_paid')
    ).filter(balance__gt=0).order_by('-balance')[:3]
    
    context = {
        'fee_stats': fee_stats,
        'collection_progress': collection_progress,
        'monthly_data': monthly_data,
        'payment_method_labels': payment_method_labels,
        'payment_method_data': payment_method_data,
        'grade_labels': grade_labels,
        'grade_collected': grade_collected,
        'grade_pending': grade_pending,
        'recent_payments': recent_payments,
        'students_with_pending': students_with_pending,
        'admission_fee': ADMISSION_FEE,
        'admission_fees_collected': admission_fees_collected,
        'admission_fees_expected': total_admission_fees_expected,
        'admission_fees_pending': admission_fees_pending,
        'title': 'Fees Dashboard',
        'page_title': 'Fees Dashboard'
    }
    
    return render(request, 'schools/fees_dashboard.html', context)


@login_required
def import_teachers(request):
    """Import teachers from Excel/CSV file"""
    if request.method == 'POST':
        try:
            file = request.FILES['file']
            # Process the uploaded file (Excel/CSV)
            # Add your file processing logic here
            
            messages.success(request, 'Teachers imported successfully')
            return redirect('teacher_list')
        except Exception as e:
            messages.error(request, f'Error importing teachers: {str(e)}')
            return redirect('teacher_list')
    return redirect('teacher_list')

@login_required
def download_teacher_template(request):
    """Download teacher import template"""
    # Create response with Excel/CSV template
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="teacher_import_template.xlsx"'
    
    # Create workbook and add headers
    wb = Workbook()
    ws = wb.active
    headers = ['First Name', 'Last Name', 'Email', 'Phone', 'Subject', 'Grade']
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    wb.save(response)
    return response

@login_required
def assessment_create(request):
    """Create a new assessment with subject results"""
    if request.method == 'POST':
        form = AssessmentForm(request.POST)
        if form.is_valid():
            assessment = form.save(commit=False)
            assessment.recorded_by = request.user
            assessment.save()
            
            # Create formset for assessment results
            formset = AssessmentResultFormSet(request.POST, instance=assessment)
            if formset.is_valid():
                formset.save()
                messages.success(request, 'Assessment created successfully')
                return redirect('assessment_detail', pk=assessment.pk)
            else:
                # If formset is invalid, delete the assessment and show errors
                assessment.delete()
                messages.error(request, 'Please correct the errors below.')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = AssessmentForm()
        formset = AssessmentResultFormSet()
    
    context = {
        'form': form,
        'result_formset': formset,
        'students': Student.objects.all(),
        'subjects': Subject.objects.all(),
        'title': 'Create Assessment',
        'page_title': 'Create Assessment'
    }
    return render(request, 'schools/assessment_form.html', context)

@login_required
def record_payment(request, student_id):
    student = get_object_or_404(Student, pk=student_id)
    
    if request.method == 'POST':
        try:
            # Get and validate amount
            amount = request.POST.get('amount', '0')
            try:
                # Remove any commas and convert to Decimal
                amount = Decimal(amount.replace(',', '').strip())
                if amount <= 0:
                    raise ValueError("Amount must be greater than zero")
                if amount > 1000000:
                    raise ValueError("Amount cannot exceed KES 1,000,000")
            except (TypeError, ValueError, InvalidOperation) as e:
                messages.error(request, f'Invalid amount: {str(e)}')
                # Check if user came from finance page
                if 'finance' in request.META.get('HTTP_REFERER', ''):
                    return redirect('student_finance', pk=student_id)
                return redirect('student_detail', pk=student_id)
                
            # Validate payment method
            payment_method = request.POST.get('payment_method')
            if not payment_method:
                messages.error(request, 'Payment method is required')
                # Check if user came from finance page
                if 'finance' in request.META.get('HTTP_REFERER', ''):
                    return redirect('student_finance', pk=student_id)
                return redirect('student_detail', pk=student_id)
            
            # Get reference number and validate based on payment method
            reference_number = None
            if payment_method == 'MPESA':
                reference_number = request.POST.get('mpesa_reference_number') or request.POST.get('reference_number')
            elif payment_method == 'BANK':
                reference_number = request.POST.get('bank_reference_number') or request.POST.get('reference_number')
            elif payment_method == 'CHEQUE':
                reference_number = request.POST.get('cheque_reference_number') or request.POST.get('reference_number')
            
            if payment_method in ['MPESA', 'BANK', 'CHEQUE'] and not reference_number:
                messages.error(request, f'{payment_method} reference number is required')
                # Check if user came from finance page
                if 'finance' in request.META.get('HTTP_REFERER', ''):
                    return redirect('student_finance', pk=student_id)
                return redirect('student_detail', pk=student_id)
            
            # For cash payments, generate a reference number
            if payment_method == 'CASH':
                reference_number = f'CASH-{timezone.now().strftime("%Y%m%d%H%M%S")}-{student_id}'
            
            # Get phone number for M-Pesa (handle both 'phone' and 'phone_number' for backward compatibility)
            phone_number = None
            if payment_method == 'MPESA':
                phone_number = request.POST.get('phone_number') or request.POST.get('phone')
                if not phone_number or not phone_number.startswith('254'):
                    messages.error(request, 'Valid phone number starting with 254 is required for M-Pesa payments')
                    # Check if user came from finance page
                    if 'finance' in request.META.get('HTTP_REFERER', ''):
                        return redirect('student_finance', pk=student_id)
                    return redirect('student_detail', pk=student_id)
            
            # Get current term
            current_term = student.current_term
            
            # Create the payment record
            payment = Payment.objects.create(
                student=student,
                amount=amount,
                payment_method=payment_method,
                reference_number=reference_number,
                phone_number=phone_number,
                term=current_term,
                status='COMPLETED'
            )
            
            # Automatically send receipt to parent
            try:
                from .utils import send_payment_receipt
                logger = logging.getLogger(__name__)
                success, result = send_payment_receipt(payment)
                if success:
                    if isinstance(result, dict):
                        if result.get('email_sent'):
                            messages.info(request, 'Receipt sent to parent email')
                        if result.get('whatsapp_sent'):
                            messages.info(request, 'Receipt notification sent to parent WhatsApp')
                else:
                    # Log error but don't fail the payment recording
                    logger.warning(f"Receipt sending failed: {result}")
            except Exception as e:
                # Log error but don't fail the payment recording
                logger = logging.getLogger(__name__)
                logger.error(f"Error sending receipt: {str(e)}")
            
            messages.success(
                request, 
                f'Payment of KES {amount:,.2f} recorded successfully. Reference: {reference_number}'
            )
            
            # Check if user came from finance page or if next parameter is provided
            next_url = request.POST.get('next') or request.GET.get('next')
            if next_url:
                return redirect(next_url)
            elif 'finance' in request.META.get('HTTP_REFERER', ''):
                return redirect('student_finance', pk=student_id)
            
            # Default redirect to payment detail page
            return redirect('payment_detail', pk=payment.pk)
            
        except Exception as e:
            messages.error(request, f'Error recording payment: {str(e)}')
            # Check if user came from finance page
            if 'finance' in request.META.get('HTTP_REFERER', ''):
                return redirect('student_finance', pk=student_id)
            return redirect('student_detail', pk=student_id)
    
    # If GET request, redirect to student detail
    return redirect('student_detail', pk=student_id)

@login_required
@accountant_required
def payment_add(request, student_id):
    student = get_object_or_404(Student, pk=student_id)
    
    if request.method == 'POST':
        form = PaymentForm(request.POST)
        if form.is_valid():
            payment = form.save(commit=False)
            payment.student = student
            payment.recorded_by = request.user
            payment.save()
            
            # Automatically send receipt to parent via WhatsApp
            try:
                from .utils import send_payment_receipt
                import logging
                logger = logging.getLogger(__name__)
                success, result = send_payment_receipt(payment)
                if success:
                    if isinstance(result, dict) and result.get('whatsapp_sent'):
                        messages.info(request, 'Receipt notification sent to parent WhatsApp')
                else:
                    logger.warning(f"Receipt sending failed: {result}")
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error sending receipt: {str(e)}")
            
            return redirect('student_detail', pk=student_id)
    else:
        form = PaymentForm()
    
    context = {
        'form': form,
        'student': student,
        'page_title': 'Add Payment'
    }
    return render(request, 'schools/payment_form.html', context)

@login_required
def assessment_form(request):
    student_id = request.GET.get('student_id')
    student = get_object_or_404(Student, pk=student_id) if student_id else None
    
    if request.method == 'POST':
        form = AssessmentForm(request.POST)
        if form.is_valid():
            assessment = form.save(commit=False)
            if student:
                assessment.student = student
            if hasattr(request.user, 'teacher'):
                assessment.teacher = request.user.teacher
            assessment.recorded_by = request.user
            assessment.save()
            
            # Create formset for assessment results
            formset = AssessmentResultFormSet(request.POST, instance=assessment)
            if formset.is_valid():
                formset.save()
                messages.success(request, 'Assessment recorded successfully')
                if student:
                    return redirect('student_detail', pk=student.id)
                else:
                    return redirect('assessment_detail', pk=assessment.pk)
            else:
                # If formset is invalid, delete the assessment and show errors
                assessment.delete()
                messages.error(request, 'Please correct the errors below.')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        initial_data = {}
        if student:
            initial_data['student'] = student
        form = AssessmentForm(initial=initial_data)
        formset = AssessmentResultFormSet()
    
    context = {
        'form': form,
        'result_formset': formset,
        'student': student,
        'subjects': Subject.objects.all(),
        'title': 'Record Assessment',
        'page_title': 'Record Assessment'
    }
    return render(request, 'schools/assessment_form.html', context)

@login_required
def mark_attendance(request, student_id=None, grade_id=None):
    """Mark attendance for a single student or all students in a grade"""
    from datetime import datetime
    
    today = timezone.now().date()
    selected_date = request.GET.get('date', today.strftime('%Y-%m-%d'))
    
    # Get grade_id from URL if not provided
    if not grade_id:
        grade_id_param = request.GET.get('grade_id')
        if grade_id_param:
            try:
                grade_id = int(grade_id_param)
            except (ValueError, TypeError):
                grade_id = None
    
    try:
        selected_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        selected_date = today
    
    # Handle POST request - save attendance
    if request.method == 'POST':
        if request.content_type == 'application/json' or 'application/json' in request.content_type:
            # Bulk attendance marking
            try:
                data = json.loads(request.body)
            except (json.JSONDecodeError, AttributeError):
                data = {}
            students_data = data.get('students', [])
            date_str = data.get('date', selected_date.strftime('%Y-%m-%d'))
            
            try:
                attendance_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except (ValueError, TypeError):
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid date format'
                }, status=400)
            
            created_count = 0
            updated_count = 0
            errors = []
            
            for student_data in students_data:
                student_id = student_data.get('student_id')
                status = student_data.get('status')
                remarks = student_data.get('remarks', '')
                
                if not student_id or not status:
                    continue
                
                try:
                    student = Student.objects.get(pk=student_id)
                    attendance, created = Attendance.objects.update_or_create(
                        student=student,
                        date=attendance_date,
                        defaults={
                            'status': status,
                            'remarks': remarks,
                            'recorded_by': request.user
                        }
                    )
                    if created:
                        created_count += 1
                    else:
                        updated_count += 1
                except Exception as e:
                    errors.append(f"Error for student {student_id}: {str(e)}")
            
            return JsonResponse({
                'success': True,
                'message': f'Attendance marked: {created_count} created, {updated_count} updated',
                'created': created_count,
                'updated': updated_count,
                'errors': errors
            })
        else:
            # Single student attendance
            student = get_object_or_404(Student, pk=student_id)
            status = request.POST.get('status')
            date_str = request.POST.get('date')
            remarks = request.POST.get('remarks', '')
            
            try:
                attendance_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except (ValueError, TypeError):
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid date format'
                }, status=400)
            
            attendance, created = Attendance.objects.update_or_create(
                student=student,
                date=attendance_date,
                defaults={
                    'status': status,
                    'remarks': remarks,
                    'recorded_by': request.user
                }
            )
            
            return JsonResponse({
                'success': True,
                'message': 'Attendance marked successfully',
                'created': created
            })
    
    # Handle GET request - display attendance page
    grade = None
    students = []
    attendance_records = {}
    
    if grade_id:
        try:
            grade = get_object_or_404(Grade, pk=grade_id)
            students = Student.objects.filter(grade=grade).order_by('first_name', 'last_name')
            # Get attendance records for selected date
            attendance_records = {
                record.student_id: record
                for record in Attendance.objects.filter(
                    student__in=students,
                    date=selected_date
                ).select_related('student')
            }
        except (ValueError, TypeError):
            grade = None
    elif student_id:
        try:
            student = get_object_or_404(Student, pk=student_id)
            students = [student]
            attendance = Attendance.objects.filter(
                student=student,
                date=selected_date
            ).first()
            if attendance:
                attendance_records[student.id] = attendance
        except (ValueError, TypeError):
            pass
    
    # Calculate statistics from saved attendance records
    present_count = 0
    absent_count = 0
    late_count = 0
    
    for record in attendance_records.values():
        if record.status == 'PRESENT':
            present_count += 1
        elif record.status == 'ABSENT':
            absent_count += 1
        elif record.status == 'LATE':
            late_count += 1
    
    # Get calendar events for the month
    calendar_events = []
    if students:
        start_of_month = selected_date.replace(day=1)
        if selected_date.month == 12:
            end_of_month = selected_date.replace(year=selected_date.year + 1, month=1, day=1)
        else:
            end_of_month = selected_date.replace(month=selected_date.month + 1, day=1)
        
        month_attendance = Attendance.objects.filter(
            student__in=students,
            date__gte=start_of_month,
            date__lt=end_of_month
        ).select_related('student')
        
        for att in month_attendance:
            calendar_events.append({
                'title': f"{att.student.get_full_name()} - {att.get_status_display()}",
                'start': att.date.isoformat(),
                'status': att.status,
                'student_id': att.student.id,
                'color': '#10b981' if att.status == 'PRESENT' else '#ef4444' if att.status == 'ABSENT' else '#f59e0b'
            })
    
    context = {
        'grade': grade,
        'students': students,
        'attendance_records': attendance_records,
        'selected_date': selected_date,
        'today': today,
        'calendar_events': json.dumps(calendar_events),
        'attendance_choices': Attendance.ATTENDANCE_STATUS,
        'present_count': present_count,
        'absent_count': absent_count,
        'late_count': late_count,
        'title': f'Mark Attendance - {grade.name if grade else "Student"}' if grade or student_id else 'Mark Attendance',
        'page_title': f'Mark Attendance - {grade.name if grade else "Student"}' if grade or student_id else 'Mark Attendance'
    }
    return render(request, 'schools/mark_attendance.html', context)

@login_required
def grade_subjects_api(request, grade_id):
    grade = get_object_or_404(Grade, pk=grade_id)
    subjects = Subject.objects.filter(grade=grade)
    data = {
        'subjects': [{'id': subject.id, 'name': subject.name} for subject in subjects]
    }
    return JsonResponse(data)

@login_required
def assessment_record(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    subjects = Subject.objects.all()
    
    # Get existing assessment if any
    existing_assessment = None
    if request.GET.get('assessment_id'):
        existing_assessment = get_object_or_404(Assessment, id=request.GET.get('assessment_id'))
    
    if request.method == 'POST':
        # Get data directly from POST since we're not using the form anymore
        assessment_type = request.POST.get('assessment_type')
        term = request.POST.get('term')
        date = request.POST.get('date')
        remarks = request.POST.get('remarks')
        
        # Debug: Print POST data
        print(f"POST data: {dict(request.POST)}")
        print(f"Assessment Type: {assessment_type}")
        print(f"Term: {term}")
        print(f"Date: {date}")
        
        # Validate required fields
        if not all([assessment_type, term, date]):
            if not assessment_type:
                messages.error(request, 'Please select an assessment type.')
            if not term:
                messages.error(request, 'Please select a term.')
            if not date:
                messages.error(request, 'Please select a date.')
            print(f"Validation failed: assessment_type={assessment_type}, term={term}, date={date}")
        else:
            # Check if at least one subject has performance data
            has_subject_data = False
            for subject in subjects:
                if request.POST.get(subject.code):
                    has_subject_data = True
                    break
            
            if not has_subject_data:
                messages.error(request, 'Please provide performance data for at least one subject.')
                print("Validation failed: No subject performance data provided")
            else:
                try:
                    # Create or update assessment
                    if existing_assessment:
                        assessment = existing_assessment
                        assessment.assessment_type = assessment_type
                        assessment.term = int(term)
                        assessment.date = date
                        assessment.remarks = remarks
                        assessment.save()
                        print(f"Updated existing assessment: {assessment.id}")
                    else:
                        assessment = Assessment.objects.create(
                            student=student,
                            assessment_type=assessment_type,
                            term=int(term),
                            date=date,
                            remarks=remarks,
                            recorded_by=request.user
                        )
                        print(f"Created new assessment: {assessment.id}")
                    
                    # Save results for each subject that has marks
                    for subject in subjects:
                        marks = request.POST.get(subject.code)  # Using subject code as field name
                        subject_remarks = request.POST.get(f'remarks_{subject.code}')
                        
                        print(f"Subject {subject.code}: marks={marks}, remarks={subject_remarks}")
                        
                        # Only create result if marks are provided
                        if marks:
                            # Convert performance level to decimal score (1-4 scale to 0-1 scale)
                            score = float(marks) / 4.0  # Convert 1-4 scale to 0-1 scale
                            
                            # Check if result already exists for this subject
                            result, created = AssessmentResult.objects.get_or_create(
                                assessment=assessment,
                                subject=subject,
                                defaults={
                                    'performance_level': marks,
                                    'weekly_score': score if assessment.assessment_type == 'weekly' else None,
                                    'opener_score': score if assessment.assessment_type == 'opener' else None,
                                    'midpoint_score': score if assessment.assessment_type == 'mid-term' else None,
                                    'endpoint_score': score if assessment.assessment_type == 'end-term' else None,
                                    'remarks': subject_remarks
                                }
                            )
                            
                            # Update existing result if it exists
                            if not created:
                                result.performance_level = marks
                                result.weekly_score = score if assessment.assessment_type == 'weekly' else result.weekly_score
                                result.opener_score = score if assessment.assessment_type == 'opener' else result.opener_score
                                result.midpoint_score = score if assessment.assessment_type == 'mid-term' else result.midpoint_score
                                result.endpoint_score = score if assessment.assessment_type == 'end-term' else result.endpoint_score
                                result.remarks = subject_remarks
                                result.save()
                            
                            print(f"Saved result for {subject.code}: performance_level={marks}, score={score}")
                    
                    messages.success(request, 'Assessment recorded successfully.')
                    return redirect('student_detail', pk=student.id)
                    
                except Exception as e:
                    messages.error(request, f'Error saving assessment: {str(e)}')
                    print(f"Error: {str(e)}")
                    import traceback
                    traceback.print_exc()
    else:
        form = AssessmentForm()
    
    # Get existing results for the assessment if editing
    existing_results = {}
    if existing_assessment:
        for result in existing_assessment.results.all():
            existing_results[result.subject.id] = {
                'marks': result.performance_level,
                'remarks': result.remarks
            }
    
    # Always define form variable to avoid UnboundLocalError
    if 'form' not in locals():
        form = AssessmentForm()
    
    context = {
        'student': student,
        'subjects': subjects,
        'form': form,
        'existing_assessment': existing_assessment,
        'existing_results': existing_results
    }
    return render(request, 'schools/assessment_record.html', context)

@login_required
def weekly_assessment_record(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    subjects = Subject.objects.all()
    
    # Get existing assessment if any
    existing_assessment = None
    if request.GET.get('assessment_id'):
        existing_assessment = get_object_or_404(Assessment, id=request.GET.get('assessment_id'))
    
    if request.method == 'POST':
        assessment_type = 'weekly'  # Always weekly for this view
        term = request.POST.get('term')
        date = request.POST.get('date')
        week_number = request.POST.get('week_number')
        remarks = request.POST.get('remarks')
        
        # Use existing assessment or create new one
        if existing_assessment:
            assessment = existing_assessment
            # Update assessment details if changed
            assessment.assessment_type = assessment_type
            assessment.term = term
            assessment.date = date
            assessment.remarks = remarks
            assessment.week_number = week_number
            assessment.save()
        else:
            # Prevent duplicates for same student/term/week by updating existing
            assessment, _ = Assessment.objects.update_or_create(
                student=student,
                assessment_type=assessment_type,
                term=term,
                week_number=week_number,
                defaults={
                    'date': date,
                    'recorded_by': request.user,
                    'remarks': remarks,
                }
            )
        
        # Save results for each subject that has marks
        for subject in subjects:
            marks = request.POST.get(f'marks_{subject.id}')
            subject_remarks = request.POST.get(f'remarks_{subject.id}')
            
            # Only create result if marks are provided
            if marks:
                # Convert percentage marks to decimal (e.g., 25% becomes 0.25)
                score = float(marks) / 10  # Marks are out of 10
                
                # Check if result already exists for this subject
                result, created = AssessmentResult.objects.get_or_create(
                    assessment=assessment,
                    subject=subject,
                    defaults={
                        'marks': marks,
                        'weekly_score': score,
                        'remarks': subject_remarks
                    }
                )
                
                # Update existing result if it exists
                if not created:
                    result.marks = marks
                    result.weekly_score = score
                    result.remarks = subject_remarks
                    result.save()
        
        messages.success(request, 'Weekly assessment recorded successfully.')
        return redirect('student_detail', pk=student.id)
    
    # Get existing results for the assessment if editing
    existing_results = {}
    if existing_assessment:
        for result in existing_assessment.results.all():
            existing_results[result.subject.id] = {
                'marks': result.marks,
                'remarks': result.remarks
            }
    
    # Calculate current week number
    from datetime import datetime
    today = datetime.now()
    start_of_year = datetime(today.year, 1, 1)
    days = (today - start_of_year).days
    current_week = min((days + start_of_year.weekday() + 1) // 7, 14)  # Cap at 14 weeks
    
    context = {
        'student': student,
        'subjects': subjects,
        'existing_assessment': existing_assessment,
        'existing_results': existing_results,
        'current_week': current_week,
        'current_date': today.strftime('%Y-%m-%d')
    }
    return render(request, 'schools/weekly_assessment_record.html', context)

@login_required
def weekly_assessment_detail(request, student_id, pk):
    student = get_object_or_404(Student, id=student_id)
    assessment = get_object_or_404(Assessment, pk=pk, student=student, assessment_type='weekly')
    results = assessment.results.all().select_related('subject')
    
    # Calculate weekly assessment statistics
    total_marks = 0
    subject_count = 0
    subject_stats = []
    
    for result in results:
        if result.marks is not None:
            marks = float(result.marks)
            total_marks += marks
            subject_count += 1
            
            # Determine performance level
            if marks >= 8:
                performance_level = 4  # Exceeding
                performance_color = 'success'
            elif marks >= 6:
                performance_level = 3  # Meeting
                performance_color = 'primary'
            elif marks >= 4:
                performance_level = 2  # Approaching
                performance_color = 'warning'
            else:
                performance_level = 1  # Below
                performance_color = 'danger'
            
            subject_stats.append({
                'subject': result.subject,
                'marks': marks,
                'performance_level': performance_level,
                'performance_color': performance_color,
                'remarks': result.remarks
            })
    
    # Calculate average score
    average_score = total_marks / subject_count if subject_count > 0 else 0
    
    # Calculate percentage for progress bar (convert from 0-10 scale to 0-100%)
    average_percentage = (average_score / 10) * 100
    
    # Determine overall performance
    if average_score >= 8:
        overall_performance = 'Exceeding Expectations'
        overall_color = 'success'
    elif average_score >= 6:
        overall_performance = 'Meeting Expectations'
        overall_color = 'primary'
    elif average_score >= 4:
        overall_performance = 'Approaching Expectations'
        overall_color = 'warning'
    else:
        overall_performance = 'Below Expectations'
        overall_color = 'danger'
    
    # Calculate week number
    from datetime import datetime, date
    assessment_date = assessment.date
    start_of_year = date(assessment_date.year, 1, 1)
    days = (assessment_date - start_of_year).days
    week_number = min((days + start_of_year.weekday() + 1) // 7, 14)
    
    context = {
        'student': student,
        'assessment': assessment,
        'results': results,
        'subject_stats': subject_stats,
        'average_score': average_score,
        'average_percentage': average_percentage,
        'overall_performance': overall_performance,
        'overall_color': overall_color,
        'week_number': week_number,
        'total_subjects': subject_count,
        'is_weekly': True
    }
    return render(request, 'schools/weekly_assessment_detail.html', context)

@login_required
def assessment_detail(request, pk):
    assessment = get_object_or_404(Assessment, pk=pk)
    results = assessment.results.all().select_related('subject')
    
    # Calculate statistics based on assessment type
    if assessment.assessment_type == 'weekly':
        # Weekly assessment statistics
        total_marks = 0
        subject_count = 0
        subject_stats = []
        
        for result in results:
            if result.marks is not None:
                marks = float(result.marks)
                total_marks += marks
                subject_count += 1
                
                # Calculate percentage score (marks out of 10)
                percentage = (marks / 10) * 100
                
                subject_stats.append({
                    'subject': result.subject.name,
                    'marks': marks,
                    'percentage': round(percentage, 1),
                    'remarks': result.remarks
                })
        
        # Calculate overall average
        average_percentage = round((total_marks / (subject_count * 30)) * 100, 1) if subject_count > 0 else 0
        
        # Determine overall performance category
        if average_percentage >= 80:
            performance_category = 'Excellent'
            category_color = 'success'
        elif average_percentage >= 60:
            performance_category = 'Good'
            category_color = 'primary'
        elif average_percentage >= 40:
            performance_category = 'Average'
            category_color = 'warning'
        else:
            performance_category = 'Needs Improvement'
            category_color = 'danger'
        
        context = {
            'assessment': assessment,
            'results': results,
            'is_weekly': True,
            'subject_stats': subject_stats,
            'average_percentage': average_percentage,
            'performance_category': performance_category,
            'category_color': category_color,
            'title': f'Weekly Assessment - {assessment.date}',
            'page_title': 'Assessment Details'
        }
    else:
        # Other assessment types statistics
        performance_counts = {
            '4': 0,  # Exceeding
            '3': 0,  # Meeting
            '2': 0,  # Approaching
            '1': 0   # Below
        }
        
        for result in results:
            if result.performance_level:
                performance_counts[result.performance_level] += 1
        
        context = {
            'assessment': assessment,
            'results': results,
            'is_weekly': False,
            'performance_counts': performance_counts,
            'title': f'{assessment.get_assessment_type_display()} - {assessment.date}',
            'page_title': 'Assessment Details'
        }
    
    return render(request, 'schools/assessment_detail.html', context)

@login_required
def assessment_edit(request, pk):
    assessment = get_object_or_404(Assessment.objects.prefetch_related(
        'results',
        'results__subject'
    ), pk=pk)
    
    if request.method == 'POST':
        try:
            # Update results for each subject
            for result in assessment.results.all():
                performance_level = request.POST.get(f'subject_{result.subject.id}')
                
                if performance_level:
                    result.performance_level = performance_level
                    # Update the appropriate score based on assessment type
                    if assessment.assessment_type == 'opener':
                        result.opener_score = performance_level
                    elif assessment.assessment_type == 'mid-term':
                        result.midpoint_score = performance_level
                    elif assessment.assessment_type == 'end-term':
                        result.endpoint_score = performance_level
                    result.save()
            
            messages.success(request, 'Assessment results updated successfully!')
        except Exception as e:
            messages.error(request, f'Error updating assessment: {str(e)}')
        
        return redirect('student_detail', pk=assessment.student.id)
    
    context = {
        'assessment': assessment,
        'title': 'Edit Assessment',
        'performance_levels': [
            {'value': '4', 'label': 'Exceeding Expectations'},
            {'value': '3', 'label': 'Meeting Expectations'},
            {'value': '2', 'label': 'Approaching Expectations'},
            {'value': '1', 'label': 'Below Expectations'}
        ],
        'page_title': 'Edit Assessment'
    }
    return render(request, 'schools/assessment_edit.html', context)

@login_required
def assessment_delete(request, pk):
    if request.method == 'POST':
        assessment = get_object_or_404(Assessment, pk=pk)
        student_id = assessment.student.id
        
        # Check if user has permission to delete this assessment
        can_delete = (
            request.user.is_superuser or  # Admin can delete any assessment
            (hasattr(request.user, 'teacher') and 
             request.user.teacher.is_class_teacher and 
             request.user.teacher.grade == assessment.student.grade)  # Class teacher can delete assessments for their class
        )
        
        if can_delete:
            assessment.delete()
            messages.success(request, 'Assessment deleted successfully.')
        else:
            messages.error(request, 'You do not have permission to delete this assessment.')
        
        return redirect('student_detail', pk=student_id)
    return redirect('dashboard')

@login_required
@admin_required
def teacher_create(request):
    if request.method == 'POST':
        form = TeacherForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                teacher = form.save()
                messages.success(request, f'Teacher {teacher.get_full_name()} has been created successfully!')
                return redirect('teacher_list')
            except Exception as e:
                messages.error(request, f'Error creating teacher: {str(e)}')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = TeacherForm()
    
    context = {
        'form': form,
        'title': 'Add New Teacher',
        'page_title': 'Add New Teacher'
    }
    return render(request, 'schools/teacher_form.html', context)

@login_required
def payment_list(request, student_id=None):
    # Handle all payments case
    if student_id is None:
        payments = Payment.objects.select_related('student').order_by('-date')
        context = {
            'payments': payments,
            'title': 'All Payments',
            'page_title': 'All Payments'
        }
        return render(request, 'schools/payment_list.html', context)
    
    # Handle student-specific payments case
    student = get_object_or_404(Student, pk=student_id)
    payments = Payment.objects.filter(student=student).order_by('-date')
    
    fee_stats = calculate_fee_statistics(student=student)
    
    context = {
        'student': student,
        'payments': payments,
        'total_fees': fee_stats['total_fees'],
        'total_paid': fee_stats['total_collected'],
        'balance': fee_stats['total_fees'] - fee_stats['total_collected'],
        'payment_percent': fee_stats['total_collected'] / fee_stats['total_fees'] * 100 if fee_stats['total_fees'] > 0 else 0,
        'title': f'Payments - {student.get_full_name()}',
        'page_title': f'Payments - {student.get_full_name()}'
    }
    return render(request, 'schools/payment_list.html', context)
 
@login_required
def all_payments_list(request):
    """View for listing all payments"""
    payments = Payment.objects.select_related('student').order_by('-date')
    
    context = {
        'payments': payments,
        'title': 'All Payments',
        'page_title': 'All Payments'
    }
    return render(request, 'schools/payment_list.html', context)

@login_required
def payment_create(request):
    if request.method == 'POST':
        form = PaymentForm(request.POST)
        if form.is_valid():
            payment = form.save(commit=False)
            payment.recorded_by = request.user
            payment.save()
            
            # Automatically send receipt to parent via WhatsApp
            try:
                from .utils import send_payment_receipt
                import logging
                logger = logging.getLogger(__name__)
                success, result = send_payment_receipt(payment)
                if success:
                    if isinstance(result, dict) and result.get('whatsapp_sent'):
                        messages.info(request, 'Receipt notification sent to parent WhatsApp')
                else:
                    logger.warning(f"Receipt sending failed: {result}")
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error sending receipt: {str(e)}")
            
            messages.success(request, 'Payment recorded successfully')
            return redirect('payment_detail', pk=payment.pk)
    else:
        # Pre-fill student if provided in URL
        initial = {}
        student_id = request.GET.get('student')
        if student_id:
            student = get_object_or_404(Student, pk=student_id)
            initial['student'] = student
        form = PaymentForm(initial=initial)
    
    return render(request, 'schools/payment_form.html', {'form': form})

@login_required
def payment_detail(request, pk):
    payment = get_object_or_404(Payment, pk=pk)
    
    # Generate QR code
    qr_code = generate_payment_qr(payment)
    
    context = {
        'payment': payment,
        'qr_code': qr_code,
    }
    return render(request, 'schools/payment_detail.html', context)

@login_required
def export_fee_report(request):
    report_type = request.GET.get('report_type')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    format_type = request.GET.get('format', 'pdf')
    
    # Convert dates to datetime objects
    start_date = datetime.strptime(start_date, '%Y-%m-%d')
    end_date = datetime.strptime(end_date, '%Y-%m-%d')
    
    # Base queryset for payments
    payments = Payment.objects.select_related('student', 'student__grade').filter(
        date__range=[start_date, end_date]
    ).order_by('-date')
    
    # Get students with fee information
    students = Student.objects.select_related('grade').annotate(
        total_term_fees=F('grade__term1_fees') + F('grade__term2_fees') + F('grade__term3_fees'),
        total_paid=Coalesce(
            Subquery(
                Payment.objects.filter(student=OuterRef('pk'))
                .values('student')
                .annotate(total=Sum('amount'))
                .values('total')
            ),
            0,
            output_field=DecimalField(max_digits=10, decimal_places=2)
        ),
        balance=F('total_term_fees') - F('total_paid')
    )

    if format_type == 'excel':
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        
        if report_type == 'payments':
            ws.title = "Payment History"
            
            # Headers
            headers = ['Date', 'Student Name', 'Grade', 'Amount', 'Payment Method', 'Reference']
            ws.append(headers)
            
            # Add payment data
            for payment in payments:
                ws.append([
                    payment.date.strftime('%Y-%m-%d'),
                    payment.student.get_full_name(),
                    payment.student.grade.name,
                    float(payment.amount),
                    payment.payment_method,
                    payment.reference_number or ''
                ])
                
        elif report_type == 'defaulters':
            ws.title = "Fee Defaulters"
            
            # Headers
            headers = ['Student Name', 'Grade', 'Total Fees', 'Amount Paid', 'Balance']
            ws.append(headers)
            
            # Add defaulters data
            defaulters = students.filter(balance__gt=0).order_by('-balance')
            for student in defaulters:
                ws.append([
                    student.get_full_name(),
                    student.grade.name,
                    float(student.total_term_fees),
                    float(student.total_paid),
                    float(student.balance)
                ])
                
        else:  # summary
            ws.title = "Fee Summary"
            
            # Add summary statistics
            stats = calculate_fee_statistics()
            
            # Summary section
            ws.append(['Fee Collection Summary'])
            ws.append(['Total Expected', float(stats['total_expected'])])
            ws.append(['Total Collected', float(stats['total_collected'])])
            ws.append(['Total Pending', float(stats['total_pending'])])
            ws.append(['Collection Rate', f"{stats['collection_percentage']}%"])
            ws.append([])
            
            # Grade-wise breakdown
            ws.append(['Grade-wise Collection'])
            ws.append(['Grade', 'Expected', 'Collected', 'Pending', 'Collection Rate'])
            
            grades = Grade.objects.all()
            for grade in grades:
                collected = Payment.objects.filter(student__grade=grade).aggregate(
                    total=Coalesce(Sum('amount'), 0, output_field=DecimalField())
                )['total']
                expected = grade.term1_fees + grade.term2_fees + grade.term3_fees
                pending = max(0, expected - collected)
                rate = round((float(collected) / float(expected) * 100) if expected > 0 else 0, 1)
                
                ws.append([
                    grade.name,
                    float(expected),
                    float(collected),
                    float(pending),
                    f"{rate}%"
                ])
        
        # Style the worksheet
        for col in range(1, len(headers) + 1):
            col_letter = get_column_letter(col)
            # Style header
            header_cell = ws[f'{col_letter}1']
            header_cell.font = Font(bold=True)
            header_cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            
            # Adjust column width
            ws.column_dimensions[col_letter].width = 15
            
            # Format amount columns
            if any(amount_header in headers[col-1].lower() for amount_header in ['amount', 'fees', 'paid', 'balance']):
                for row in range(2, len(students) + 2):
                    cell = ws[f'{col_letter}{row}']
                    cell.number_format = '#,##0.00'
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=fee_report_{report_type}_{start_date.strftime("%Y%m%d")}.xlsx'
        wb.save(response)
        return response
        
    else:  # PDF format
        # Create the PDF document
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=fee_report_{report_type}_{start_date.strftime("%Y%m%d")}.pdf'
        
        doc = SimpleDocTemplate(
            response,
            pagesize=letter,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=72
        )
        
        # Container for the 'Flowable' objects
        elements = []
        
        # Define styles
        styles = getSampleStyleSheet()
        title_style = styles['Heading1']
        subtitle_style = styles['Heading2']
        normal_style = styles['Normal']
        
        # Add school header
        elements.append(Paragraph('School Fee Report', title_style))
        elements.append(Spacer(1, 12))
        elements.append(Paragraph(f'Period: {start_date.strftime("%B %d, %Y")} - {end_date.strftime("%B %d, %Y")}', subtitle_style))
        elements.append(Spacer(1, 12))
        
        if report_type == 'payments':
            # Payment History Report
            elements.append(Paragraph('Payment History', subtitle_style))
            elements.append(Spacer(1, 12))
            
            # Create table data
            data = [['Date', 'Student Name', 'Grade', 'Amount', 'Method']]
            for payment in payments:
                data.append([
                    payment.date.strftime('%Y-%m-%d'),
                    payment.student.get_full_name(),
                    payment.student.grade.name,
                    f'KES {payment.amount:,.2f}',
                    payment.payment_method
                ])
                
        elif report_type == 'defaulters':
            # Fee Defaulters Report
            elements.append(Paragraph('Fee Defaulters', subtitle_style))
            elements.append(Spacer(1, 12))
            
            # Create table data
            data = [['Student Name', 'Grade', 'Total Fees', 'Paid', 'Balance']]
            defaulters = students.filter(balance__gt=0).order_by('-balance')
            for student in defaulters:
                data.append([
                    student.get_full_name(),
                    student.grade.name,
                    f'KES {student.total_term_fees:,.2f}',
                    f'KES {student.total_paid:,.2f}',
                    f'KES {student.balance:,.2f}'
                ])
                
        else:  # summary
            # Fee Summary Report
            elements.append(Paragraph('Fee Collection Summary', subtitle_style))
            elements.append(Spacer(1, 12))
            
            stats = calculate_fee_statistics()
            
            # Summary table
            summary_data = [
                ['Total Expected', f'KES {stats["total_expected"]:,.2f}'],
                ['Total Collected', f'KES {stats["total_collected"]:,.2f}'],
                ['Total Pending', f'KES {stats["total_pending"]:,.2f}'],
                ['Collection Rate', f'{stats["collection_percentage"]}%']
            ]
            summary_table = Table(summary_data)
            summary_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ]))
            elements.append(summary_table)
            elements.append(Spacer(1, 20))
            
            # Grade-wise breakdown
            elements.append(Paragraph('Grade-wise Collection', subtitle_style))
            elements.append(Spacer(1, 12))
            
            data = [['Grade', 'Expected', 'Collected', 'Pending', 'Rate']]
            grades = Grade.objects.all()
            for grade in grades:
                collected = Payment.objects.filter(student__grade=grade).aggregate(
                    total=Coalesce(Sum('amount'), 0, output_field=DecimalField())
                )['total']
                expected = grade.term1_fees + grade.term2_fees + grade.term3_fees
                pending = max(0, expected - collected)
                rate = round((float(collected) / float(expected) * 100) if expected > 0 else 0, 1)
                
                data.append([
                    grade.name,
                    f'KES {expected:,.2f}',
                    f'KES {collected:,.2f}',
                    f'KES {pending:,.2f}',
                    f'{rate}%'
                ])
        
        # Create and style the table
        table = Table(data)
        table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)
        
        # Build the PDF document
        doc.build(elements)
        return response

    return redirect('fees_dashboard')

@login_required
def payment_edit(request, pk):
    payment = get_object_or_404(Payment.objects.select_related('student'), pk=pk)
    
    if request.method == 'POST':
        form = PaymentForm(request.POST, instance=payment)
        if form.is_valid():
            payment = form.save()
            
            # Automatically send receipt to parent after update
            try:
                from .utils import send_payment_receipt
                import logging
                logger = logging.getLogger(__name__)
                success, result = send_payment_receipt(payment)
                if success:
                    if isinstance(result, dict):
                        if result.get('whatsapp_sent'):
                            messages.info(request, 'Updated receipt notification sent to parent WhatsApp')
                else:
                    logger.warning(f"Receipt sending failed: {result}")
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error sending receipt: {str(e)}")
            
            messages.success(request, 'Payment updated successfully')
            return redirect('payment_detail', pk=payment.pk)
    else:
        form = PaymentForm(instance=payment)
    
    context = {
        'form': form,
        'payment': payment,
        'student': payment.student,  # Add student to context
        'title': 'Edit Payment',
        'page_title': 'Edit Payment'
    }
    return render(request, 'schools/payment_form.html', context)

@login_required
def payment_delete(request, pk):
    payment = get_object_or_404(Payment, pk=pk)
    
    if request.method == 'POST':
        student = payment.student
        payment.delete()
        messages.success(request, 'Payment deleted successfully')
        return redirect('student_detail', pk=student.pk)
    
    context = {
        'payment': payment,
        'title': 'Delete Payment',
        'message': 'Are you sure you want to delete this payment?',
        'page_title': 'Delete Payment'
    }
    return render(request, 'schools/confirm_delete.html', context)


@login_required
def student_assessment_detail(request, pk):
    assessment = get_object_or_404(Assessment.objects.prefetch_related(
        'results',
        'results__subject'
    ), pk=pk)
    
    # Check if user has permission to view this assessment
    if not request.user.is_superuser and assessment.student.grade.teacher != request.user.teacher:
        raise PermissionDenied
    
    # Get assessment data organized by subject
    assessment_data = []
    for result in assessment.results.all():
        # Get the appropriate score based on assessment type
        if assessment.assessment_type == 'openner':
            score = result.opener_score
        elif assessment.assessment_type == 'mid-term':
            score = result.midpoint_score
        elif assessment.assessment_type == 'end-term':
            score = result.endpoint_score
        else:
            score = 0
            
        data = {
            'subject': result.subject.name,
            'score': float(score) if score is not None else 0,
            'performance_level': result.get_performance_level_display(),
            'color': get_performance_color(result.performance_level)
        }
        assessment_data.append(data)
    
    context = {
        'assessment': assessment,
        'assessment_data': assessment_data,
        'title': f'{assessment.get_assessment_type_display()} Assessment Details',
        'page_title': f'{assessment.get_assessment_type_display()} Assessment Details'
    }
    return render(request, 'schools/student_assessment_detail.html', context)

def get_performance_color(level):
    """Return appropriate color for performance level"""
    colors = {
        '4': '#047857',  # Exceeding - Green
        '3': '#1e40af',  # Meeting - Blue
        '2': '#b45309',  # Approaching - Orange
        '1': '#b91c1c',  # Below - Red
    }
    return colors.get(str(level), '#64748b')  # Default gray
    if request.method == 'POST':
        form = SupportStaffForm(request.POST, request.FILES)
        if form.is_valid():
            staff = form.save()
            messages.success(request, 'Support staff created successfully')
            return redirect('employee_detail', pk=staff.id)
    else:
        form = SupportStaffForm()
    
    context = {
        'form': form,
        'title': 'Add Support Staff',
        'page_title': 'Add Support Staff'
    }
    return render(request, 'schools/hr/support_staff_form.html', context)

@admin_required
def admin_create(request):
    if request.method == 'POST':
        form = AdminStaffForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                admin = form.save()
                messages.success(request, f'Admin staff {admin.get_full_name()} created successfully')
                return redirect('employee_list')
            except Exception as e:
                messages.error(request, f'Error creating admin staff: {str(e)}')
    else:
        form = AdminStaffForm()
    
    return render(request, 'schools/hr/admin_form.html', {
        'form': form,
        'title': 'Create Admin Staff'
    })






@login_required
def other_staff_create(request):
    """Create other type of staff member"""
    if request.method == 'POST':
        form = OtherStaffForm(request.POST, request.FILES)
        if form.is_valid():
            staff = form.save()
            messages.success(request, 'Staff member created successfully')
            return redirect('employee_detail', pk=staff.id)
    else:
        form = OtherStaffForm()
    
    context = {
        'form': form,
        'title': 'Add Other Staff',
        'page_title': 'Add Other Staff'
    }
    return render(request, 'schools/hr/other_staff_form.html', context)

@login_required
def employee_update(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    
    if request.method == 'POST':
        form = EmployeeForm(request.POST, request.FILES, instance=employee)
        if form.is_valid():
            try:
                employee = form.save(commit=False)
                if not employee.date_joined:
                    employee.date_joined = timezone.now()
                employee.save()
                messages.success(request, 'Employee updated successfully')
                return redirect('employee_detail', pk=employee.id)
            except Exception as e:
                messages.error(request, f'Error updating employee: {str(e)}')
                return redirect('employee_update', pk=pk)
    else:
        form = EmployeeForm(instance=employee)
    
    context = {
        'form': form,
        'employee': employee,
        'title': f'Update Employee: {employee.get_full_name()}',
        'page_title': f'Update Employee: {employee.get_full_name()}'
    }
    return render(request, 'schools/hr/employee_form.html', context)

@login_required
def verify_payment(request, reference):
    """Verify payment by reference number"""
    payment = get_object_or_404(Payment.objects.select_related(
        'student',
        'recorded_by'
    ), pk=reference)
    
    # Generate QR code again for verification page
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=0,
    )
    
    verification_url = request.build_absolute_uri(
        f'/verify-payment/{payment.reference_number}/'
    )
    
    qr.add_data(verification_url)
    qr.make(fit=True)
    
    img_buffer = BytesIO()
    qr_image = qr.make_image(fill_color="#1e40af", back_color="white")
    qr_image.save(img_buffer, format='PNG')
    qr_code_data = base64.b64encode(img_buffer.getvalue()).decode()

    context = {
        'payment': payment,
        'qr_code': qr_code_data,
        'is_valid': True,
        'verification_date': timezone.now(),
        'title': 'Payment Verification',
        'page_title': 'Payment Verification'
    }
    return render(request, 'schools/payment_verify.html', context)

@login_required
def request_leave(request, teacher_id):
    teacher = get_object_or_404(Teacher, pk=teacher_id)
    
    if request.method == 'POST':
        try:
            leave = Leave.objects.create(
                employee=teacher,
                leave_type=request.POST.get('leave_type'),
                start_date=request.POST.get('start_date'),
                end_date=request.POST.get('end_date'),
                reason=request.POST.get('reason'),
                status='PENDING'
            )
            messages.success(request, 'Leave request submitted successfully')
        except Exception as e:
            messages.error(request, f'Error submitting leave request: {str(e)}')
    
    return redirect('teacher_detail', pk=teacher_id)


@login_required
def schedule_create(request, teacher_pk):
    teacher = get_object_or_404(Teacher, pk=teacher_pk)
    
    # Allow teachers to create their own schedules or admins to create for any teacher
    if not (request.user.is_superuser or 
            request.user.is_staff or 
            hasattr(request.user, 'teacher') and request.user.teacher == teacher):
        raise PermissionDenied
    
    if request.method == 'POST':
        form = ScheduleForm(request.POST)
        if form.is_valid():
            schedule = form.save(commit=False)
            schedule.teacher = teacher
            schedule.save()
            messages.success(request, 'Schedule created successfully')
            return redirect('teacher_detail', pk=teacher_pk)
    else:
        form = ScheduleForm()
    
    context = {
        'form': form,
        'teacher': teacher,
        'title': 'Add New Class',
        'page_title': 'Add New Class'
    }
    return render(request, 'schools/hr/schedule_form.html', context)


@login_required
def leave_request(request, teacher_id):
    teacher = get_object_or_404(Teacher, pk=teacher_id)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                leave = Leave.objects.create(
                    employee=teacher,
                    leave_type=request.POST['leave_type'],
                    start_date=request.POST['start_date'],
                    end_date=request.POST['end_date'],
                    reason=request.POST['reason'],
                    status='PENDING'
                )
                messages.success(request, 'Leave request submitted successfully')
                return redirect('teacher_detail', pk=teacher_id)
        except Exception as e:
            messages.error(request, f'Error submitting leave request: {str(e)}')
    
    return redirect('teacher_detail', pk=teacher_id)

@login_required
@admin_required
def schedule_edit(request, pk):
    schedule = get_object_or_404(Schedule, pk=pk)
    if request.method == 'POST':
        form = ScheduleForm(request.POST, instance=schedule)
        if form.is_valid():
            form.save()
            messages.success(request, 'Schedule updated successfully')
            return redirect('teacher_detail', pk=schedule.teacher.pk)
    else:
        form = ScheduleForm(instance=schedule)
    
    context = {
        'form': form,
        'schedule': schedule,
        'teacher': schedule.teacher,
        'title': 'Edit Schedule',
        'page_title': 'Edit Schedule'
    }
    return render(request, 'schools/hr/schedule_form.html', context)

@login_required
@admin_required
def schedule_delete(request, pk):
    schedule = get_object_or_404(Schedule, pk=pk)
    teacher_id = schedule.teacher.pk
    if request.method == 'POST':
        schedule.delete()
        messages.success(request, 'Schedule deleted successfully')
        return redirect('teacher_detail', pk=teacher_id)
    return redirect('teacher_detail', pk=teacher_id)

@login_required
def class_students(request, grade_id):
    grade = get_object_or_404(Grade, pk=grade_id)
    
    # Check if the user is authorized to view this class
    if not (request.user.is_superuser or 
            request.user.is_staff or 
            (hasattr(request.user, 'teacher') and 
             request.user.teacher.is_class_teacher and 
             request.user.teacher.grade_id == grade_id)):
        raise PermissionDenied
    
    students = Student.objects.filter(grade=grade).order_by('first_name', 'last_name')
    
    context = {
        'grade': grade,
        'students': students,
        'title': f'{grade.name} Students',
        'page_title': f'{grade.name} Students'
    }
    return render(request, 'schools/class_students.html', context)


@csrf_exempt
@require_http_methods(["GET"])
@login_required
def api_sms_list(request):
    """API endpoint to list SMS messages"""
    try:
        from config.models import SchoolConfig
        school = SchoolConfig.get_config(user=request.user, request=request)
        
        messages = SMSMessage.objects.filter(school=school).select_related(
            'sent_by', 'specific_grade', 'specific_student', 'specific_employee'
        ).order_by('-created_at')
        
        data = []
        for msg in messages:
            data.append({
                'id': msg.id,
                'message': msg.message,
                'recipient_type': msg.recipient_type,
                'recipient_type_display': msg.get_recipient_type_display(),
                'recipients_count': msg.recipients_count,
                'status': msg.status,
                'sent_at': msg.sent_at.strftime('%Y-%m-%d %H:%M') if msg.sent_at else None,
                'sent_by': msg.sent_by.get_full_name() if msg.sent_by else 'System',
                'created_at': msg.created_at.strftime('%Y-%m-%d %H:%M'),
            })
            
        return JsonResponse({'messages': data})
    except Exception as e:
        import traceback
        return JsonResponse({'error': str(e), 'traceback': traceback.format_exc()}, status=500)

@csrf_exempt
@require_http_methods(["POST"])
@login_required
def api_sms_send_bulk(request):
    """API endpoint to send bulk SMS messages with advanced filtering and personalization"""
    try:
        data = json.loads(request.body)
        from config.models import SchoolConfig
        school = SchoolConfig.get_config(user=request.user, request=request)
        
        # Helper to convert empty strings to None
        def clean_val(val, as_int=False):
            if val == '' or val is None:
                return None
            return int(val) if as_int else val

        sms = SMSMessage.objects.create(
            school=school,
            message=data.get('message'),
            recipient_type=data.get('recipient_type'),
            specific_grade_id=clean_val(data.get('specific_grade')),
            specific_student_id=clean_val(data.get('specific_student')),
            specific_employee_id=clean_val(data.get('specific_employee')),
            fee_min=data.get('fee_min') if data.get('fee_min') != '' else None,
            fee_max=data.get('fee_max') if data.get('fee_max') != '' else None,
            assessment_term=clean_val(data.get('assessment_term'), as_int=True),
            assessment_type=data.get('assessment_type'),
            sent_by=request.user
        )
        
        # Trigger sending (this now handles personalization)
        success, response = sms.send()
        
        if success:
            return JsonResponse({
                'success': True, 
                'message': 'SMS sent successfully',
                'recipients': sms.recipients_count,
                'response': response
            })
        else:
            return JsonResponse({
                'success': False, 
                'error': f'Failed to send SMS: {response}',
                'recipients': sms.recipients_count
            }, status=400)
            
    except Exception as e:
        import traceback
        return JsonResponse({'error': str(e), 'traceback': traceback.format_exc()}, status=500)

@csrf_exempt
@login_required
@require_http_methods(["GET"])
def api_communication_template_list(request):
    """API endpoint to list and get communication templates"""
    try:
        from config.models import SchoolConfig
        school = SchoolConfig.get_config(user=request.user, request=request)
        
        templates = CommunicationTemplate.objects.filter(Q(school=school) | Q(school__isnull=True))
        
        # If no templates exist, create some defaults for this school
        if not templates.exists():
            DEFAULT_TEMPLATES = [
                {
                    'name': 'Professional Fee Reminder',
                    'category': 'FEES',
                    'message': 'Dear {parent_name}, this is a reminder from {school_name} that {student_name} has a fee balance of {balance}. Please clear as soon as possible. Thank you.'
                },
                {
                    'name': 'Assessment Results (Compact)',
                    'category': 'RESULTS',
                    'message': 'Dear Parent, {student_name} {exam_type} {term} results: {results}. Keep it up!'
                },
                {
                    'name': 'Staff Meeting Alert',
                    'category': 'STAFF',
                    'message': 'Dear {staff_name}, there will be a staff meeting on Friday at 4pm. Agenda: Term Planning. Regard, Admin.'
                }
            ]
            for t in DEFAULT_TEMPLATES:
                CommunicationTemplate.objects.create(school=school, **t)
            templates = CommunicationTemplate.objects.filter(Q(school=school) | Q(school__isnull=True))

        data = [{
            'id': t.id,
            'name': t.name,
            'category': t.category,
            'category_display': t.get_category_display(),
            'message': t.message,
            'is_default': t.is_default,
        } for t in templates]
        
        return JsonResponse({'templates': data})
    except Exception as e:
        import traceback
        return JsonResponse({'error': str(e), 'traceback': traceback.format_exc()}, status=500)

@csrf_exempt
@login_required
def api_communication_template_create(request):
    """API endpoint to create a new communication template"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST method required'}, status=405)
        
    try:
        data = json.loads(request.body)
        from config.models import SchoolConfig
        school = SchoolConfig.get_config(user=request.user, request=request)
        
        template = CommunicationTemplate.objects.create(
            school=school,
            name=data.get('name'),
            category=data.get('category'),
            message=data.get('message')
        )
        
        return JsonResponse({
            'success': True,
            'template': {
                'id': template.id,
                'name': template.name,
                'category': template.category
            }
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@admin_required
def send_sms(request):
    if request.method == 'POST':
        form = SMSMessageForm(request.POST)
        if form.is_valid():
            sms = form.save(commit=False)
            sms.sent_by = request.user
            sms.save()
            
            # Send the SMS
            success, response = sms.send()
            
            if success:
                messages.success(request, 'SMS sent successfully')
            else:
                messages.error(request, f'Failed to send SMS: {response}')
                
            return redirect('sms_list')
    else:
        form = SMSMessageForm()
    
    context = {
        'form': form,
        'title': 'Send SMS',
        'page_title': 'Send SMS'
    }
    return render(request, 'schools/sms_form.html', context)

@login_required
@admin_required
def sms_list(request):
    sms_messages = SMSMessage.objects.select_related(
        'sent_by', 'specific_grade', 'specific_student'
    ).order_by('-created_at')
    
    context = {
        'sms_messages': sms_messages,
        'title': 'SMS Messages',
        'page_title': 'SMS Messages'
    }
    return render(request, 'schools/sms_list.html', context)

@login_required
@admin_required
def update_all_fees(request):
    # Playgroup fees
    Grade.objects.filter(name='PG').update(term_fees=5500)
    
    # PP1 & PP2 fees
    Grade.objects.filter(name__in=['PP1', 'PP2']).update(term_fees=6500)
    
    # Grade 1-3 fees
    Grade.objects.filter(name__in=['G1', 'G2', 'G3']).update(term_fees=7000)
    
    # Grade 4-12 fees
    Grade.objects.filter(
        name__in=['G4', 'G5', 'G6', 'G7', 'G8', 'G9', 'G10', 'G11', 'G12']
    ).update(term_fees=8000)
    
    messages.success(request, 'School fees updated successfully')
    return redirect('fees_dashboard')

@login_required
def update_student_term(request, student_id):
    if request.method == 'POST':
        student = get_object_or_404(Student, id=student_id)
        new_term = request.POST.get('current_term')
        
        if new_term in ['1', '2', '3']:
            student.current_term = int(new_term)
            student.save()
            messages.success(request, f'Term updated successfully to Term {new_term}')
        else:
            messages.error(request, 'Invalid term selection')
            
    return redirect('student_detail', pk=student_id)

@login_required
@admin_required
def auto_advance_academic(request):
    """View to trigger automatic term advancement and student promotion"""
    from django.core.management import call_command
    from io import StringIO
    import sys
    
    if request.method == 'POST':
        action = request.POST.get('action')
        dry_run = request.POST.get('dry_run') == 'on'
        
        # Capture command output
        old_stdout = sys.stdout
        sys.stdout = output = StringIO()
        
        try:
            if action == 'advance_term':
                call_command('advance_term', dry_run=dry_run)
            elif action == 'promote_students':
                call_command('promote_students', dry_run=dry_run)
            elif action == 'auto_advance':
                call_command('auto_advance_academic', dry_run=dry_run)
            else:
                messages.error(request, 'Invalid action')
                return redirect('dashboard')
            
            output_value = output.getvalue()
            sys.stdout = old_stdout
            
            if dry_run:
                messages.info(request, 'Dry run completed. Check output below.')
            else:
                messages.success(request, 'Academic progression completed successfully!')
            
            # Store output in session to display
            request.session['command_output'] = output_value
            
        except Exception as e:
            sys.stdout = old_stdout
            messages.error(request, f'Error: {str(e)}')
        
        return redirect('academic_progression')
    
    # Get current status
    current_term_obj = Term.objects.filter(is_current=True).first()
    if current_term_obj:
        current_term = current_term_obj.number
        current_year = current_term_obj.year
    else:
        sample_student = Student.objects.first()
        if sample_student:
            current_term = sample_student.current_term
        else:
            current_term = 1
        current_year = timezone.now().year
    
    # Get student statistics
    students_by_term = Student.objects.values('current_term').annotate(
        count=Count('id')
    ).order_by('current_term')
    
    students_by_grade = Student.objects.filter(
        grade__isnull=False
    ).values('grade__name').annotate(
        count=Count('id')
    ).order_by('grade__name')
    
    # Get command output if available
    command_output = request.session.pop('command_output', None)
    
    context = {
        'current_term': current_term,
        'current_year': current_year,
        'is_end_of_year': current_term == 3,
        'students_by_term': students_by_term,
        'students_by_grade': students_by_grade,
        'command_output': command_output,
    }
    
    return render(request, 'schools/academic_progression.html', context)

def generate_receipt_qr(data):
    """Generate QR code for receipt verification"""
    import qrcode
    import json
    
    # Convert data to JSON string
    json_data = json.dumps(data)
    
    # Create QR code instance
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    
    # Add data
    qr.add_data(json_data)
    qr.make(fit=True)
    
    # Create an image from the QR Code
    return qr.make_image(fill_color="black", back_color="white")

@login_required
def payment_receipt(request, payment_id):
    try:
        # Get the payment object with related student data
        payment = get_object_or_404(Payment.objects.select_related('student', 'student__grade'), id=payment_id)
        
        # Generate QR code for verification
        qr_data = {
            'payment_id': payment.id,
            'amount': str(payment.amount),
            'date': payment.date.strftime('%Y-%m-%d'),
            'reference': payment.reference_number
        }
        qr_code = generate_receipt_qr(qr_data)
        
        # Generate PDF receipt
        pdf_content = generate_payment_receipt(payment, qr_code)
        
        if not pdf_content:
            raise ValueError("PDF generation failed - no content generated")
        
        # Create the HTTP response with PDF content
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="receipt_{payment.reference_number}.pdf"'
        response['Content-Length'] = len(pdf_content)
        response.write(pdf_content)
        
        return response
        
    except Exception as e:
        error_msg = str(e)
        print(f"Error in payment_receipt: {error_msg}")
        import traceback
        print(traceback.format_exc())
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'error': 'Error generating PDF',
                'details': error_msg
            }, status=500)
        
        messages.error(request, f'Error generating PDF: {error_msg}')
        return redirect('payment_detail', pk=payment_id)


@login_required
def verify_payment(request, reference):
    payment = get_object_or_404(Payment, reference_number=reference)
    student = payment.student
    
    # Create QR data for verification
    qr_data = {
        'type': 'verification',
        'receipt_no': payment.reference_number,
        'amount': str(payment.amount),
        'student': student.admission_number,
        'date': payment.date.strftime('%Y-%m-%d')
    }
    
    qr_code_data = generate_receipt_qr(qr_data)
    
    context = {
        'payment': payment,
        'student': student,
        'qr_code': qr_code_data,
        'verification_date': timezone.now(),
        'title': 'Payment Verification',
        'page_title': 'Payment Verification'
    }
    return render(request, 'schools/payment_verify.html', context)

@login_required
@require_http_methods(["POST"])
def record_attendance(request):
    """Record attendance for a student via AJAX"""
    try:
        # Parse JSON data
        try:
            data = json.loads(request.body)
        except (json.JSONDecodeError, ValueError) as e:
            return JsonResponse({
                'success': False, 
                'error': 'Invalid JSON data',
                'message': str(e)
            }, status=400)
        
        student_id = data.get('student_id')
        status = data.get('status')
        date = data.get('date')
        
        # Validate required fields
        if not student_id:
            return JsonResponse({
                'success': False,
                'error': 'Student ID is required'
            }, status=400)
        
        if not status:
            return JsonResponse({
                'success': False,
                'error': 'Status is required'
            }, status=400)
        
        if not date:
            return JsonResponse({
                'success': False,
                'error': 'Date is required'
            }, status=400)
        
        # Validate status
        valid_statuses = ['PRESENT', 'ABSENT', 'LATE', 'EXCUSED']
        if status not in valid_statuses:
            return JsonResponse({
                'success': False,
                'error': f'Invalid status. Must be one of: {", ".join(valid_statuses)}'
            }, status=400)
        
        # Get student
        try:
            student = Student.objects.get(id=student_id)
        except Student.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Student not found'
            }, status=404)
        
        # Parse date
        try:
            from datetime import datetime
            attendance_date = datetime.strptime(date, '%Y-%m-%d').date()
        except (ValueError, TypeError) as e:
            return JsonResponse({
                'success': False,
                'error': 'Invalid date format. Use YYYY-MM-DD',
                'message': str(e)
            }, status=400)
        
        # Create or update attendance
        attendance, created = Attendance.objects.update_or_create(
            student=student,
            date=attendance_date,
            defaults={
                'status': status,
                'recorded_by': request.user
            }
        )
        
        return JsonResponse({
            'success': True, 
            'created': created,
            'message': f'Attendance marked as {status} successfully'
        })
        
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        return JsonResponse({
            'success': False, 
            'error': str(e),
            'message': 'An error occurred while recording attendance'
        }, status=500)

def attendance_events(request, student_id):
    student = get_object_or_404(Student, pk=student_id)
    attendance_records = Attendance.objects.filter(student=student)
    
    events = []
    for record in attendance_records:
        events.append({
            'title': record.get_status_display(),
            'start': record.date.isoformat(),
            'status': record.status,
            'recordedBy': record.recorded_by.get_full_name()
        })
    
    return JsonResponse(events, safe=False)


def index(request):
    """Landing page view"""
    if request.user.is_authenticated:
        # Check if user is a student
        if hasattr(request.user, 'student_profile'):
            return redirect('student_detail', pk=request.user.student_profile.pk)
        # Check if user is a teacher
        elif hasattr(request.user, 'teacher'):
            return redirect('teacher_detail', pk=request.user.teacher.pk)
        # For admin and other staff
        else:
            return redirect('dashboard')
    return render(request, 'schools/index.html')


def generate_payment_qr(payment):
    """Generate QR code with payment details"""
    # Create payment details string
    qr_data = (
        f"Receipt No: {payment.reference_number}\n"
        f"Student: {payment.student.get_full_name()}\n"
        f"Amount: KES {payment.amount}\n"
        f"Date: {payment.date.strftime('%Y-%m-%d')}\n"
        f"Term: {payment.term}\n"
        f"ID: {payment.id}"
    )
    
    # Generate QR code
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(qr_data)
    qr.make(fit=True)

    # Create QR image
    img = qr.make_image(fill_color="black", back_color="white")
    
    # Convert to base64
    buffer = BytesIO()
    img.save(buffer, format='PNG')
    return base64.b64encode(buffer.getvalue()).decode()

def payment_detail(request, pk):
    payment = get_object_or_404(Payment, pk=pk)
    
    # Generate QR code
    qr_code = generate_payment_qr(payment)
    
    context = {
        'payment': payment,
        'qr_code': qr_code,
    }
    return render(request, 'schools/payment_detail.html', context)

@login_required
def download_payment_pdf(request, pk):
    try:
        # Get the payment object with related student data
        payment = get_object_or_404(Payment.objects.select_related('student', 'student__grade'), pk=pk)
        
        # Generate QR code for verification
        qr_code = generate_payment_qr(payment)
        
        # Generate PDF receipt
        pdf_content = generate_payment_receipt(payment, qr_code)
        
        if not pdf_content:
            raise ValueError("PDF generation failed - no content generated")
        
        # Create the HTTP response with PDF content
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="receipt_{payment.reference_number}.pdf"'
        response.write(pdf_content)
        
        return response
        
    except Exception as e:
        error_msg = str(e)
        print(f"Error in download_payment_pdf: {error_msg}")
        import traceback
        print(traceback.format_exc())
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'error': 'Error generating PDF',
                'details': error_msg
            }, status=500)
        
        messages.error(request, f'Error generating PDF: {error_msg}')
        return redirect('payment_detail', pk=pk)

@csrf_exempt
@require_http_methods(["POST"])
def stk_push(request):
    try:
        data = json.loads(request.body)
        print(f"DEBUG: stk_push received data: {data}")
        student_id = data.get('student_id')
        phone = data.get('phone')
        amount = data.get('amount')
        
        if not phone:
            return JsonResponse({'error': 'M-Pesa phone number is required'}, status=400)
        if amount is None:
            return JsonResponse({'error': 'Amount is required'}, status=400)
            
        try:
            amount = float(amount)
        except (ValueError, TypeError):
            return JsonResponse({'error': 'Invalid amount format'}, status=400)

        account_ref = data.get('account_ref')
        description = data.get('description')
        is_subscription = data.get('is_subscription', False)
            
        from config.models import SchoolConfig
        school_config = SchoolConfig.get_config(request=request)
        # Generate a unique reference number for Paystack
        timestamp = timezone.now().strftime('%Y%m%d%H%M%S')
        
        if is_subscription:
            from config.models import SchoolConfig
            # For subscriptions, we must be absolutely sure we're upgrading the right school
            if request.user.is_authenticated:
                school_config = SchoolConfig.objects.filter(admin=request.user).first()
                if not school_config:
                    school_config = SchoolConfig.get_config(request=request)
            else:
                school_config = SchoolConfig.get_config(request=request)
                
            target_plan = data.get('target_plan', 'Enterprise')
            # Format: SUB_UP_{plan}_{school_id}_{timestamp}
            unique_ref = f"SUB_UP_{target_plan}_{school_config.id if school_config else '0'}_{timestamp}"
            email = school_config.school_email if (school_config and school_config.school_email) else "billing@edumanage.com"
            if not email or '@' not in email:
                email = "billing@edumanage.com"
            school_name = school_config.school_name if (school_config and school_config.school_name) else "EduManage"
        else:
            if not student_id:
                return JsonResponse({'error': 'Student ID is required for fee payments'}, status=400)
            try:
                student = Student.objects.get(id=student_id)
            except Student.DoesNotExist:
                return JsonResponse({'error': 'Student not found'}, status=404)
                
            # Format: STU_FEE_{student_id}_{timestamp}
            unique_ref = f"STU_FEE_{student_id}_{timestamp}"
            email = student.email if (student.email and '@' in student.email) else f"student_{student_id}@edumanage.com"
            school_name = student.school.school_name if student.school else "EduManage"

        # Initiate STK push via Paystack
        paystack_ref, error = initiate_paystack_stk(
            phone=phone,
            amount=amount,
            email=email,
            reference=unique_ref,
            school_name=school_name
        )
        
        if error:
            # If error is a dict (raw response), extract message
            error_msg = error
            if isinstance(error, dict):
                error_msg = error.get('message', str(error))
            return JsonResponse({
                'error': str(error_msg),
                'success': False
            }, status=400)
            
        if not is_subscription:
            # Create pending payment record for students
            payment = Payment.objects.create(
                student=student,
                amount=amount,
                payment_method='MPESA', # Keep as MPESA for UI labeling
                phone_number=phone,
                status='PENDING',
                checkout_request_id=paystack_ref, # Store paystack ref here
                reference_number=unique_ref,
                term=student.current_term,
                school=student.school
            )
        else:
            print(f"DEBUG: Paystack Subscription payment initiated for {school_config.school_name}")
        
        return JsonResponse({
            'success': True,
            'message': 'Payment request sent! Please check your phone for a PIN prompt.',
            'reference': paystack_ref
        })
        
    except Student.DoesNotExist:
        return JsonResponse({'error': 'Student not found'}, status=404)
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["POST"])
def paystack_webhook(request):
    """
    Handle Paystack Webhooks
    """
    import hmac
    import hashlib
    
    # Verify Paystack Signature
    signature = request.headers.get('x-paystack-signature')
    if not signature:
        return HttpResponse(status=400)
        
    payload = request.body
    secret = settings.PAYSTACK_SECRET_KEY.encode('utf-8')
    hash = hmac.new(secret, payload, hashlib.sha512).hexdigest()
    
    if hash != signature:
        return HttpResponse(status=401)
        
    data = json.loads(payload)
    event = data.get('event')
    
    if event == 'charge.success':
        tx_data = data.get('data', {})
        reference = tx_data.get('reference', '')
        
        print(f"DEBUG: Paystack Webhook Success - Ref: {reference}")
        
        # 1. Handle Subscription Upgrade
        if reference.startswith('SUB_UP_'):
            try:
                # Extract plan and school ID: SUB_UP_{plan}_{school_id}_{timestamp}
                parts = reference.split('_')
                if len(parts) >= 4:
                    target_plan = parts[2]
                    school_id = parts[3]
                    from config.models import SchoolConfig, SubscriptionPayment
                    from django.utils import timezone
                    from datetime import timedelta
                    from decimal import Decimal
                    
                    config = SchoolConfig.objects.get(id=school_id)
                    config.subscription_status = 'Active'
                    config.subscription_plan = target_plan
                    # Reset trial/Extend dates
                    config.trial_end_date = timezone.now() + timedelta(days=31) # 1 month sub
                    config.save()
                    
                    # Log the payment for Super Admin revenue tracking
                    amount_paid = Decimal(str(tx_data.get('amount', 0))) / 100 # Paystack amounts are in kobo/cents
                    SubscriptionPayment.objects.update_or_create(
                        reference=reference,
                        defaults={
                            'school': config,
                            'amount': amount_paid,
                            'plan': target_plan,
                            'status': 'COMPLETED',
                            'transaction_id': tx_data.get('id')
                        }
                    )
                    
                    print(f"DEBUG: School '{config.school_name}' upgraded to {target_plan} successfully via Paystack Webhook! Amount: {amount_paid}")
            except Exception as e:
                print(f"ERROR Upgrading School (Webhook): {str(e)}")
                
        # 2. Handle Student Fee Payment
        elif reference.startswith('STU_FEE_'):
            try:
                payment = Payment.objects.get(reference_number=reference)
                payment.status = 'COMPLETED'
                payment.transaction_id = tx_data.get('id')
                payment.save()
                print(f"DEBUG: Student Payment {reference} marked as COMPLETED")
            except Exception as e:
                 print(f"ERROR Updating Payment: {str(e)}")
                 
    return HttpResponse(status=200)


@csrf_exempt
def mpesa_callback(request):
    try:
        data = json.loads(request.body)
        result = data.get('Body', {}).get('stkCallback', {})
        checkout_id = result.get('CheckoutRequestID')
        
        payment = Payment.objects.get(checkout_request_id=checkout_id)
        
        if result.get('ResultCode') == 0:
            # Payment successful
            payment.status = 'COMPLETED'
            payment.transaction_id = result.get('CallbackMetadata', {}).get('Item', [{}])[1].get('Value')
            payment.save()
            return JsonResponse({'message': 'success'})
        else:
            # Payment failed
            payment.status = 'FAILED'
            payment.save()
            return JsonResponse({'error': 'Payment failed'}, status=400)
            
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

# Transport Views
@login_required
@admin_required
def transport_dashboard(request):
    """Transport management dashboard"""
    from django.db.models import Count, Sum, Q
    from django.utils import timezone
    from datetime import timedelta
    import json
    
    vehicles = Vehicle.objects.all()
    routes = Route.objects.filter(is_active=True)
    assignments = StudentTransportAssignment.objects.filter(is_active=True)
    transport_fees = TransportFee.objects.all()
    
    # Basic Statistics
    total_vehicles = vehicles.count()
    active_vehicles = vehicles.filter(status='ACTIVE').count()
    inactive_vehicles = vehicles.filter(status='INACTIVE').count()
    maintenance_vehicles = vehicles.filter(status='MAINTENANCE').count()
    total_routes = routes.count()
    total_students = assignments.count()
    total_capacity = vehicles.aggregate(total=Sum('capacity'))['total'] or 0
    used_capacity = sum(v.get_current_students_count() for v in vehicles)
    capacity_percentage = (used_capacity / total_capacity * 100) if total_capacity > 0 else 0
    
    # Revenue Statistics
    total_revenue = transport_fees.filter(status='COMPLETED').aggregate(total=Sum('amount'))['total'] or 0
    pending_payments = transport_fees.filter(status='PENDING').aggregate(total=Sum('amount'))['total'] or 0
    completed_payments = transport_fees.filter(status='COMPLETED').count()
    pending_payments_count = transport_fees.filter(status='PENDING').count()
    failed_payments_count = transport_fees.filter(status='FAILED').count()
    
    # Vehicle Status Distribution (for pie chart)
    vehicle_status_data = {
        'ACTIVE': vehicles.filter(status='ACTIVE').count(),
        'INACTIVE': vehicles.filter(status='INACTIVE').count(),
        'MAINTENANCE': vehicles.filter(status='MAINTENANCE').count(),
    }
    
    # Vehicle Type Distribution (for bar chart)
    vehicle_type_data = {}
    for vtype, vname in Vehicle.VEHICLE_TYPES:
        vehicle_type_data[vname] = vehicles.filter(vehicle_type=vtype).count()
    
    # Route Popularity (top routes by student count)
    route_popularity = []
    for route in routes:
        student_count = route.get_students_count()
        revenue = transport_fees.filter(route=route, status='COMPLETED').aggregate(total=Sum('amount'))['total'] or 0
        route_popularity.append({
            'name': route.name,
            'students': student_count,
            'revenue': float(revenue)
        })
    route_popularity = sorted(route_popularity, key=lambda x: x['students'], reverse=True)[:10]
    
    # Payment Status Distribution
    payment_status_data = {
        'COMPLETED': completed_payments,
        'PENDING': pending_payments_count,
        'FAILED': failed_payments_count,
    }
    
    # Monthly Revenue (last 6 months)
    monthly_revenue = []
    for i in range(5, -1, -1):
        month_start = timezone.now().replace(day=1) - timedelta(days=30 * i)
        month_end = (month_start + timedelta(days=30)).replace(day=1) - timedelta(days=1)
        revenue = transport_fees.filter(
            date__gte=month_start,
            date__lte=month_end,
            status='COMPLETED'
        ).aggregate(total=Sum('amount'))['total'] or 0
        monthly_revenue.append({
            'month': month_start.strftime('%b %Y'),
            'revenue': float(revenue)
        })
    
    # Vehicle Capacity Utilization
    vehicle_utilization = []
    for vehicle in vehicles.filter(status='ACTIVE')[:10]:
        used = vehicle.get_current_students_count()
        util_percent = (used / vehicle.capacity * 100) if vehicle.capacity > 0 else 0
        vehicle_utilization.append({
            'name': vehicle.vehicle_number,
            'used': used,
            'capacity': vehicle.capacity,
            'percentage': round(util_percent, 1)
        })
    vehicle_utilization = sorted(vehicle_utilization, key=lambda x: x['percentage'], reverse=True)
    
    # Payment Method Distribution
    payment_method_data = {}
    for method, method_name in TransportFee.PAYMENT_METHODS:
        payment_method_data[method_name] = transport_fees.filter(
            payment_method=method,
            status='COMPLETED'
        ).count()
    
    stats = {
        'total_vehicles': total_vehicles,
        'active_vehicles': active_vehicles,
        'inactive_vehicles': inactive_vehicles,
        'maintenance_vehicles': maintenance_vehicles,
        'total_routes': total_routes,
        'total_students': total_students,
        'total_capacity': total_capacity,
        'used_capacity': used_capacity,
        'capacity_percentage': round(capacity_percentage, 1),
        'total_revenue': total_revenue,
        'pending_payments': pending_payments,
        'completed_payments': completed_payments,
        'pending_payments_count': pending_payments_count,
    }
    
    context = {
        'vehicles': vehicles[:5],
        'routes': routes[:5],
        'assignments': assignments.select_related('student', 'route', 'vehicle')[:10],
        'stats': stats,
        'vehicle_status_data': json.dumps(vehicle_status_data),
        'vehicle_type_data': json.dumps(vehicle_type_data),
        'route_popularity': json.dumps(route_popularity),
        'payment_status_data': json.dumps(payment_status_data),
        'monthly_revenue': json.dumps(monthly_revenue),
        'vehicle_utilization': json.dumps(vehicle_utilization),
        'payment_method_data': json.dumps(payment_method_data),
        'title': 'Transport Dashboard',
        'page_title': 'Transport Management'
    }
    return render(request, 'schools/transport/dashboard.html', context)

@login_required
@admin_required
def vehicle_list(request):
    """List all vehicles"""
    from django.db.models import Count, Q
    import json
    
    vehicles = Vehicle.objects.select_related('driver').all()
    
    # Filtering
    status_filter = request.GET.get('status', '')
    type_filter = request.GET.get('type', '')
    location_filter = request.GET.get('location', '')
    search_query = request.GET.get('search', '')
    
    if status_filter:
        vehicles = vehicles.filter(status=status_filter)
    if type_filter:
        vehicles = vehicles.filter(vehicle_type=type_filter)
    if location_filter:
        vehicles = vehicles.filter(location=location_filter)
    if search_query:
        vehicles = vehicles.filter(
            Q(vehicle_number__icontains=search_query) |
            Q(make__icontains=search_query) |
            Q(model__icontains=search_query) |
            Q(driver__first_name__icontains=search_query) |
            Q(driver__last_name__icontains=search_query)
        )
    
    # Statistics
    total_vehicles = Vehicle.objects.count()
    active_vehicles = Vehicle.objects.filter(status='ACTIVE').count()
    maintenance_vehicles = Vehicle.objects.filter(status='MAINTENANCE').count()
    inactive_vehicles = Vehicle.objects.filter(status='INACTIVE').count()
    
    # Capacity statistics
    total_capacity = vehicles.aggregate(total=Sum('capacity'))['total'] or 0
    total_used = sum(v.get_current_students_count() for v in vehicles)
    avg_utilization = (total_used / total_capacity * 100) if total_capacity > 0 else 0
    
    # Vehicle type distribution
    type_distribution = {}
    for vtype, vname in Vehicle.VEHICLE_TYPES:
        type_distribution[vname] = Vehicle.objects.filter(vehicle_type=vtype).count()
    
    # Status distribution
    status_distribution = {
        'ACTIVE': active_vehicles,
        'MAINTENANCE': maintenance_vehicles,
        'INACTIVE': inactive_vehicles,
    }
    
    context = {
        'vehicles': vehicles,
        'total_vehicles': total_vehicles,
        'active_vehicles': active_vehicles,
        'maintenance_vehicles': maintenance_vehicles,
        'inactive_vehicles': inactive_vehicles,
        'total_capacity': total_capacity,
        'total_used': total_used,
        'avg_utilization': round(avg_utilization, 1),
        'type_distribution': json.dumps(type_distribution),
        'status_distribution': json.dumps(status_distribution),
        'status_filter': status_filter,
        'type_filter': type_filter,
        'location_filter': location_filter,
        'search_query': search_query,
        'title': 'Vehicles',
        'page_title': 'Vehicle Management'
    }
    return render(request, 'schools/transport/vehicle_list.html', context)

@login_required
@admin_required
def vehicle_create(request):
    """Create a new vehicle"""
    if request.method == 'POST':
        form = VehicleForm(request.POST)
        if form.is_valid():
            vehicle = form.save()
            messages.success(request, f'Vehicle {vehicle.vehicle_number} created successfully')
            return redirect('vehicle_detail', pk=vehicle.pk)
    else:
        form = VehicleForm()
    
    context = {
        'form': form,
        'title': 'Add Vehicle',
        'page_title': 'Add New Vehicle'
    }
    return render(request, 'schools/transport/vehicle_form.html', context)

@login_required
@admin_required
def vehicle_detail(request, pk):
    """Vehicle detail view"""
    vehicle = get_object_or_404(Vehicle, pk=pk)
    assignments = vehicle.student_assignments.filter(is_active=True).select_related('student', 'route')
    
    context = {
        'vehicle': vehicle,
        'assignments': assignments,
        'title': f'Vehicle: {vehicle.vehicle_number}',
        'page_title': f'Vehicle Details - {vehicle.vehicle_number}'
    }
    return render(request, 'schools/transport/vehicle_detail.html', context)

@login_required
@admin_required
def vehicle_edit(request, pk):
    """Edit vehicle"""
    vehicle = get_object_or_404(Vehicle, pk=pk)
    if request.method == 'POST':
        form = VehicleForm(request.POST, instance=vehicle)
        if form.is_valid():
            form.save()
            messages.success(request, 'Vehicle updated successfully')
            return redirect('vehicle_detail', pk=vehicle.pk)
    else:
        form = VehicleForm(instance=vehicle)
    
    context = {
        'form': form,
        'vehicle': vehicle,
        'title': f'Edit Vehicle: {vehicle.vehicle_number}',
        'page_title': f'Edit Vehicle - {vehicle.vehicle_number}'
    }
    return render(request, 'schools/transport/vehicle_form.html', context)

@login_required
@admin_required
def route_list(request):
    """List all routes"""
    from django.db.models import Q, Sum, Count
    
    routes = Route.objects.prefetch_related('student_assignments').all()
    
    # Filtering
    status_filter = request.GET.get('status', '')
    location_filter = request.GET.get('location', '')
    search_query = request.GET.get('search', '')
    
    if status_filter == 'active':
        routes = routes.filter(is_active=True)
    elif status_filter == 'inactive':
        routes = routes.filter(is_active=False)
    
    if location_filter:
        routes = routes.filter(location=location_filter)
    
    if search_query:
        routes = routes.filter(
            Q(name__icontains=search_query) |
            Q(start_location__icontains=search_query) |
            Q(end_location__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    # Statistics
    total_routes = Route.objects.count()
    active_routes = Route.objects.filter(is_active=True).count()
    inactive_routes = Route.objects.filter(is_active=False).count()
    
    # Calculate total students and revenue from filtered routes
    total_students = 0
    total_revenue = 0
    for route in routes:
        student_count = route.get_students_count()
        total_students += student_count
        total_revenue += float(student_count * route.fee_per_term)
    
    context = {
        'routes': routes,
        'total_routes': total_routes,
        'active_routes': active_routes,
        'inactive_routes': inactive_routes,
        'total_students': total_students,
        'total_revenue': total_revenue,
        'status_filter': status_filter,
        'location_filter': location_filter,
        'search_query': search_query,
        'title': 'Routes',
        'page_title': 'Route Management'
    }
    return render(request, 'schools/transport/route_list.html', context)

@login_required
@admin_required
def route_create(request):
    """Create a new route"""
    from schools.utils.maps import geocode_address
    
    if request.method == 'POST':
        form = RouteForm(request.POST)
        if form.is_valid():
            route = form.save(commit=False)
            
            # Geocode locations if coordinates not provided
            if not route.start_latitude and form.cleaned_data.get('start_location'):
                lat, lng = geocode_address(form.cleaned_data['start_location'])
                if lat and lng:
                    route.start_latitude = lat
                    route.start_longitude = lng
            
            if not route.end_latitude and form.cleaned_data.get('end_location'):
                lat, lng = geocode_address(form.cleaned_data['end_location'])
                if lat and lng:
                    route.end_latitude = lat
                    route.end_longitude = lng
            
            # If school coordinates not set, use start location as default
            if not route.school_latitude:
                route.school_latitude = route.start_latitude
                route.school_longitude = route.start_longitude
            
            route.save()
            messages.success(request, f'Route {route.name} created successfully')
            return redirect('route_detail', pk=route.pk)
    else:
        form = RouteForm()
    
    context = {
        'form': form,
        'title': 'Add Route',
        'page_title': 'Add New Route',
    }
    return render(request, 'schools/transport/route_form.html', context)

@login_required
@admin_required
def route_detail(request, pk):
    """Route detail view with map"""
    route = get_object_or_404(Route, pk=pk)
    assignments = route.student_assignments.filter(is_active=True).select_related('student', 'vehicle')
    
    # Prepare pickup locations for map
    pickup_locations = []
    for assignment in assignments:
        if assignment.pickup_location and assignment.pickup_latitude and assignment.pickup_longitude:
            pickup_locations.append({
                'student': assignment.student.get_full_name(),
                'location': assignment.pickup_location,
                'lat': float(assignment.pickup_latitude),
                'lng': float(assignment.pickup_longitude),
                'time': assignment.pickup_time.strftime('%I:%M %p') if assignment.pickup_time else 'Not set',
                'admission_number': assignment.student.admission_number,
            })
    
    # School location (default to route start if school coordinates not set)
    school_lat = float(route.school_latitude) if route.school_latitude else (float(route.start_latitude) if route.start_latitude else None)
    school_lng = float(route.school_longitude) if route.school_longitude else (float(route.start_longitude) if route.start_longitude else None)
    
    context = {
        'route': route,
        'assignments': assignments,
        'pickup_locations': pickup_locations,
        'school_lat': school_lat,
        'school_lng': school_lng,
        'title': f'Route: {route.name}',
        'page_title': f'Route Details - {route.name}',
    }
    return render(request, 'schools/transport/route_detail.html', context)

@login_required
@admin_required
def route_edit(request, pk):
    """Edit route"""
    from schools.utils.maps import geocode_address
    
    route = get_object_or_404(Route, pk=pk)
    if request.method == 'POST':
        form = RouteForm(request.POST, instance=route)
        if form.is_valid():
            route = form.save(commit=False)
            
            # Geocode locations if coordinates not provided and address changed
            if not route.start_latitude and form.cleaned_data.get('start_location'):
                lat, lng = geocode_address(form.cleaned_data['start_location'])
                if lat and lng:
                    route.start_latitude = lat
                    route.start_longitude = lng
            
            if not route.end_latitude and form.cleaned_data.get('end_location'):
                lat, lng = geocode_address(form.cleaned_data['end_location'])
                if lat and lng:
                    route.end_latitude = lat
                    route.end_longitude = lng
            
            # If school coordinates not set, use start location as default
            if not route.school_latitude:
                route.school_latitude = route.start_latitude
                route.school_longitude = route.start_longitude
            
            route.save()
            messages.success(request, 'Route updated successfully')
            return redirect('route_detail', pk=route.pk)
    else:
        form = RouteForm(instance=route)
    
    context = {
        'form': form,
        'route': route,
        'title': f'Edit Route: {route.name}',
        'page_title': f'Edit Route - {route.name}',
    }
    return render(request, 'schools/transport/route_form.html', context)

@login_required
@admin_required
def transport_assignment_list(request):
    """List all transport assignments"""
    from django.db.models import Q
    
    assignments = StudentTransportAssignment.objects.select_related(
        'student', 'route', 'vehicle'
    ).prefetch_related('student__grade').all()
    
    # Filtering
    status_filter = request.GET.get('status', '')
    location_filter = request.GET.get('location', '')
    search_query = request.GET.get('search', '')
    
    if status_filter == 'active':
        assignments = assignments.filter(is_active=True)
    elif status_filter == 'inactive':
        assignments = assignments.filter(is_active=False)
    
    if location_filter:
        assignments = assignments.filter(
            Q(student__location=location_filter) |
            Q(route__location=location_filter)
        )
    
    if search_query:
        assignments = assignments.filter(
            Q(student__first_name__icontains=search_query) |
            Q(student__last_name__icontains=search_query) |
            Q(student__admission_number__icontains=search_query) |
            Q(route__name__icontains=search_query) |
            Q(vehicle__vehicle_number__icontains=search_query) |
            Q(pickup_location__icontains=search_query) |
            Q(dropoff_location__icontains=search_query)
        )
    
    # Statistics
    total_assignments = StudentTransportAssignment.objects.count()
    active_assignments = StudentTransportAssignment.objects.filter(is_active=True).count()
    inactive_assignments = StudentTransportAssignment.objects.filter(is_active=False).count()
    assigned_vehicles = StudentTransportAssignment.objects.filter(vehicle__isnull=False, is_active=True).count()
    unassigned_vehicles = StudentTransportAssignment.objects.filter(vehicle__isnull=True, is_active=True).count()
    
    context = {
        'assignments': assignments,
        'total_assignments': total_assignments,
        'active_assignments': active_assignments,
        'inactive_assignments': inactive_assignments,
        'assigned_vehicles': assigned_vehicles,
        'unassigned_vehicles': unassigned_vehicles,
        'status_filter': status_filter,
        'location_filter': location_filter,
        'search_query': search_query,
        'title': 'Transport Assignments',
        'page_title': 'Student Transport Assignments'
    }
    return render(request, 'schools/transport/assignment_list.html', context)

@login_required
@admin_required
def transport_assignment_create(request):
    """Create a new transport assignment"""
    from schools.utils.maps import geocode_address
    
    if request.method == 'POST':
        form = StudentTransportAssignmentForm(request.POST)
        if form.is_valid():
            try:
                assignment = form.save(commit=False)
                
                # Geocode pickup location if coordinates not provided
                if not assignment.pickup_latitude and form.cleaned_data.get('pickup_location'):
                    lat, lng = geocode_address(form.cleaned_data['pickup_location'])
                    if lat and lng:
                        assignment.pickup_latitude = lat
                        assignment.pickup_longitude = lng
                
                # Geocode dropoff location if coordinates not provided
                if not assignment.dropoff_latitude and form.cleaned_data.get('dropoff_location'):
                    lat, lng = geocode_address(form.cleaned_data['dropoff_location'])
                    if lat and lng:
                        assignment.dropoff_latitude = lat
                        assignment.dropoff_longitude = lng
                
                assignment.save()
                messages.success(request, f'Transport assignment created for {assignment.student.get_full_name()}')
                return redirect('transport_assignment_detail', pk=assignment.pk)
            except ValidationError as e:
                messages.error(request, str(e))
    else:
        form = StudentTransportAssignmentForm()
    
    context = {
        'form': form,
        'title': 'Assign Transport',
        'page_title': 'Assign Student to Transport',
    }
    return render(request, 'schools/transport/assignment_form.html', context)

@login_required
@admin_required
def transport_assignment_detail(request, pk):
    """Transport assignment detail view"""
    assignment = get_object_or_404(StudentTransportAssignment, pk=pk)
    transport_fees = TransportFee.objects.filter(
        student=assignment.student,
        route=assignment.route
    ).order_by('-date')
    
    # Get route school location for map
    route = assignment.route
    route_school_lat = float(route.school_latitude) if route.school_latitude else (float(route.start_latitude) if route.start_latitude else None)
    route_school_lng = float(route.school_longitude) if route.school_longitude else (float(route.start_longitude) if route.start_longitude else None)
    
    context = {
        'assignment': assignment,
        'transport_fees': transport_fees,
        'route_school_lat': route_school_lat,
        'route_school_lng': route_school_lng,
        'title': f'Transport Assignment: {assignment.student.get_full_name()}',
        'page_title': f'Transport Assignment Details'
    }
    return render(request, 'schools/transport/assignment_detail.html', context)

@login_required
@admin_required
def transport_assignment_edit(request, pk):
    """Edit transport assignment"""
    from schools.utils.maps import geocode_address
    
    assignment = get_object_or_404(StudentTransportAssignment, pk=pk)
    if request.method == 'POST':
        form = StudentTransportAssignmentForm(request.POST, instance=assignment)
        if form.is_valid():
            try:
                assignment = form.save(commit=False)
                
                # Geocode pickup location if coordinates not provided
                if not assignment.pickup_latitude and form.cleaned_data.get('pickup_location'):
                    lat, lng = geocode_address(form.cleaned_data['pickup_location'])
                    if lat and lng:
                        assignment.pickup_latitude = lat
                        assignment.pickup_longitude = lng
                
                # Geocode dropoff location if coordinates not provided
                if not assignment.dropoff_latitude and form.cleaned_data.get('dropoff_location'):
                    lat, lng = geocode_address(form.cleaned_data['dropoff_location'])
                    if lat and lng:
                        assignment.dropoff_latitude = lat
                        assignment.dropoff_longitude = lng
                
                assignment.save()
                messages.success(request, 'Transport assignment updated successfully')
                return redirect('transport_assignment_detail', pk=assignment.pk)
            except ValidationError as e:
                messages.error(request, str(e))
    else:
        form = StudentTransportAssignmentForm(instance=assignment)
    
    context = {
        'form': form,
        'assignment': assignment,
        'title': f'Edit Transport Assignment',
        'page_title': f'Edit Transport Assignment',
    }
    return render(request, 'schools/transport/assignment_form.html', context)

@login_required
@admin_required
def transport_fee_list(request):
    """List all transport fees with analytics"""
    from django.db.models import Q, Sum, Count
    from django.utils import timezone
    from datetime import timedelta
    import json
    
    fees = TransportFee.objects.select_related('student', 'route').all()
    
    # Filtering
    term_filter = request.GET.get('term', '')
    status_filter = request.GET.get('status', '')
    search_query = request.GET.get('search', '')
    
    if term_filter:
        fees = fees.filter(term=term_filter)
    
    if status_filter:
        fees = fees.filter(status=status_filter)
    
    if search_query:
        fees = fees.filter(
            Q(student__first_name__icontains=search_query) |
            Q(student__last_name__icontains=search_query) |
            Q(student__admission_number__icontains=search_query) |
            Q(route__name__icontains=search_query) |
            Q(reference_number__icontains=search_query) |
            Q(transaction_id__icontains=search_query)
        )
    
    # Statistics
    total_fees = TransportFee.objects.count()
    completed_fees = TransportFee.objects.filter(status='COMPLETED').count()
    pending_fees = TransportFee.objects.filter(status='PENDING').count()
    failed_fees = TransportFee.objects.filter(status='FAILED').count()
    
    # Revenue Statistics
    total_revenue = TransportFee.objects.filter(status='COMPLETED').aggregate(total=Sum('amount'))['total'] or 0
    pending_revenue = TransportFee.objects.filter(status='PENDING').aggregate(total=Sum('amount'))['total'] or 0
    failed_revenue = TransportFee.objects.filter(status='FAILED').aggregate(total=Sum('amount'))['total'] or 0
    
    # Payment Status Distribution
    payment_status_data = {
        'COMPLETED': completed_fees,
        'PENDING': pending_fees,
        'FAILED': failed_fees
    }
    
    # Payment Method Distribution
    payment_method_data = {}
    for method, method_name in TransportFee.PAYMENT_METHODS:
        count = TransportFee.objects.filter(payment_method=method).count()
        if count > 0:
            payment_method_data[method_name] = count
    
    # Monthly Revenue (last 6 months)
    monthly_revenue = []
    for i in range(5, -1, -1):
        month_start = timezone.now().replace(day=1) - timedelta(days=30 * i)
        month_end = (month_start + timedelta(days=30)).replace(day=1) - timedelta(days=1)
        revenue = TransportFee.objects.filter(
            date__gte=month_start,
            date__lte=month_end,
            status='COMPLETED'
        ).aggregate(total=Sum('amount'))['total'] or 0
        monthly_revenue.append({
            'month': month_start.strftime('%b %Y'),
            'revenue': float(revenue)
        })
    
    # Revenue by Term
    term_revenue = []
    for term_num in [1, 2, 3]:
        revenue = TransportFee.objects.filter(
            term=term_num,
            status='COMPLETED'
        ).aggregate(total=Sum('amount'))['total'] or 0
        term_revenue.append({
            'term': f'Term {term_num}',
            'revenue': float(revenue)
        })
    
    # Revenue by Route (top 10)
    route_revenue = []
    routes = Route.objects.all()
    for route in routes:
        revenue = TransportFee.objects.filter(
            route=route,
            status='COMPLETED'
        ).aggregate(total=Sum('amount'))['total'] or 0
        if revenue > 0:
            route_revenue.append({
                'name': route.name,
                'revenue': float(revenue)
            })
    route_revenue = sorted(route_revenue, key=lambda x: x['revenue'], reverse=True)[:10]
    
    # Daily Revenue (last 30 days)
    daily_revenue = []
    for i in range(29, -1, -1):
        day = timezone.now().date() - timedelta(days=i)
        revenue = TransportFee.objects.filter(
            date__date=day,
            status='COMPLETED'
        ).aggregate(total=Sum('amount'))['total'] or 0
        daily_revenue.append({
            'day': day.strftime('%b %d'),
            'revenue': float(revenue)
        })
    
    context = {
        'fees': fees,
        'total_fees': total_fees,
        'completed_fees': completed_fees,
        'pending_fees': pending_fees,
        'failed_fees': failed_fees,
        'total_revenue': total_revenue,
        'pending_revenue': pending_revenue,
        'failed_revenue': failed_revenue,
        'payment_status_data': json.dumps(payment_status_data),
        'payment_method_data': json.dumps(payment_method_data),
        'monthly_revenue': json.dumps(monthly_revenue),
        'term_revenue': json.dumps(term_revenue),
        'route_revenue': json.dumps(route_revenue),
        'daily_revenue': json.dumps(daily_revenue),
        'term_filter': term_filter,
        'status_filter': status_filter,
        'search_query': search_query,
        'title': 'Transport Fees',
        'page_title': 'Transport Fee Management'
    }
    return render(request, 'schools/transport/fee_list.html', context)

@login_required
@admin_required
def get_student_transport_info(request, student_id):
    """API endpoint to get student transport information"""
    from django.http import JsonResponse
    from django.db.models import Sum, Q
    
    try:
        student = Student.objects.get(pk=student_id)
        
        # Get active transport assignment
        assignment = StudentTransportAssignment.objects.filter(
            student=student,
            is_active=True
        ).select_related('route').first()
        
        if not assignment:
            return JsonResponse({
                'error': 'No active transport assignment found for this student',
                'has_assignment': False
            })
        
        route = assignment.route
        
        # Get term from request or use current term
        term = request.GET.get('term', get_current_term())
        term = int(term) if term else get_current_term()
        
        # Calculate transport fees for this term
        term_fee = route.fee_per_term
        
        # Get total paid for this term
        total_paid = TransportFee.objects.filter(
            student=student,
            route=route,
            term=term,
            status='COMPLETED'
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        # Calculate balance
        balance = float(term_fee) - float(total_paid)
        
        # Get payment history for this term
        payments = TransportFee.objects.filter(
            student=student,
            route=route,
            term=term
        ).order_by('-date')
        
        payment_history = []
        for payment in payments:
            payment_history.append({
                'date': payment.date.strftime('%b %d, %Y'),
                'amount': float(payment.amount),
                'status': payment.get_status_display(),
                'method': payment.get_payment_method_display()
            })
        
        return JsonResponse({
            'has_assignment': True,
            'route_id': route.id,
            'route_name': route.name,
            'route_fee': float(term_fee),
            'total_paid': float(total_paid),
            'balance': balance,
            'term': term,
            'payment_history': payment_history
        })
    except Student.DoesNotExist:
        return JsonResponse({'error': 'Student not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@admin_required
def transport_fee_create(request):
    """Create a new transport fee payment"""
    if request.method == 'POST':
        form = TransportFeeForm(request.POST)
        if form.is_valid():
            fee = form.save()
            
            # Automatically send receipt to parent
            try:
                from .utils import send_transport_fee_receipt
                import logging
                logger = logging.getLogger(__name__)
                success, result = send_transport_fee_receipt(fee)
                if success:
                    if isinstance(result, dict):
                        if result.get('email_sent'):
                            messages.info(request, 'Receipt sent to parent email')
                        if result.get('whatsapp_sent'):
                            messages.info(request, 'Receipt notification sent to parent WhatsApp')
                else:
                    # Log error but don't fail the payment recording
                    logger.warning(f"Receipt sending failed: {result}")
            except Exception as e:
                # Log error but don't fail the payment recording
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error sending receipt: {str(e)}")
            
            messages.success(request, f'Transport fee payment recorded for {fee.student.get_full_name()}')
            return redirect('transport_fee_detail', pk=fee.pk)
    else:
        form = TransportFeeForm()
    
    context = {
        'form': form,
        'title': 'Record Transport Fee',
        'page_title': 'Record Transport Fee Payment'
    }
    return render(request, 'schools/transport/fee_form.html', context)

@login_required
@admin_required
def transport_fee_detail(request, pk):
    """Transport fee detail view"""
    fee = get_object_or_404(TransportFee, pk=pk)
    
    context = {
        'fee': fee,
        'title': f'Transport Fee: {fee.student.get_full_name()}',
        'page_title': f'Transport Fee Details'
    }
    return render(request, 'schools/transport/fee_detail.html', context)

# Meal Payment Management Views
@login_required
@admin_required
def meal_dashboard(request):
    """Meal payment management dashboard"""
    from django.db.models import Count, Sum, Q
    from django.utils import timezone
    from datetime import timedelta
    import json
    
    meal_payments = StudentMealPayment.objects.all()
    meal_pricing = MealPricing.objects.filter(is_active=True)
    
    # Location filter
    location_filter = request.GET.get('location', '')
    if location_filter:
        meal_payments = meal_payments.filter(location=location_filter)
        meal_pricing = meal_pricing.filter(location=location_filter)
    
    # Basic Statistics
    total_payments = meal_payments.count()
    completed_payments = meal_payments.filter(status='COMPLETED').count()
    pending_payments = meal_payments.filter(status='PENDING').count()
    failed_payments = meal_payments.filter(status='FAILED').count()
    
    # Revenue Statistics
    total_revenue = meal_payments.filter(status='COMPLETED').aggregate(total=Sum('amount'))['total'] or 0
    pending_revenue = meal_payments.filter(status='PENDING').aggregate(total=Sum('amount'))['total'] or 0
    failed_revenue = meal_payments.filter(status='FAILED').aggregate(total=Sum('amount'))['total'] or 0
    
    # Meal Type Distribution
    meal_type_data = {}
    for meal_type, meal_name in StudentMealPayment.MEAL_TYPES:
        count = meal_payments.filter(meal_type=meal_type, status='COMPLETED').count()
        revenue = meal_payments.filter(meal_type=meal_type, status='COMPLETED').aggregate(total=Sum('amount'))['total'] or 0
        meal_type_data[meal_name] = {
            'count': count,
            'revenue': float(revenue)
        }
    
    # Payment Method Distribution
    payment_method_data = {}
    for method, method_name in StudentMealPayment.PAYMENT_METHODS:
        count = meal_payments.filter(payment_method=method, status='COMPLETED').count()
        payment_method_data[method_name] = count
    
    # Monthly Revenue (last 12 months)
    monthly_revenue = []
    for i in range(11, -1, -1):
        month_start = timezone.now().replace(day=1) - timedelta(days=30*i)
        month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        revenue = meal_payments.filter(
            payment_date__gte=month_start,
            payment_date__lte=month_end,
            status='COMPLETED'
        ).aggregate(total=Sum('amount'))['total'] or 0
        monthly_revenue.append({
            'month': month_start.strftime('%b %Y'),
            'revenue': float(revenue)
        })
    
    # Daily Revenue (last 30 days)
    daily_revenue = []
    for i in range(29, -1, -1):
        day = timezone.now().date() - timedelta(days=i)
        revenue = meal_payments.filter(
            payment_date=day,
            status='COMPLETED'
        ).aggregate(total=Sum('amount'))['total'] or 0
        daily_revenue.append({
            'day': day.strftime('%b %d'),
            'revenue': float(revenue)
        })
    
    # Students with meal payments
    students_with_payments = meal_payments.filter(status='COMPLETED').values('student').distinct().count()
    
    # Today's revenue from meals served today
    today = timezone.now().date()
    today_consumptions = MealConsumption.objects.filter(consumption_date=today).select_related('meal_payment')
    today_revenue = Decimal('0.00')
    today_meals_served = today_consumptions.count()
    
    for consumption in today_consumptions:
        price_per_day = consumption.meal_payment.get_price_per_day()
        today_revenue += Decimal(str(price_per_day))
    
    # Today's payment revenue (payments made today)
    today_payment_revenue = meal_payments.filter(
        payment_date=today,
        status='COMPLETED'
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    
    context = {
        'total_payments': total_payments,
        'completed_payments': completed_payments,
        'pending_payments': pending_payments,
        'failed_payments': failed_payments,
        'total_revenue': total_revenue,
        'pending_revenue': pending_revenue,
        'failed_revenue': failed_revenue,
        'today_revenue': today_revenue,
        'today_payment_revenue': today_payment_revenue,
        'today_meals_served': today_meals_served,
        'meal_type_data': json.dumps(meal_type_data),
        'payment_method_data': json.dumps(payment_method_data),
        'monthly_revenue': json.dumps(monthly_revenue),
        'daily_revenue': json.dumps(daily_revenue),
        'students_with_payments': students_with_payments,
        'meal_pricing': meal_pricing,
        'title': 'Meal Payment Dashboard',
        'page_title': 'Food Management Dashboard'
    }
    return render(request, 'schools/meal/dashboard.html', context)

@login_required
@admin_required
def meal_pricing_list(request):
    """List and manage meal pricing"""
    pricing = MealPricing.objects.all().order_by('meal_type')
    
    if request.method == 'POST':
        # Handle bulk update or create
        for meal_type, _ in StudentMealPayment.MEAL_TYPES:
            price_key = f'price_{meal_type}'
            if price_key in request.POST:
                try:
                    price = Decimal(request.POST[price_key])
                    pricing_obj, created = MealPricing.objects.get_or_create(
                        meal_type=meal_type,
                        defaults={'price_per_day': price}
                    )
                    if not created:
                        pricing_obj.price_per_day = price
                        pricing_obj.is_active = True
                        pricing_obj.save()
                    messages.success(request, f'{pricing_obj.get_meal_type_display()} pricing updated')
                except (ValueError, InvalidOperation):
                    messages.error(request, f'Invalid price for {meal_type}')
        
        return redirect('meal_pricing_list')
    
    # Create a dictionary mapping meal_type to pricing object for easier template access
    pricing_dict = {p.meal_type: p for p in pricing}
    
    # Create list of meal types with their pricing info
    meal_types_with_pricing = []
    for meal_type, meal_name in StudentMealPayment.MEAL_TYPES:
        meal_types_with_pricing.append({
            'type': meal_type,
            'name': meal_name,
            'pricing': pricing_dict.get(meal_type)
        })
    
    context = {
        'pricing': pricing,
        'meal_types': StudentMealPayment.MEAL_TYPES,
        'meal_types_with_pricing': meal_types_with_pricing,
        'title': 'Meal Pricing',
        'page_title': 'Manage Meal Pricing'
    }
    return render(request, 'schools/meal/pricing_list.html', context)

@login_required
@admin_required
def meal_payment_list(request):
    """List all meal payments with filtering and analytics"""
    from django.db.models import Q, Sum, Count
    from django.utils import timezone
    from datetime import timedelta
    import json
    
    payments = StudentMealPayment.objects.select_related('student').all()
    
    # Filtering
    meal_filter = request.GET.get('meal_type', '')
    status_filter = request.GET.get('status', '')
    search_query = request.GET.get('search', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    if meal_filter:
        payments = payments.filter(meal_type=meal_filter)
    
    if status_filter:
        payments = payments.filter(status=status_filter)
    
    if search_query:
        payments = payments.filter(
            Q(student__first_name__icontains=search_query) |
            Q(student__last_name__icontains=search_query) |
            Q(student__admission_number__icontains=search_query) |
            Q(reference_number__icontains=search_query) |
            Q(transaction_id__icontains=search_query)
        )
    
    if date_from:
        payments = payments.filter(payment_date__gte=date_from)
    
    if date_to:
        payments = payments.filter(payment_date__lte=date_to)
    
    # Statistics
    total_payments = StudentMealPayment.objects.count()
    completed_payments = StudentMealPayment.objects.filter(status='COMPLETED').count()
    pending_payments = StudentMealPayment.objects.filter(status='PENDING').count()
    failed_payments = StudentMealPayment.objects.filter(status='FAILED').count()
    
    # Revenue Statistics
    total_revenue = StudentMealPayment.objects.filter(status='COMPLETED').aggregate(total=Sum('amount'))['total'] or 0
    pending_revenue = StudentMealPayment.objects.filter(status='PENDING').aggregate(total=Sum('amount'))['total'] or 0
    failed_revenue = StudentMealPayment.objects.filter(status='FAILED').aggregate(total=Sum('amount'))['total'] or 0
    
    # Payment Status Distribution
    payment_status_data = {
        'COMPLETED': completed_payments,
        'PENDING': pending_payments,
        'FAILED': failed_payments
    }
    
    # Payment Method Distribution
    payment_method_data = {}
    for method, method_name in StudentMealPayment.PAYMENT_METHODS:
        count = StudentMealPayment.objects.filter(payment_method=method, status='COMPLETED').count()
        payment_method_data[method_name] = count
    
    # Meal Type Distribution
    meal_type_data = {}
    for meal_type, meal_name in StudentMealPayment.MEAL_TYPES:
        count = StudentMealPayment.objects.filter(meal_type=meal_type, status='COMPLETED').count()
        revenue = StudentMealPayment.objects.filter(meal_type=meal_type, status='COMPLETED').aggregate(total=Sum('amount'))['total'] or 0
        meal_type_data[meal_name] = {
            'count': count,
            'revenue': float(revenue)
        }
    
    # Group payments by payment_group to show grouped payments as one entry
    grouped_payments = {}
    individual_payments = []
    
    for payment in payments.order_by('-payment_date', '-created_at'):
        if payment.payment_group:
            # Group payments with the same payment_group
            if payment.payment_group not in grouped_payments:
                # Extract base reference number (remove -2, -3 suffixes for display)
                base_ref = payment.reference_number
                if '-' in base_ref and base_ref.rsplit('-', 1)[-1].isdigit():
                    base_ref = base_ref.rsplit('-', 1)[0]
                
                grouped_payments[payment.payment_group] = {
                    'payments': [],
                    'total_amount': 0,
                    'payment_date': payment.payment_date,
                    'student': payment.student,
                    'reference_number': base_ref,  # Use base reference for display
                    'payment_method': payment.get_payment_method_display(),
                    'status': payment.status,
                    'transaction_id': payment.transaction_id,
                    'phone_number': payment.phone_number,
                    'created_at': payment.created_at
                }
            grouped_payments[payment.payment_group]['payments'].append(payment)
            grouped_payments[payment.payment_group]['total_amount'] += payment.amount
        else:
            # Individual payments (no group)
            individual_payments.append(payment)
    
    # Convert grouped payments to a list format for display
    grouped_payment_list = []
    for group_id, group_data in grouped_payments.items():
        grouped_payment_list.append({
            'is_grouped': True,
            'group_id': group_id,
            'payments': group_data['payments'],
            'total_amount': group_data['total_amount'],
            'payment_date': group_data['payment_date'],
            'student': group_data['student'],
            'reference_number': group_data['reference_number'],
            'payment_method': group_data['payment_method'],
            'status': group_data['status'],
            'transaction_id': group_data['transaction_id'],
            'phone_number': group_data['phone_number'],
            'created_at': group_data['created_at'],
            'meal_types': [p.get_meal_type_display() for p in group_data['payments']],
            'number_of_days': group_data['payments'][0].number_of_days if group_data['payments'] else 0
        })
    
    # Combine grouped and individual payments, sort by date
    all_payments = grouped_payment_list + [{'is_grouped': False, 'payment': p} for p in individual_payments]
    
    def get_sort_date(item):
        if item.get('is_grouped'):
            return item.get('payment_date') or item.get('created_at')
        else:
            payment = item.get('payment')
            if payment:
                return payment.payment_date or payment.created_at
        return None
    
    all_payments.sort(key=get_sort_date, reverse=True)
    
    # Pagination
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    paginator = Paginator(all_payments, 50)
    page = request.GET.get('page', 1)
    try:
        payments_page = paginator.page(page)
    except (EmptyPage, PageNotAnInteger):
        payments_page = paginator.page(1)
    
    context = {
        'payments': payments_page,
        'total_payments': total_payments,
        'completed_payments': completed_payments,
        'pending_payments': pending_payments,
        'failed_payments': failed_payments,
        'total_revenue': total_revenue,
        'pending_revenue': pending_revenue,
        'failed_revenue': failed_revenue,
        'payment_status_data': json.dumps(payment_status_data),
        'payment_method_data': json.dumps(payment_method_data),
        'meal_type_data': json.dumps(meal_type_data),
        'meal_filter': meal_filter,
        'status_filter': status_filter,
        'search_query': search_query,
        'date_from': date_from,
        'date_to': date_to,
        'title': 'Meal Payments',
        'page_title': 'Meal Payment Management'
    }
    return render(request, 'schools/meal/payment_list.html', context)

@login_required
@admin_required
def get_student_location(request, student_id):
    """API endpoint to get student location"""
    from django.http import JsonResponse
    try:
        student = Student.objects.get(pk=student_id)
        return JsonResponse({
            'location': student.location,
            'location_display': student.get_location_display()
        })
    except Student.DoesNotExist:
        return JsonResponse({'error': 'Student not found'}, status=404)

@login_required
@admin_required
def meal_payment_create(request):
    """Create a new meal payment"""
    if request.method == 'POST':
        form = StudentMealPaymentForm(request.POST)
        if form.is_valid():
            try:
                payment = form.save()
                
                # Automatically send receipt to parent
                try:
                    from .utils import send_meal_payment_receipt
                    success, result = send_meal_payment_receipt(payment)
                    if success:
                        if isinstance(result, dict):
                            if result.get('email_sent'):
                                messages.info(request, 'Receipt sent to parent email')
                            if result.get('whatsapp_sent'):
                                messages.info(request, 'Receipt notification sent to parent WhatsApp')
                    else:
                        # Log error but don't fail the payment recording
                        logger = logging.getLogger(__name__)
                        logger.warning(f"Receipt sending failed: {result}")
                except Exception as e:
                    # Log error but don't fail the payment recording
                    logger = logging.getLogger(__name__)
                    logger.error(f"Error sending receipt: {str(e)}")
                
                messages.success(request, f'Meal payment recorded successfully for {payment.student.get_full_name()}')
                return redirect('meal_payment_detail', pk=payment.pk)
            except Exception as e:
                import traceback
                error_details = traceback.format_exc()
                logger = logging.getLogger(__name__)
                logger.error(f"Error creating meal payment: {str(e)}\n{error_details}")
                messages.error(request, f'Error creating payment: {str(e)}')
        else:
            # Log form errors for debugging
            logger = logging.getLogger(__name__)
            logger.error(f"Form validation errors: {form.errors}")
            messages.error(request, 'Please correct the errors below.')
    else:
        form = StudentMealPaymentForm()
    
    # Get meal pricing for calculation
    meal_pricing = MealPricing.objects.filter(is_active=True)
    pricing_dict = {p.meal_type: float(p.price_per_day) for p in meal_pricing}
    import json
    
    context = {
        'form': form,
        'meal_pricing': pricing_dict,
        'meal_pricing_json': json.dumps(pricing_dict),
        'title': 'Record Meal Payment',
        'page_title': 'Record Meal Payment'
    }
    return render(request, 'schools/meal/payment_form.html', context)

@login_required
@admin_required
def meal_payment_detail(request, pk):
    """Meal payment detail view"""
    payment = get_object_or_404(StudentMealPayment.objects.select_related('student'), pk=pk)
    
    # Get all payments in the group (including current payment) for receipt generation
    related_payments = []
    all_group_payments = []
    if payment.payment_group:
        all_group_payments = StudentMealPayment.objects.filter(
            payment_group=payment.payment_group
        ).select_related('student').order_by('meal_type')
        related_payments = all_group_payments.exclude(pk=payment.pk)
    
    context = {
        'payment': payment,
        'related_payments': related_payments,
        'all_group_payments': all_group_payments,
        'is_grouped': bool(payment.payment_group),
        'title': f'Meal Payment: {payment.student.get_full_name()}',
        'page_title': f'Meal Payment Details'
    }
    return render(request, 'schools/meal/payment_detail.html', context)

@login_required
@admin_required
def meal_payment_edit(request, pk):
    """Edit meal payment"""
    payment = get_object_or_404(StudentMealPayment, pk=pk)
    
    if request.method == 'POST':
        form = StudentMealPaymentForm(request.POST, instance=payment)
        if form.is_valid():
            form.save()
            messages.success(request, 'Meal payment updated successfully')
            return redirect('meal_payment_detail', pk=payment.pk)
    else:
        form = StudentMealPaymentForm(instance=payment)
        # Set initial meals for the form
        form.initial['meals'] = [payment.meal_type]
    
    context = {
        'form': form,
        'payment': payment,
        'title': 'Edit Meal Payment',
        'page_title': 'Edit Meal Payment'
    }
    return render(request, 'schools/meal/payment_form.html', context)

@login_required
@admin_required
def meal_payment_report(request):
    """Meal serving and tracking page - redesigned with meal serving functionality"""
    from django.db.models import Q, Sum, Count, F
    from datetime import datetime, timedelta, date
    
    # Get filter parameters
    serving_date = request.GET.get('serving_date', timezone.now().date().strftime('%Y-%m-%d'))
    meal_type_filter = request.GET.get('meal_type', '')
    grade_id = request.GET.get('grade', '')
    search_query = request.GET.get('search', '')
    
    try:
        serving_date = datetime.strptime(serving_date, '%Y-%m-%d').date()
    except:
        serving_date = timezone.now().date()
    
    # Get all active payments first (for statistics)
    all_active_payments = StudentMealPayment.objects.filter(
        status='COMPLETED',
        is_active=True
    )
    
    # Get students with active meal payments (filtered)
    active_payments = StudentMealPayment.objects.select_related(
        'student', 'student__grade'
    ).filter(
        status='COMPLETED',
        is_active=True,
        days_remaining__gt=0
    )
    
    # Apply filters
    if meal_type_filter:
        active_payments = active_payments.filter(meal_type=meal_type_filter)
    if grade_id:
        active_payments = active_payments.filter(student__grade_id=grade_id)
    if search_query:
        active_payments = active_payments.filter(
            Q(student__first_name__icontains=search_query) |
            Q(student__last_name__icontains=search_query) |
            Q(student__admission_number__icontains=search_query)
        )
    
    # Group by student and meal type to show summary
    students_data = {}
    for payment in active_payments:
        # Initialize days_remaining and balance if not set
        if payment.days_remaining == 0 and payment.days_consumed == 0:
            # Fix invalid payment_frequency if needed
            valid_frequencies = [choice[0] for choice in StudentMealPayment.PAYMENT_FREQUENCY]
            if payment.payment_frequency not in valid_frequencies:
                payment.payment_frequency = 'MULTIPLE_DAYS'
            
            payment.days_remaining = payment.number_of_days
            payment.balance = payment.amount
            try:
                payment.save(update_fields=['days_remaining', 'balance', 'payment_frequency'])
            except ValidationError:
                # Skip this payment if validation fails
                continue
        
        student_id = payment.student.id
        if student_id not in students_data:
            students_data[student_id] = {
                'student': payment.student,
                'meals': {},
                'total_balance': Decimal('0.00')
            }
        
        meal_type = payment.meal_type
        if meal_type not in students_data[student_id]['meals']:
            students_data[student_id]['meals'][meal_type] = {
                'payment': payment,
                'days_remaining': 0,
                'balance': Decimal('0.00'),
                'already_served_today': False
            }
        
        # Sum up days and balance for same meal type
        students_data[student_id]['meals'][meal_type]['days_remaining'] += payment.days_remaining
        students_data[student_id]['meals'][meal_type]['balance'] += payment.balance
        students_data[student_id]['total_balance'] += payment.balance
        
        # Check if already served today
        already_served = MealConsumption.objects.filter(
            student=payment.student,
            meal_type=meal_type,
            consumption_date=serving_date
        ).exists()
        students_data[student_id]['meals'][meal_type]['already_served_today'] = already_served
    
    # Get today's consumption statistics
    today_consumptions = MealConsumption.objects.filter(consumption_date=serving_date).select_related('meal_payment')
    total_served_today = today_consumptions.count()
    
    # Calculate accurate revenue - sum of price per day for each consumption
    total_revenue_today = Decimal('0.00')
    for consumption in today_consumptions:
        price_per_day = consumption.meal_payment.get_price_per_day()
        total_revenue_today += Decimal(str(price_per_day))
    
    # Get consumption by meal type for today
    consumption_by_meal = today_consumptions.values('meal_type').annotate(
        count=Count('id')
    )
    
    # Additional statistics (from all active payments, not filtered)
    total_active_payments = all_active_payments.filter(
        days_remaining__gt=0
    ).count()
    
    # Initialize days_remaining and balance for payments that don't have them set
    # Also fix invalid payment_frequency values
    for payment in all_active_payments.filter(days_remaining=0, days_consumed=0)[:100]:  # Limit to avoid performance issues
        # Fix invalid payment_frequency if needed
        valid_frequencies = [choice[0] for choice in StudentMealPayment.PAYMENT_FREQUENCY]
        if payment.payment_frequency not in valid_frequencies:
            payment.payment_frequency = 'MULTIPLE_DAYS'  # Default to MULTIPLE_DAYS
        
        payment.days_remaining = payment.number_of_days
        payment.balance = payment.amount
        try:
            payment.save(update_fields=['days_remaining', 'balance', 'payment_frequency'])
        except ValidationError:
            # If validation still fails, skip this payment
            continue
    
    total_days_remaining = all_active_payments.aggregate(
        total=Sum('days_remaining')
    )['total'] or 0
    
    total_balance_remaining = all_active_payments.aggregate(
        total=Sum('balance')
    )['total'] or Decimal('0.00')
    
    # Total revenue from all completed payments
    total_revenue_all_time = StudentMealPayment.objects.filter(
        status='COMPLETED'
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    
    # Count unique students with active payments
    unique_students_count = all_active_payments.values('student').distinct().count()
    
    context = {
        'students_data': students_data,
        'serving_date': serving_date,
        'meal_type_filter': meal_type_filter,
        'grade_id': grade_id,
        'search_query': search_query,
        'total_served_today': total_served_today,
        'total_revenue_today': total_revenue_today,
        'total_active_payments': total_active_payments,
        'total_days_remaining': total_days_remaining,
        'total_balance_remaining': total_balance_remaining,
        'total_revenue_all_time': total_revenue_all_time,
        'unique_students_count': unique_students_count,
        'consumption_by_meal': consumption_by_meal,
        'grades': Grade.objects.all(),
        'meal_types': StudentMealPayment.MEAL_TYPES,
        'title': 'Meal Serving & Reports',
        'page_title': 'Meal Serving & Tracking'
    }
    return render(request, 'schools/meal/payment_report.html', context)

@login_required
@admin_required
@require_http_methods(["POST"])
def mark_meal_served(request):
    """Mark a meal as served for a student - AJAX endpoint"""
    from django.http import JsonResponse
    from datetime import datetime
    
    try:
        student_id = request.POST.get('student_id')
        meal_type = request.POST.get('meal_type')
        serving_date = request.POST.get('serving_date', timezone.now().date().strftime('%Y-%m-%d'))
        
        if not student_id or not meal_type:
            return JsonResponse({
                'success': False,
                'error': 'Student ID and meal type are required'
            }, status=400)
        
        student = get_object_or_404(Student, pk=student_id)
        
        # Parse date
        try:
            serving_date = datetime.strptime(serving_date, '%Y-%m-%d').date()
        except:
            serving_date = timezone.now().date()
        
        # Check if already served today
        existing = MealConsumption.objects.filter(
            student=student,
            meal_type=meal_type,
            consumption_date=serving_date
        ).first()
        
        if existing:
            return JsonResponse({
                'success': False,
                'error': f'{student.get_full_name()} has already been served {existing.get_meal_type_display()} today'
            }, status=400)
        
        # Find active payment for this meal type
        active_payment = StudentMealPayment.objects.filter(
            student=student,
            meal_type=meal_type,
            status='COMPLETED',
            is_active=True,
            days_remaining__gt=0
        ).order_by('payment_date').first()
        
        if not active_payment:
            return JsonResponse({
                'success': False,
                'error': f'No active payment found for {student.get_full_name()} - {dict(StudentMealPayment.MEAL_TYPES).get(meal_type, meal_type)}'
            }, status=404)
        
        # Create consumption record
        consumption = MealConsumption.objects.create(
            student=student,
            meal_payment=active_payment,
            meal_type=meal_type,
            consumption_date=serving_date,
            served_by=request.user
        )
        
        # Consume one day from the payment
        active_payment.consume_day()
        
        return JsonResponse({
            'success': True,
            'message': f'{student.get_full_name()} marked as served for {consumption.get_meal_type_display()}',
            'days_remaining': active_payment.days_remaining,
            'balance': float(active_payment.balance),
            'consumption_id': consumption.id
        })
        
    except Student.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Student not found'
        }, status=404)
    except Exception as e:
        import traceback
        return JsonResponse({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc() if request.user.is_superuser else None
        }, status=500)


@login_required
def sync_status_view(request):
    """View to display sync status page"""
    from schools.utils.sync import sync_service
    
    stats = sync_service.get_sync_statistics()
    pending_items = SyncQueue.objects.filter(status='PENDING').order_by('-created_at')[:50]
    failed_items = SyncQueue.objects.filter(status='FAILED').order_by('-created_at')[:50]
    
    context = {
        'stats': stats,
        'pending_items': pending_items,
        'failed_items': failed_items,
    }
    
    return render(request, 'schools/sync_status.html', context)


@login_required
def sync_status_api(request):
    """API endpoint to get current sync status (JSON)"""
    from schools.utils.sync import sync_service
    
    stats = sync_service.get_sync_statistics()
    
    return JsonResponse({
        'is_online': stats['is_online'],
        'pending': stats['pending'],
        'failed': stats['failed'],
        'last_sync': stats['last_sync'].isoformat() if stats['last_sync'] else None,
    })


@login_required
def manual_sync(request):
    """Manually trigger a sync operation"""
    from schools.utils.sync import sync_service
    
    if request.method == 'POST':
        max_items = request.POST.get('max_items', None)
        if max_items:
            try:
                max_items = int(max_items)
            except ValueError:
                max_items = None
        
        results = sync_service.sync_pending_operations(max_items=max_items)
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse(results)
        else:
            if results['synced'] > 0:
                messages.success(request, f'Successfully synced {results["synced"]} item(s)')
            if results['failed'] > 0:
                messages.warning(request, f'Failed to sync {results["failed"]} item(s)')
            if results['synced'] == 0 and results['failed'] == 0:
                messages.info(request, 'No pending operations to sync')
            
            return redirect('sync_status')
    
    return redirect('sync_status')

@csrf_exempt
@require_http_methods(["GET"])
@login_required
def api_departments(request):
    """API endpoint to list departments"""
    try:
        from config.models import SchoolConfig
        school = SchoolConfig.get_config(user=request.user, request=request)
        
        departments_qs = Department.objects.filter(school=school)
        
        # Seed default departments if none exist
        if not departments_qs.exists():
            default_depts = [
                'Languages', 'Sciences', 'Mathematics', 'Humanities', 
                'Technical & Applied', 'Physical Education', 'Special Education'
            ]
            for dept_name in default_depts:
                Department.objects.create(school=school, name=dept_name)
            # Re-fetch
            departments_qs = Department.objects.filter(school=school)
            
        departments = departments_qs.values('id', 'name').order_by('name')
        return JsonResponse({'departments': list(departments)})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(['POST'])
@login_required
def api_attendance_mark_batch(request):
    try:
        import json
        data = json.loads(request.body)
        
        attendance_list = data.get('attendance', [])
        date_str = data.get('date')
        
        if not attendance_list or not date_str:
            return JsonResponse({'error': 'Attendance data and date are required'}, status=400)
            
        saved_count = 0
        
        with transaction.atomic():
            for item in attendance_list:
                student_id = item.get('student_id')
                status = item.get('status')
                remarks = item.get('remarks', '')
                
                if not student_id or not status:
                    continue
                    
                student = Student.objects.get(pk=student_id)
                
                Attendance.objects.update_or_create(
                    student=student,
                    date=date_str,
                    defaults={
                        'status': status,
                        'remarks': remarks,
                        'recorded_by': request.user
                    }
                )
                saved_count += 1
                
        return JsonResponse({'success': True, 'saved_count': saved_count})
        
    except Student.DoesNotExist:
        return JsonResponse({'error': 'One or more students not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def api_grade_attendance(request, grade_id):
    date_str = request.GET.get('date')
    if not date_str:
        return JsonResponse({'error': 'Date is required'}, status=400)
    
    try:
        attendance_records = Attendance.objects.filter(
            student__grade_id=grade_id,
            date=date_str
        ).select_related('student')
        
        data = {}
        for record in attendance_records:
            data[record.student.id] = {
                'status': record.status,
                'remarks': record.remarks or ''
            }
            
        return JsonResponse({'attendance': data})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def api_student_attendance(request, student_id):
    try:
        student = Student.objects.get(pk=student_id)
        attendance_records = Attendance.objects.filter(student=student).order_by('-date')
        
        data = []
        for record in attendance_records:
            # Use the actual time it was marked
            time_in = timezone.localtime(record.updated_at).strftime('%I:%M %p') if record.updated_at else "08:00 AM" 
                 
            data.append({
                'date': record.date.strftime('%Y-%m-%d'),
                'status': record.status,
                'remarks': record.remarks,
                'time': time_in,
                'time_in': time_in,
                'day': record.date.strftime('%a')
            })
            
        return JsonResponse({'attendance': data})
    except Student.DoesNotExist:
        return JsonResponse({'error': 'Student not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(['POST'])
@login_required
def api_student_attendance_mark_single(request, student_id):
    try:
        data = json.loads(request.body)
        date_str = data.get('date')
        status = data.get('status')
        remarks = data.get('remark', '') or data.get('remarks', '')
        
        if not date_str or not status:
            return JsonResponse({'error': 'Date and status are required'}, status=400)
            
        student = Student.objects.get(pk=student_id)
        
        # Use student's current term if not specified
        term = str(student.current_term)
        
        with transaction.atomic():
            attendance, created = Attendance.objects.update_or_create(
                student=student,
                date=date_str,
                defaults={
                    'status': status.upper(),
                    'remarks': remarks,
                    'recorded_by': request.user,
                    'term': term
                }
            )
        
        return JsonResponse({
            'success': True, 
            'message': 'Attendance recorded successfully',
            'created': created
        })
        
    except Student.DoesNotExist:
        return JsonResponse({'error': 'Student not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

# reload
# force reload 2

@never_cache
@csrf_exempt
@require_http_methods(["GET"])
def api_school_config(request):
    """API endpoint to fetch school configuration"""
    try:
        from config.models import SchoolConfig
        # Pass request to get_config so it can read X-Portal-Slug header
        config = SchoolConfig.get_config(
            user=request.user if request.user.is_authenticated else None,
            request=request
        )
        
        data = {
            'school_name': config.school_name,
            'school_code': config.school_code,
            'school_email': config.school_email,
            'school_phone': config.school_phone,
            'school_address': config.school_address,
            'school_logo': request.build_absolute_uri(config.school_logo.url) if config.school_logo else None,
            'portal_slug': config.portal_slug,
            'current_term': config.current_term,
            'current_year': config.current_year,
            'admission_number_format': config.admission_number_format,
            'admission_counter': config.admission_counter,
            'currency': config.default_currency,
            'admission_fee': float(config.admission_fee) if hasattr(config, 'admission_fee') else 0,
            'subscription': {
                'plan': config.subscription_plan,
                'status': config.subscription_status,
                'trial_start': config.trial_start_date.isoformat() if config.trial_start_date else None,
                'trial_end': config.trial_end_date.isoformat() if config.trial_end_date else None,
                'is_test_mode': settings.PAYSTACK_SECRET_KEY.startswith('sk_test'),
            }
        }
        
        if request.user.is_authenticated and request.user.is_superuser:
            # Unlock the portal for superusers regardless of actual status
            # data['subscription']['status'] = 'Active'
            # data['subscription']['plan'] = 'Enterprise'
            
            data.update({
                'teacher_portal_password': config.teacher_portal_password,
                'student_portal_password': config.student_portal_password,
                'accountant_portal_password': config.accountant_portal_password,
                'food_portal_password': config.food_portal_password,
                'transport_portal_password': config.transport_portal_password,
                'driver_portal_password': config.driver_portal_password,
            })
            
        return JsonResponse(data)  # Return data directly, not nested
    except Exception as e:
        import traceback
        logging.error(f"Config API Error: {str(e)}")
        logging.error(traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["POST"])
@login_required
def api_school_config_update(request):
    """API endpoint to update school configuration"""
    try:
        from config.models import SchoolConfig
        config = SchoolConfig.get_config(request.user, request=request)
        
        # Check permissions (superuser or school admin)
        if not (request.user.is_superuser or config.admin == request.user):
             return JsonResponse({'error': 'Permission denied. Only school administrators can update settings.'}, status=403)
        
        # Handle multipart/form-data
        if request.content_type.startswith('multipart/form-data'):
             data = request.POST
             files = request.FILES
        else:
             import json
             data = json.loads(request.body)
             files = {}
             
        if 'school_name' in data: 
            config.school_name = data['school_name']
            # Sync with User first_name if user is the admin
            if config.admin == request.user:
                request.user.first_name = data['school_name']
                request.user.save()

        if 'school_code' in data: config.school_code = data['school_code']
        
        if 'school_email' in data: 
            config.school_email = data['school_email']
            # Sync with User email/username if user is the admin and they were the same
            if config.admin == request.user:
                email = data['school_email']
                if not User.objects.filter(email=email).exclude(pk=request.user.pk).exists():
                    old_email = request.user.email
                    request.user.email = email
                    if request.user.username == old_email:
                        request.user.username = email
                    request.user.save()

        if 'school_phone' in data: config.school_phone = data['school_phone']
        if 'school_address' in data: config.school_address = data['school_address']
        
        if 'school_logo' in files:
            config.school_logo = files['school_logo']
            
        if 'current_term' in data: config.current_term = data['current_term']
        if 'current_year' in data: config.current_year = data['current_year']
        if 'admission_number_format' in data: config.admission_number_format = data['admission_number_format']
        if 'admission_fee' in data: config.admission_fee = data['admission_fee']
        
        # Passwords
        if 'teacher_portal_password' in data: config.teacher_portal_password = data['teacher_portal_password']
        if 'student_portal_password' in data: config.student_portal_password = data['student_portal_password']
        if 'accountant_portal_password' in data: config.accountant_portal_password = data['accountant_portal_password']
        if 'food_portal_password' in data: config.food_portal_password = data['food_portal_password']
        if 'transport_portal_password' in data: config.transport_portal_password = data['transport_portal_password']
        if 'driver_portal_password' in data: config.driver_portal_password = data['driver_portal_password']
        
        config.save()
        
        # Return the updated config so frontend can display it immediately
        updated_data = {
            'school_name': config.school_name,
            'school_code': config.school_code,
            'school_email': config.school_email,
            'school_phone': config.school_phone,
            'school_address': config.school_address,
            'school_logo': request.build_absolute_uri(config.school_logo.url) if config.school_logo else None,
            'portal_slug': config.portal_slug,
            'current_term': config.current_term,
            'current_year': config.current_year,
            'admission_number_format': config.admission_number_format,
            'admission_counter': config.admission_counter,
            'currency': config.default_currency,
            'admission_fee': float(config.admission_fee) if hasattr(config, 'admission_fee') else 0,
        }
        
        return JsonResponse({
            'success': True, 
            'message': 'Configuration updated successfully', 
            'config': updated_data,
            'user': {
                'id': request.user.id,
                'username': request.user.username,
                'email': request.user.email,
                'first_name': request.user.first_name,
                'last_name': request.user.last_name
            }
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["POST"])
@login_required
def api_regenerate_portal_slug(request):
    """API endpoint to regenerate portal slug"""
    try:
        from config.models import SchoolConfig
        config = SchoolConfig.get_config(request.user, request=request)
        
        if not (request.user.is_superuser or config.admin == request.user):
             return JsonResponse({'error': 'Permission denied'}, status=403)
             
        import uuid
        config.portal_slug = str(uuid.uuid4())[:12]
        config.save()
        return JsonResponse({'success': True, 'portal_slug': config.portal_slug})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
@csrf_exempt
@require_http_methods(["POST"])
@login_required
def api_user_profile_update(request):
    """API endpoint to update user profile details"""
    try:
        data = json.loads(request.body)
        user = request.user
        
        user.first_name = data.get('first_name', user.first_name)
        user.last_name = data.get('last_name', user.last_name)
        
        email = data.get('email')
        if email and email != user.email:
            if User.objects.filter(email=email).exclude(pk=user.pk).exists():
                return JsonResponse({'error': 'Email already in use'}, status=400)
            
            old_email = user.email
            user.email = email
            # If username was email, update it
            if user.username == old_email:
                user.username = email
        
        user.save()
        return JsonResponse({
            'success': True, 
            'message': 'Profile updated successfully',
            'user': {
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'first_name': user.first_name,
                'last_name': user.last_name
            }
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["POST"])
@login_required
def api_change_password(request):
    """API endpoint to change user password"""
    try:
        data = json.loads(request.body)
        user = request.user
        
        current_password = data.get('current_password')
        new_password = data.get('new_password')
        
        if not user.check_password(current_password):
            return JsonResponse({'error': 'Incorrect current password'}, status=400)
            
        user.set_password(new_password)
        user.save()
        
        # Update session to keep user logged in
        from django.contrib.auth import update_session_auth_hash
        update_session_auth_hash(request, user)
        
        return JsonResponse({'success': True, 'message': 'Password changed successfully'})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
@csrf_exempt
@require_http_methods(["POST"])
def verify_subscription_payment(request):
    """
    Manually verify a subscription payment with Paystack
    Called by frontend to check if payment was successful
    """
    try:
        data = json.loads(request.body)
        reference = data.get('reference')
        
        if not reference:
            return JsonResponse({'error': 'Reference is required'}, status=400)
        
        # Call Paystack verification API
        import requests
        headers = {
            'Authorization': f'Bearer {settings.PAYSTACK_SECRET_KEY}',
            'Content-Type': 'application/json'
        }
        
        verify_url = f'https://api.paystack.co/transaction/verify/{reference}'
        response = requests.get(verify_url, headers=headers)
        
        if response.status_code == 200:
            result = response.json()
            if result.get('status') and result['data'].get('status') == 'success':
                # Payment was successful, process the upgrade
                if reference.startswith('SUB_UP_'):
                    try:
                        # Extract plan and school ID: SUB_UP_{plan}_{school_id}_{timestamp}
                        parts = reference.split('_')
                        if len(parts) >= 4:
                            target_plan = parts[2]
                            school_id = parts[3]
                            from config.models import SchoolConfig
                            from django.utils import timezone
                            from datetime import timedelta
                            
                            config = SchoolConfig.objects.get(id=school_id)
                            
                            # Update fields
                            config.subscription_status = 'Active'
                            config.subscription_plan = target_plan
                            
                            # If they have remaining trial days, add to them, otherwise start fresh month
                            if config.trial_end_date and config.trial_end_date > timezone.now():
                                config.trial_end_date = config.trial_end_date + timedelta(days=31)
                            else:
                                config.trial_end_date = timezone.now() + timedelta(days=31)
                                
                            config.save()

                            # Log the payment for Super Admin tracking
                            from config.models import SubscriptionPayment
                            from decimal import Decimal
                            tx_data = result.get('data', {})
                            amount_paid = Decimal(str(tx_data.get('amount', 0))) / 100
                            
                            SubscriptionPayment.objects.update_or_create(
                                reference=reference,
                                defaults={
                                    'school': config,
                                    'amount': amount_paid,
                                    'plan': target_plan,
                                    'status': 'COMPLETED',
                                    'transaction_id': tx_data.get('id')
                                }
                            )
                            
                            print(f"DEBUG: School '{config.school_name}' upgraded to {target_plan} via Manual Verification. Amount: {amount_paid}")
                            
                            return JsonResponse({
                                'success': True,
                                'upgraded': True,
                                'plan': target_plan,
                                'status': 'Active',
                                'message': 'Subscription upgraded successfully'
                            })
                    except SchoolConfig.DoesNotExist:
                        print(f"ERROR: School ID {school_id} not found")
                        return JsonResponse({'error': 'School not found'}, status=404)
                    except Exception as e:
                        print(f"ERROR Upgrading School: {str(e)}")
                        return JsonResponse({'error': str(e)}, status=500)
                
                print(f"DEBUG: Reference {reference} verified but not a SUB_UP reference")
                return JsonResponse({'success': True, 'upgraded': False})
            else:
                print(f"DEBUG: Payment verification failed or status not success: {result.get('data', {}).get('status')}")
                return JsonResponse({'success': False, 'message': 'Payment not confirmed yet'})
        else:
            print(f"DEBUG: Paystack returned error code: {response.status_code}")
            return JsonResponse({'error': 'Failed to verify payment'}, status=500)
            
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)
